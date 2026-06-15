from __future__ import annotations

from dataclasses import asdict, is_dataclass
from pathlib import Path
from typing import Any, Iterable, Mapping
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RAIZ = Path(__file__).resolve().parents[1]
if str(RAIZ) not in sys.path:
    sys.path.insert(0, str(RAIZ))

from nucleo.contexto_operacional_canonico import carregar_contexto_operacional_canonico
from nucleo.identidade_baseline import VERSAO_BASELINE, VERSAO_SLUG, caminho_artifact, caminho_saida_operacional, nome_relatorio_operacional


class PacoteSaidaObservavelOficialAusente(RuntimeError):
    pass


ABAS = {
    'extrato_passado': 'Extrato Passado',
    'extrato_futuro': 'Extrato Futuro',
    'switching': 'Switching',
    'carteira': 'Carteira',
    'situacao_atual': 'Situação Atual',
}

HEADERS = {
    'extrato_passado': ['Data', 'Conta', 'Despesa ID', 'Lote', 'Saldo Antes', 'Bruto', 'Imposto', 'Líquido', 'Saldo Remanescente'],
    'extrato_futuro': ['Data', 'Conta', 'Despesa ID', 'Valor', 'Lote sugerido', 'Fonte técnica', 'Saldo Antes', 'Bruto', 'Imposto', 'Líquido', 'Saldo Remanescente', 'Cobertura integral', 'Pacote do dia', 'Pacote técnico', 'Motivo bloqueio lote', 'Status recomendação'],
    'switching': ['Data sugerida', 'Lote origem', 'Produto origem', 'Produto destino switching', 'Ganho estimado', 'Valor líquido origem', 'Status'],
}


def _cfg(contexto: Any) -> Mapping[str, Any]:
    conteudo = getattr(getattr(contexto, 'pacote_config', None), 'conteudo', {}) or {}
    saidas = conteudo.get('saidas') if isinstance(conteudo, Mapping) else {}
    cfg = saidas.get('planilha_operacional') if isinstance(saidas, Mapping) else {}
    return cfg if isinstance(cfg, Mapping) else {}


def _nome_aba(contexto: Any, chave: str) -> str:
    abas = _cfg(contexto).get('abas')
    if isinstance(abas, Mapping) and str(abas.get(chave) or '').strip():
        return str(abas[chave])
    return ABAS[chave]


def _nome_arquivo(contexto: Any) -> str:
    cfg = _cfg(contexto)
    nome = str(cfg.get('arquivo') or cfg.get('nome_arquivo') or nome_relatorio_operacional()).strip()
    try:
        nome = nome.format(versao=VERSAO_BASELINE, versao_slug=VERSAO_SLUG)
    except Exception:
        pass
    return nome if nome.lower().endswith('.xlsx') else f'{nome}.xlsx'


def _caminhos(contexto: Any) -> tuple[Path, Path]:
    nome = _nome_arquivo(contexto)
    return caminho_saida_operacional(RAIZ, nome), caminho_artifact(nome)


def _serializar(valor: Any) -> Any:
    if is_dataclass(valor):
        return asdict(valor)
    if isinstance(valor, (dict, list, tuple, set)):
        return str(valor)
    if hasattr(valor, 'isoformat') and not isinstance(valor, str):
        try:
            return valor.isoformat()
        except Exception:
            return str(valor)
    return valor


def _linha(item: Any) -> dict[str, Any]:
    if is_dataclass(item):
        item = asdict(item)
    if not isinstance(item, Mapping):
        return {'valor': item}
    return {str(k): _serializar(v) for k, v in dict(item).items()}


def _linhas(itens: Iterable[Any]) -> list[dict[str, Any]]:
    return [_linha(item) for item in list(itens or [])]


def _headers(linhas: list[dict[str, Any]]) -> list[str]:
    vistos: set[str] = set()
    cols: list[str] = []
    for linha in linhas:
        for chave in linha:
            if chave not in vistos:
                cols.append(chave)
                vistos.add(chave)
    return cols or ['status']


def _abas_oficiais(pacote: Any) -> dict[str, list[dict[str, Any]]]:
    bloco = getattr(pacote, 'bloco_xlsx', None)
    abas = getattr(bloco, 'abas', None)
    if not isinstance(abas, Mapping):
        return {}
    return {str(nome): _linhas(linhas) for nome, linhas in dict(abas).items()}


def _situacao_blocos(pacote: Any) -> list[dict[str, Any]]:
    bloco = getattr(pacote, 'bloco_xlsx', None)
    blocos = []
    for item in list(getattr(bloco, 'situacao_atual_blocos', []) or []):
        if not isinstance(item, Mapping):
            continue
        titulo = str(item.get('titulo') or '').strip()
        headers = list(item.get('headers') or [])
        linhas = _linhas(item.get('linhas') or [])
        if titulo and headers:
            blocos.append({'titulo': titulo, 'headers': headers, 'linhas': linhas})
    return blocos


def _validar_pacote(pacote: Any) -> None:
    if pacote is None:
        raise PacoteSaidaObservavelOficialAusente('Geracao XLSX oficial exige PacoteSaidaObservavelOficial.')
    if not getattr(pacote, 'preparado', False):
        raise PacoteSaidaObservavelOficialAusente('PacoteSaidaObservavelOficial nao preparado.')
    if getattr(pacote, 'bloco_xlsx', None) is None:
        raise PacoteSaidaObservavelOficialAusente('PacoteSaidaObservavelOficial sem bloco_xlsx.')
    abas = _abas_oficiais(pacote)
    faltantes = [nome for nome in ['Extrato Passado', 'Extrato Futuro', 'Switching', 'Carteira'] if not abas.get(nome)]
    if faltantes:
        raise PacoteSaidaObservavelOficialAusente(f'Abas oficiais ausentes ou vazias: {faltantes}.')
    if not _situacao_blocos(pacote) and not abas.get('Situação Atual'):
        raise PacoteSaidaObservavelOficialAusente('Situacao Atual oficial ausente.')


def _estilo(ws, headers: list[str], linhas: list[dict[str, Any]], *, start_row: int = 1, title: str | None = None, congelar_painel: bool = True) -> int:
    ws.sheet_view.showGridLines = False
    fill = PatternFill('solid', fgColor='D9EAF7')
    font = Font(color='1F1F1F', bold=True)
    title_fill = PatternFill('solid', fgColor='EDF4FA')
    title_font = Font(color='1F1F1F', bold=True, size=12)
    thin = Side(style='thin', color='D9E1F2')
    header_row = start_row + 1 if title else start_row
    if title:
        ws.cell(start_row, 1, title).fill = title_fill
        ws.cell(start_row, 1).font = title_font
    for col, header in enumerate(headers, 1):
        cell = ws.cell(header_row, col, header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(bottom=thin)
    for lin, item in enumerate(linhas, header_row + 1):
        for col, header in enumerate(headers, 1):
            valor = item.get(header)
            cell = ws.cell(lin, col, valor)
            cell.alignment = Alignment(horizontal='right' if isinstance(valor, (int, float)) else 'left', vertical='center')
            cell.border = Border(bottom=thin)
    if headers:
        ws.auto_filter.ref = f'A{header_row}:{get_column_letter(len(headers))}{max(header_row + len(linhas), header_row)}'
    if congelar_painel:
        ws.freeze_panes = f'A{header_row + 1}'
    for col, header in enumerate(headers, 1):
        letter = get_column_letter(col)
        largura = max([len(str(header))] + [len(str(linha.get(header))) for linha in linhas if linha.get(header) is not None])
        ws.column_dimensions[letter].width = min(max(largura + 2, 10), 42)
    return header_row + len(linhas)


def _aba(wb: Workbook, nome: str, headers: list[str], linhas: list[dict[str, Any]]) -> None:
    ws = wb.active if len(wb.worksheets) == 1 and wb.active.title == 'Sheet' else wb.create_sheet(nome)
    ws.title = nome
    _estilo(ws, headers, linhas)


def _aba_situacao(wb: Workbook, contexto: Any, pacote: Any) -> None:
    ws = wb.create_sheet(_nome_aba(contexto, 'situacao_atual'))
    blocos = _situacao_blocos(pacote)
    if not blocos:
        linhas = _abas_oficiais(pacote).get('Situação Atual', [])
        _estilo(ws, _headers(linhas), linhas, congelar_painel=False)
        return
    row = 1
    for idx, bloco in enumerate(blocos):
        row = _estilo(ws, bloco['headers'], bloco['linhas'], start_row=row if idx == 0 else row + 3, title=bloco['titulo'], congelar_painel=False)


def main(*, contexto: Any = None, saida: Any = None, pacote_saida_observavel_oficial: Any = None, estado_temporal_inicial: Any = None, incluir_abas_diagnosticas: bool | None = None, modo_artefato: str = 'oficial') -> Path:
    _ = saida, estado_temporal_inicial, incluir_abas_diagnosticas, modo_artefato
    _validar_pacote(pacote_saida_observavel_oficial)
    if contexto is None:
        contexto = carregar_contexto_operacional_canonico(raiz_repositorio=RAIZ, instalar_automaticamente=False)
    saida_interna, saida_externa = _caminhos(contexto)
    abas = _abas_oficiais(pacote_saida_observavel_oficial)
    wb = Workbook()
    _aba(wb, _nome_aba(contexto, 'extrato_passado'), HEADERS['extrato_passado'], abas['Extrato Passado'])
    _aba(wb, _nome_aba(contexto, 'extrato_futuro'), HEADERS['extrato_futuro'], abas['Extrato Futuro'])
    _aba(wb, _nome_aba(contexto, 'switching'), HEADERS['switching'], abas['Switching'])
    carteira = abas['Carteira']
    _aba(wb, _nome_aba(contexto, 'carteira'), _headers(carteira), carteira)
    _aba_situacao(wb, contexto, pacote_saida_observavel_oficial)
    saida_interna.parent.mkdir(parents=True, exist_ok=True)
    wb.save(saida_interna)
    try:
        if saida_externa.parent.exists():
            wb.save(saida_externa)
    except Exception as exc:
        print(f'[AVISO] copia externa nao gerada: {type(exc).__name__}:{exc}')
    return saida_interna


if __name__ == '__main__':
    print(main())
