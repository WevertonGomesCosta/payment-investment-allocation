from __future__ import annotations

from dataclasses import asdict, dataclass, field, is_dataclass
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import re
from pathlib import Path
from typing import Any, Iterable, Mapping

from nucleo.saida_observavel_oficial import PacoteSaidaObservavelOficial


ARTEFATO_PARIDADE = 'ResultadoParidadeRenderizacaoOficial'
ENTRADA_FORMAL = 'PacoteSaidaObservavelOficial'
MODULO_PARIDADE = 'nucleo/paridade_renderizacao_oficial.py'
ETAPA_PARIDADE = 10
PREFIXO_ABA_OBSERVAVEL = 'Obs '

ABAS_XLSX_OFICIAIS_CONTRATUAIS: dict[str, list[str]] = {
    'Extrato Passado': [
        'Data', 'Conta', 'Despesa ID', 'Lote', 'Saldo Antes',
        'Bruto', 'Imposto', 'Líquido', 'Saldo Remanescente',
    ],
    'Extrato Futuro': [
        'Data', 'Conta', 'Despesa ID', 'Valor', 'Lote sugerido',
        'Fonte técnica', 'Saldo Antes', 'Bruto', 'Imposto', 'Líquido',
        'Saldo Remanescente', 'Cobertura integral', 'Pacote do dia',
        'Pacote técnico', 'Motivo bloqueio lote', 'Status recomendação',
    ],
    'Switching': [
        'Data sugerida', 'Lote origem', 'Produto origem',
        'Produto destino switching', 'Ganho estimado',
        'Valor líquido origem', 'Status',
    ],
    'Carteira': [
        'Rank', 'Produto', 'Score Final', 'Proxy Terminal',
        'Retorno Proxy aa', 'Liquidez Dias', 'Carência Dias',
        'Aplicação Mínima', 'Aplicação Máxima', 'Tipo Produto',
        'Somente Combo', 'Status Confirmação', 'Campos Pendentes',
    ],
    'Situação Atual': [],
}

BLOCOS_SITUACAO_ATUAL_OBRIGATORIOS: tuple[str, ...] = (
    'Lotes exauridos — identificação',
    'Lotes exauridos — valores e patrimônio',
    'Lotes ativos — identificação',
    'Lotes ativos — valores e patrimônio',
    'Origens migradas por switching — reconciliação patrimonial',
    'Patrimônio total dos lotes',
    'Recebidos auditáveis',
    'Fechamento econômico',
    'Resumo de recebidos',
)

SECOES_CONSOLE_PARIDADE_FORTE: tuple[str, ...] = (
    'resumo_operacional',
    'ultimos_pagamentos',
    'pagamentos_data_referencia',
    'proximos_pagamentos',
    'pagamentos_por_fonte',
    'ranking_metricas',
    'ranking_amostra',
    'switchings_metricas',
    'switchings_amostra',
    'switchings_resumo_operacional',
    'switchings_realizados_operacionais',
    'situacao_atual_fechamento',
    'situacao_atual_lotes_exauridos_id',
    'situacao_atual_lotes_exauridos_valores',
    'situacao_atual_lotes_ativos_id',
    'situacao_atual_lotes_ativos_valores',
    'situacao_atual_origens_migradas',
    'situacao_atual_patrimonio_total',
    'situacao_atual_recebidos_auditaveis',
    'situacao_atual_resumo_recebidos',
    'situacao_atual_blocos',
)

ROTULOS_SECOES_CONSOLE_PARIDADE_FORTE: dict[str, str] = {
    'resumo_operacional': 'Saída Observável Oficial / resumo operacional',
    'ultimos_pagamentos': 'Pagamentos — últimos pagamentos realizados',
    'pagamentos_data_referencia': 'Pagamentos — data de referência',
    'proximos_pagamentos': 'Pagamentos — próximos pagamentos',
    'pagamentos_por_fonte': 'Pagamentos — valores por fonte',
    'ranking_metricas': 'Ranking/Carteira — métricas',
    'ranking_amostra': 'Ranking/Carteira — amostra',
    'switchings_metricas': 'Switching — métricas',
    'switchings_amostra': 'Switching — amostra',
    'switchings_resumo_operacional': 'Switching — resumo operacional',
    'switchings_realizados_operacionais': 'Switching — realizados operacionais',
    'situacao_atual_fechamento': 'Situação Atual — fechamento',
    'situacao_atual_lotes_exauridos_id': 'Situação Atual — lotes exauridos identificação',
    'situacao_atual_lotes_exauridos_valores': 'Situação Atual — lotes exauridos valores',
    'situacao_atual_lotes_ativos_id': 'Situação Atual — lotes ativos identificação',
    'situacao_atual_lotes_ativos_valores': 'Situação Atual — lotes ativos valores',
    'situacao_atual_origens_migradas': 'Situação Atual — origens migradas',
    'situacao_atual_patrimonio_total': 'Situação Atual — patrimônio total',
    'situacao_atual_recebidos_auditaveis': 'Situação Atual — recebidos auditáveis',
    'situacao_atual_resumo_recebidos': 'Situação Atual — resumo de recebidos',
    'situacao_atual_blocos': 'Situação Atual — blocos observáveis',
}
CATEGORIAS_DIVERGENCIA = (
    'PARIDADE_OK',
    'ARTEFATO_RENDERIZADO_AUSENTE',
    'ABA_XLSX_AUSENTE',
    'ABA_XLSX_EXTRA',
    'DIVERGENCIA_ESTRUTURAL',
    'DIVERGENCIA_HEADERS',
    'DIVERGENCIA_QTD_LINHAS',
    'DIVERGENCIA_CONTEUDO',
    'DIVERGENCIA_SERIALIZACAO',
    'DIVERGENCIA_NORMALIZACAO_NUMERICA',
    'DIVERGENCIA_DATA_DATETIME',
    'DIVERGENCIA_MATERIAL',
    'CONSOLE_NAO_AUDITADO',
    'CONSOLE_AUDITADO_COM_RESSALVA',
    'MELHORIA_ERGONOMICA',
)
_CATEGORIAS_MATERIAIS = {
    'ARTEFATO_RENDERIZADO_AUSENTE',
    'ABA_XLSX_AUSENTE',
    'ABA_XLSX_EXTRA',
    'DIVERGENCIA_ESTRUTURAL',
    'DIVERGENCIA_HEADERS',
    'DIVERGENCIA_QTD_LINHAS',
    'DIVERGENCIA_CONTEUDO',
    'DIVERGENCIA_MATERIAL',
}


@dataclass(slots=True)
class DivergenciaParidadeRenderizacao:
    categoria: str
    alvo: str
    mensagem: str
    material: bool = False
    aba: str | None = None
    linha: int | None = None
    coluna: str | None = None
    esperado: Any = None
    observado: Any = None
    referencias: dict[str, Any] = field(default_factory=dict)


@dataclass(slots=True)
class ResumoParidadeRenderizacaoOficial:
    status: str
    ok: bool
    qtd_divergencias: int
    qtd_divergencias_materiais: int
    qtd_ressalvas: int
    qtd_abas_esperadas: int
    qtd_abas_auditadas: int
    xlsx_auditado: bool
    console_auditado: bool


@dataclass(slots=True)
class AuditoriaParidadeXLSX:
    auditado: bool
    caminho: str | None = None
    arquivo_existe: bool = False
    abas_esperadas: list[str] = field(default_factory=list)
    abas_observadas: list[str] = field(default_factory=list)
    abas_faltantes: list[str] = field(default_factory=list)
    abas_extras: list[str] = field(default_factory=list)
    qtd_abas_esperadas: int = 0
    qtd_abas_auditadas: int = 0
    ok: bool = False
    status: str = 'nao_auditado'
    divergencias: list[DivergenciaParidadeRenderizacao] = field(default_factory=list)


@dataclass(slots=True)
class AuditoriaParidadeConsole:
    auditado: bool
    fornecido: bool = False
    ok: bool = False
    status: str = 'nao_auditado'
    secoes_esperadas: list[str] = field(default_factory=list)
    secoes_observadas: list[str] = field(default_factory=list)
    ressalvas: list[str] = field(default_factory=list)
    divergencias: list[DivergenciaParidadeRenderizacao] = field(default_factory=list)


@dataclass(slots=True)
class MetadadosParidadeRenderizacao:
    artefato: str = ARTEFATO_PARIDADE
    etapa: int = ETAPA_PARIDADE
    entrada_formal: str = ENTRADA_FORMAL
    modulo: str = MODULO_PARIDADE
    sem_reotimizacao: bool = True
    sem_revaloracao: bool = True
    sem_alteracao_decisao: bool = True
    sem_consulta_motor: bool = True
    sem_consulta_ledger: bool = True
    sem_consulta_gates: bool = True
    xlsx_auditado: bool = False
    console_auditado: bool = False
    categorias_divergencia: tuple[str, ...] = CATEGORIAS_DIVERGENCIA


@dataclass(slots=True)
class ResultadoParidadeRenderizacaoOficial:
    artefato: str
    etapa: int
    status: str
    ok: bool
    entrada_formal: str
    divergencias: list[DivergenciaParidadeRenderizacao]
    resumo: ResumoParidadeRenderizacaoOficial
    auditoria_xlsx: AuditoriaParidadeXLSX
    auditoria_console: AuditoriaParidadeConsole
    metadados: MetadadosParidadeRenderizacao


def _nova_divergencia(
    categoria: str,
    alvo: str,
    mensagem: str,
    *,
    material: bool | None = None,
    aba: str | None = None,
    linha: int | None = None,
    coluna: str | None = None,
    esperado: Any = None,
    observado: Any = None,
    referencias: dict[str, Any] | None = None,
) -> DivergenciaParidadeRenderizacao:
    categoria_final = categoria if categoria in CATEGORIAS_DIVERGENCIA else 'DIVERGENCIA_MATERIAL'
    return DivergenciaParidadeRenderizacao(
        categoria=categoria_final,
        alvo=alvo,
        mensagem=mensagem,
        material=(categoria_final in _CATEGORIAS_MATERIAIS) if material is None else bool(material),
        aba=aba,
        linha=linha,
        coluna=coluna,
        esperado=esperado,
        observado=observado,
        referencias=referencias or {},
    )


def validar_entrada_paridade_renderizacao(
    pacote_saida_observavel: Any,
) -> list[DivergenciaParidadeRenderizacao]:
    if not isinstance(pacote_saida_observavel, PacoteSaidaObservavelOficial):
        return [
            _nova_divergencia(
                'DIVERGENCIA_ESTRUTURAL',
                'entrada',
                'Entrada da Etapa 10 deve ser PacoteSaidaObservavelOficial.',
                material=True,
                esperado=ENTRADA_FORMAL,
                observado=type(pacote_saida_observavel).__name__,
            )
        ]
    bloco_xlsx = getattr(pacote_saida_observavel, 'bloco_xlsx', None)
    bloco_console = getattr(pacote_saida_observavel, 'bloco_console', None)
    divergencias: list[DivergenciaParidadeRenderizacao] = []
    if bloco_xlsx is None or not isinstance(getattr(bloco_xlsx, 'abas', None), dict):
        divergencias.append(
            _nova_divergencia(
                'DIVERGENCIA_ESTRUTURAL',
                'entrada.bloco_xlsx',
                'PacoteSaidaObservavelOficial.bloco_xlsx.abas deve estar disponível como dicionário.',
                material=True,
            )
        )
    if bloco_console is None:
        divergencias.append(
            _nova_divergencia(
                'DIVERGENCIA_ESTRUTURAL',
                'entrada.bloco_console',
                'PacoteSaidaObservavelOficial.bloco_console deve estar disponível.',
                material=True,
            )
        )
    return divergencias


def _serializar_valor_observavel(valor: Any) -> Any:
    if is_dataclass(valor):
        return asdict(valor)
    if isinstance(valor, (dict, list, tuple, set)):
        return str(valor)
    return valor


def _normalizar_linhas_observaveis_local(linhas: Iterable[Any]) -> tuple[list[str], list[dict[str, Any]]]:
    normalizadas: list[dict[str, Any]] = []
    headers: list[str] = []
    vistos: set[str] = set()
    for item in list(linhas or []):
        if is_dataclass(item):
            item = asdict(item)
        elif not isinstance(item, dict):
            item = {'valor': item}
        normalizado = {str(chave): _serializar_valor_observavel(valor) for chave, valor in dict(item).items()}
        normalizadas.append(normalizado)
        for chave in normalizado:
            if chave not in vistos:
                headers.append(chave)
                vistos.add(chave)
    if not headers:
        headers = ['status']
        normalizadas = [{'status': 'sem_registros_observaveis'}]
    return headers, normalizadas


def _nome_aba_saida_observavel_local(nome_base: str, usados: set[str]) -> str:
    prefixo = PREFIXO_ABA_OBSERVAVEL
    base = ''.join(ch if ch not in '[]:*?/\\' else '-' for ch in str(nome_base or 'Aba'))
    limite_base = 31 - len(prefixo)
    nome = f'{prefixo}{base[:limite_base]}'.strip()
    if nome not in usados:
        usados.add(nome)
        return nome
    contador = 2
    while True:
        sufixo = f' {contador}'
        nome = f'{prefixo}{base[:31 - len(prefixo) - len(sufixo)]}{sufixo}'.strip()
        if nome not in usados:
            usados.add(nome)
            return nome
        contador += 1


def extrair_blocos_esperados_do_pacote(pacote_saida_observavel: PacoteSaidaObservavelOficial) -> dict[str, Any]:
    abas_xlsx: dict[str, dict[str, Any]] = {}
    for nome, headers in ABAS_XLSX_OFICIAIS_CONTRATUAIS.items():
        if nome == 'Situação Atual':
            abas_xlsx[nome] = {
                'nome_base': nome,
                'headers': [],
                'linhas': None,
                'validacao': 'blocos_obrigatorios',
                'blocos_obrigatorios': list(BLOCOS_SITUACAO_ATUAL_OBRIGATORIOS),
            }
        else:
            abas_xlsx[nome] = {
                'nome_base': nome,
                'headers': list(headers),
                'linhas': None,
                'validacao': 'estrutura',
            }

    bloco_console = getattr(pacote_saida_observavel, 'bloco_console', None)
    return {
        'abas_xlsx': abas_xlsx,
        'bloco_console': bloco_console,
    }


def ler_renderizacao_xlsx(caminho_xlsx: Path | str | None) -> dict[str, Any]:
    if caminho_xlsx is None:
        return {'auditavel': False, 'existe': False, 'caminho': None, 'abas': {}}
    caminho = Path(caminho_xlsx)
    if not caminho.exists():
        return {'auditavel': False, 'existe': False, 'caminho': str(caminho), 'abas': {}}

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        return {
            'auditavel': False,
            'existe': True,
            'caminho': str(caminho),
            'abas': {},
            'erro': f'openpyxl indisponivel: {exc}',
        }

    try:
        wb = load_workbook(caminho, data_only=True, read_only=True)
    except Exception as exc:
        return {
            'auditavel': False,
            'existe': True,
            'caminho': str(caminho),
            'abas': {},
            'erro': f'falha ao ler XLSX: {exc}',
        }
    try:
        abas: dict[str, dict[str, Any]] = {}
        for nome in wb.sheetnames:
            ws = wb[nome]
            linhas = list(ws.iter_rows(values_only=True))
            texto_linhas = [
                ' | '.join('' if valor is None else str(valor) for valor in linha)
                for linha in linhas
            ]
            if linhas:
                headers = [str(valor) for valor in linhas[0] if valor is not None]
                registros = []
                for valores in linhas[1:]:
                    registro = {
                        header: valores[idx] if idx < len(valores) else None
                        for idx, header in enumerate(headers)
                    }
                    if any(valor is not None for valor in registro.values()):
                        registros.append(registro)
            else:
                headers = []
                registros = []
            abas[nome] = {
                'headers': headers,
                'linhas': registros,
                'texto_linhas': texto_linhas,
                'qtd_linhas_brutas': len(linhas),
            }
        return {'auditavel': True, 'existe': True, 'caminho': str(caminho), 'abas': abas}
    finally:
        wb.close()


def _objeto_para_mapping(objeto: Any) -> dict[str, Any]:
    if objeto is None:
        return {}
    if is_dataclass(objeto):
        return asdict(objeto)
    if isinstance(objeto, Mapping):
        return dict(objeto)
    return {
        chave: getattr(objeto, chave)
        for chave in dir(objeto)
        if not chave.startswith('_') and not callable(getattr(objeto, chave))
    }


def ler_renderizacao_console(console_renderizado: object | None) -> dict[str, Any]:
    if console_renderizado is None:
        return {'auditavel': False, 'fornecido': False, 'texto': None, 'estrutura': {}}
    if isinstance(console_renderizado, str):
        return {'auditavel': True, 'fornecido': True, 'texto': console_renderizado, 'estrutura': {}}
    return {
        'auditavel': True,
        'fornecido': True,
        'texto': None,
        'estrutura': _objeto_para_mapping(console_renderizado),
    }


def _normalizar_string_iso(valor: str) -> Any:
    texto = valor.strip()
    try:
        if len(texto) == 10 and texto[4] == '-' and texto[7] == '-':
            return date.fromisoformat(texto).isoformat()
        if 'T' in texto:
            dt = datetime.fromisoformat(texto)
            if dt.time().replace(tzinfo=None) == time(0, 0, 0):
                return dt.date().isoformat()
            return dt.isoformat(timespec='seconds')
    except ValueError:
        return valor
    return valor


def _decimal_ou_none(valor: Any) -> Decimal | None:
    if isinstance(valor, bool) or valor is None:
        return None
    if isinstance(valor, Decimal):
        return valor
    if isinstance(valor, int):
        return Decimal(valor)
    if isinstance(valor, float):
        return Decimal(str(valor))
    return None


def normalizar_valores_para_paridade(valor: Any) -> Any:
    if valor is None or isinstance(valor, bool):
        return valor
    if isinstance(valor, datetime):
        if valor.time().replace(tzinfo=None) == time(0, 0, 0):
            return valor.date().isoformat()
        return valor.isoformat(timespec='seconds')
    if isinstance(valor, date):
        return valor.isoformat()
    if isinstance(valor, str):
        return _normalizar_string_iso(valor)
    if isinstance(valor, Decimal):
        return valor
    if isinstance(valor, int):
        return Decimal(valor)
    if isinstance(valor, float):
        return Decimal(str(valor))
    if is_dataclass(valor):
        return {chave: normalizar_valores_para_paridade(v) for chave, v in asdict(valor).items()}
    if isinstance(valor, dict):
        return {str(chave): normalizar_valores_para_paridade(v) for chave, v in valor.items()}
    if isinstance(valor, (list, tuple)):
        return [normalizar_valores_para_paridade(v) for v in valor]
    return valor


def _valores_equivalentes(esperado: Any, observado: Any) -> bool:
    esperado_norm = normalizar_valores_para_paridade(esperado)
    observado_norm = normalizar_valores_para_paridade(observado)
    dec_esp = _decimal_ou_none(esperado_norm)
    dec_obs = _decimal_ou_none(observado_norm)
    if dec_esp is not None and dec_obs is not None:
        try:
            if abs(dec_esp - dec_obs) <= Decimal('0.005'):
                return True
            q_esp = dec_esp.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            q_obs = dec_obs.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            return q_esp == q_obs
        except (InvalidOperation, ValueError):
            return False
    return esperado_norm == observado_norm


def _categoria_conteudo(esperado: Any, observado: Any) -> str:
    if isinstance(esperado, (date, datetime)) or isinstance(observado, (date, datetime)):
        return 'DIVERGENCIA_DATA_DATETIME'
    if _decimal_ou_none(esperado) is not None or _decimal_ou_none(observado) is not None:
        return 'DIVERGENCIA_NORMALIZACAO_NUMERICA'
    if isinstance(esperado, str) != isinstance(observado, str):
        return 'DIVERGENCIA_SERIALIZACAO'
    return 'DIVERGENCIA_CONTEUDO'


def comparar_presenca_estrutura(
    esperadas: Mapping[str, Any],
    observadas: Mapping[str, Any],
    *,
    alvo: str,
) -> list[DivergenciaParidadeRenderizacao]:
    divergencias: list[DivergenciaParidadeRenderizacao] = []
    for nome in esperadas:
        if nome not in observadas:
            divergencias.append(
                _nova_divergencia('ABA_XLSX_AUSENTE', alvo, f'Aba XLSX observável ausente: {nome}.', aba=nome)
            )
    for nome in observadas:
        if nome not in esperadas:
            divergencias.append(
                _nova_divergencia('ABA_XLSX_EXTRA', alvo, f'Aba XLSX observável extra inesperada: {nome}.', aba=nome)
            )
    return divergencias


def comparar_headers(
    aba: str,
    headers_esperados: list[str],
    headers_observados: list[str],
) -> list[DivergenciaParidadeRenderizacao]:
    if list(headers_esperados) == list(headers_observados):
        return []
    return [
        _nova_divergencia(
            'DIVERGENCIA_HEADERS',
            'xlsx',
            f'Headers divergentes na aba {aba}.',
            aba=aba,
            esperado=list(headers_esperados),
            observado=list(headers_observados),
        )
    ]


def comparar_quantidade_linhas(
    aba: str,
    linhas_esperadas: list[Mapping[str, Any]],
    linhas_observadas: list[Mapping[str, Any]],
) -> list[DivergenciaParidadeRenderizacao]:
    if len(linhas_esperadas) == len(linhas_observadas):
        return []
    return [
        _nova_divergencia(
            'DIVERGENCIA_QTD_LINHAS',
            'xlsx',
            f'Quantidade de linhas divergente na aba {aba}.',
            aba=aba,
            esperado=len(linhas_esperadas),
            observado=len(linhas_observadas),
        )
    ]


def comparar_conteudo_normalizado(
    aba: str,
    headers: list[str],
    linhas_esperadas: list[Mapping[str, Any]],
    linhas_observadas: list[Mapping[str, Any]],
) -> list[DivergenciaParidadeRenderizacao]:
    divergencias: list[DivergenciaParidadeRenderizacao] = []
    limite = min(len(linhas_esperadas), len(linhas_observadas))
    for indice in range(limite):
        esperada = linhas_esperadas[indice]
        observada = linhas_observadas[indice]
        for coluna in headers:
            esperado = esperada.get(coluna)
            observado = observada.get(coluna)
            if _valores_equivalentes(esperado, observado):
                continue
            categoria = _categoria_conteudo(esperado, observado)
            divergencias.append(
                _nova_divergencia(
                    categoria,
                    'xlsx',
                    f'Conteúdo divergente na aba {aba}, linha {indice + 2}, coluna {coluna}.',
                    material=True,
                    aba=aba,
                    linha=indice + 2,
                    coluna=coluna,
                    esperado=esperado,
                    observado=observado,
                )
            )
    return divergencias


def classificar_divergencias(
    divergencias: Iterable[DivergenciaParidadeRenderizacao],
) -> list[DivergenciaParidadeRenderizacao]:
    return [
        divergencia
        if divergencia.categoria in CATEGORIAS_DIVERGENCIA
        else _nova_divergencia(
            'DIVERGENCIA_MATERIAL',
            divergencia.alvo,
            divergencia.mensagem,
            material=True,
            aba=divergencia.aba,
            linha=divergencia.linha,
            coluna=divergencia.coluna,
            esperado=divergencia.esperado,
            observado=divergencia.observado,
            referencias=divergencia.referencias,
        )
        for divergencia in divergencias
    ]


def auditar_paridade_xlsx(
    blocos_esperados: Mapping[str, Any],
    caminho_xlsx: Path | str | None = None,
) -> AuditoriaParidadeXLSX:
    esperado = dict(blocos_esperados.get('abas_xlsx', {}) or {})
    lido = ler_renderizacao_xlsx(caminho_xlsx)
    caminho = lido.get('caminho')
    if not lido['auditavel']:
        divergencias = []
        if caminho_xlsx is not None:
            divergencias.append(
                _nova_divergencia(
                    'ARTEFATO_RENDERIZADO_AUSENTE',
                    'xlsx',
                    'XLSX informado para auditoria não existe ou não está acessível.',
                    material=True,
                    esperado='arquivo XLSX existente',
                    observado=lido.get('erro') or caminho,
                )
            )
        return AuditoriaParidadeXLSX(
            auditado=False,
            caminho=caminho,
            arquivo_existe=bool(lido['existe']),
            abas_esperadas=list(esperado),
            qtd_abas_esperadas=len(esperado),
            status='bloqueado' if divergencias else 'nao_auditado',
            divergencias=divergencias,
        )

    observado = dict(lido.get('abas', {}) or {})
    divergencias = comparar_presenca_estrutura(esperado, observado, alvo='xlsx')

    for aba in sorted(set(esperado) & set(observado)):
        espec = esperado[aba]
        modo_validacao = espec.get('validacao', 'estrutura')
        observado_aba = observado[aba]

        if modo_validacao == 'blocos_obrigatorios':
            texto_aba = '\n'.join(str(linha) for linha in observado_aba.get('texto_linhas', []) or [])
            for bloco in espec.get('blocos_obrigatorios', []) or []:
                if str(bloco) not in texto_aba:
                    divergencias.append(
                        _nova_divergencia(
                            'DIVERGENCIA_ESTRUTURAL',
                            'xlsx',
                            f'Bloco obrigatório ausente na aba {aba}: {bloco}.',
                            material=True,
                            aba=aba,
                            esperado=bloco,
                        )
                    )
            continue

        headers_esperados = list(espec.get('headers') or [])
        headers_observados = list(observado_aba.get('headers') or [])
        divergencias.extend(comparar_headers(aba, headers_esperados, headers_observados))

        linhas_observadas = list(observado_aba.get('linhas') or [])
        if not linhas_observadas:
            divergencias.append(
                _nova_divergencia(
                    'DIVERGENCIA_QTD_LINHAS',
                    'xlsx',
                    f'Aba {aba} não possui linhas de dados observáveis.',
                    material=True,
                    aba=aba,
                    esperado='ao menos 1 linha de dados',
                    observado=0,
                )
            )

        linhas_esperadas = espec.get('linhas')
        if linhas_esperadas is not None:
            linhas_esperadas = list(linhas_esperadas)
            divergencias.extend(comparar_quantidade_linhas(aba, linhas_esperadas, linhas_observadas))
            if headers_esperados == headers_observados:
                divergencias.extend(
                    comparar_conteudo_normalizado(aba, headers_esperados, linhas_esperadas, linhas_observadas)
                )

    divergencias = classificar_divergencias(divergencias)
    ok = not any(div.material for div in divergencias)
    return AuditoriaParidadeXLSX(
        auditado=True,
        caminho=caminho,
        arquivo_existe=True,
        abas_esperadas=list(esperado),
        abas_observadas=list(observado),
        abas_faltantes=[nome for nome in esperado if nome not in observado],
        abas_extras=[nome for nome in observado if nome not in esperado],
        qtd_abas_esperadas=len(esperado),
        qtd_abas_auditadas=len(observado),
        ok=ok,
        status='aprovado' if ok else 'reprovado',
        divergencias=divergencias,
    )


def _secoes_console_esperadas(bloco_console: Any) -> list[str]:
    return [chave for chave, valor in _objeto_para_mapping(bloco_console).items() if valor not in (None, {}, [])]


def _texto_para_decimal_console(token: str) -> Decimal | None:
    token_limpo = token.strip()
    if not token_limpo:
        return None
    separadores = [sep for sep in ('.', ',') if sep in token_limpo]
    if len(separadores) > 1:
        return None
    if separadores:
        separador = separadores[0]
        parte_inteira, parte_decimal = token_limpo.rsplit(separador, 1)
        if len(parte_decimal) > 2:
            return None
        token_limpo = f'{parte_inteira}.{parte_decimal}' if separador == ',' else token_limpo
    try:
        return Decimal(token_limpo)
    except InvalidOperation:
        return None


def _tokens_numericos_console(texto: str) -> list[Decimal]:
    tokens = re.findall(r'(?<![\w])[-+]?\d+(?:[.,]\d+)?(?![\w])', texto)
    decimais: list[Decimal] = []
    for token in tokens:
        decimal = _texto_para_decimal_console(token)
        if decimal is not None:
            decimais.append(decimal)
    return decimais


def _texto_contem_valor(texto: str, valor: Any) -> bool:
    valor_norm = normalizar_valores_para_paridade(valor)
    valor_decimal = _decimal_ou_none(valor_norm)
    if valor_decimal is not None:
        return any(_valores_equivalentes(valor_decimal, token) for token in _tokens_numericos_console(texto))

    candidatos = {str(valor), str(valor_norm)}
    return any(candidato in texto for candidato in candidatos if candidato not in {'', 'None'})


def _rotulo_console_presente(texto_normalizado: str, campo: Any) -> bool:
    rotulo = str(campo).replace('_', ' ').lower()
    return rotulo in texto_normalizado


def _comparar_secao_console_estruturada(
    secao: str,
    esperado: Any,
    observado: Any,
) -> list[DivergenciaParidadeRenderizacao]:
    rotulo = ROTULOS_SECOES_CONSOLE_PARIDADE_FORTE.get(secao, secao)
    if not _valores_equivalentes(esperado, observado):
        return [
            _nova_divergencia(
                _categoria_conteudo(esperado, observado),
                'console',
                f'Bloco material do console diverge do PacoteSaidaObservavelOficial: {rotulo}.',
                material=True,
                coluna=secao,
                esperado=normalizar_valores_para_paridade(esperado),
                observado=normalizar_valores_para_paridade(observado),
                referencias={
                    'secao_presente': True,
                    'campo_presente': True,
                    'valor_equivalente': False,
                    'rotulo': rotulo,
                },
            )
        ]
    return []


def _auditar_console_estruturado_forte(
    bloco_console: Any,
    estrutura: Mapping[str, Any],
) -> tuple[list[str], list[DivergenciaParidadeRenderizacao]]:
    bloco_esperado = _objeto_para_mapping(bloco_console)
    secoes_observadas = [secao for secao in SECOES_CONSOLE_PARIDADE_FORTE if secao in estrutura]
    divergencias: list[DivergenciaParidadeRenderizacao] = []
    for secao in SECOES_CONSOLE_PARIDADE_FORTE:
        rotulo = ROTULOS_SECOES_CONSOLE_PARIDADE_FORTE.get(secao, secao)
        esperado = bloco_esperado.get(secao)
        if secao not in estrutura:
            divergencias.append(
                _nova_divergencia(
                    'DIVERGENCIA_ESTRUTURAL',
                    'console',
                    f'Seção material não localizada na representação estruturada de console: {rotulo}.',
                    material=True,
                    coluna=secao,
                    esperado=normalizar_valores_para_paridade(esperado),
                    referencias={
                        'secao_presente': False,
                        'campo_presente': False,
                        'valor_equivalente': False,
                        'rotulo': rotulo,
                    },
                )
            )
            continue
        divergencias.extend(_comparar_secao_console_estruturada(secao, esperado, estrutura.get(secao)))
    return secoes_observadas, divergencias


def auditar_paridade_console(
    blocos_esperados: Mapping[str, Any],
    console_renderizado: object | None = None,
) -> AuditoriaParidadeConsole:
    bloco_console = blocos_esperados.get('bloco_console')
    secoes_esperadas = _secoes_console_esperadas(bloco_console)
    lido = ler_renderizacao_console(console_renderizado)
    if not lido['auditavel']:
        divergencia = _nova_divergencia(
            'CONSOLE_NAO_AUDITADO',
            'console',
            'Console renderizado não foi fornecido; auditoria de console não executada nesta frente.',
            material=False,
        )
        return AuditoriaParidadeConsole(
            auditado=False,
            fornecido=False,
            ok=False,
            status='nao_auditado',
            secoes_esperadas=secoes_esperadas,
            ressalvas=[divergencia.mensagem],
            divergencias=[divergencia],
        )

    divergencias: list[DivergenciaParidadeRenderizacao] = []
    secoes_observadas: list[str] = []
    texto = lido.get('texto')
    estrutura = dict(lido.get('estrutura') or {})
    if texto is not None:
        texto_normalizado = texto.lower()
        secoes_observadas = [secao for secao in secoes_esperadas if secao.replace('_', ' ').lower() in texto_normalizado]
        resumo = _objeto_para_mapping(bloco_console).get('resumo_operacional', {}) or {}
        for campo, valor in dict(resumo).items():
            rotulo_presente = _rotulo_console_presente(texto_normalizado, campo)
            valor_presente = _texto_contem_valor(texto, valor)
            if not rotulo_presente or not valor_presente:
                divergencias.append(
                    _nova_divergencia(
                        'CONSOLE_AUDITADO_COM_RESSALVA',
                        'console',
                        f'Campo mínimo de resumo operacional divergente no console: {campo}.',
                        material=False,
                        coluna=str(campo),
                        esperado=valor,
                        referencias={
                            'rotulo_presente': rotulo_presente,
                            'valor_presente': valor_presente,
                        },
                    )
                )
    else:
        secoes_observadas, divergencias_fortes = _auditar_console_estruturado_forte(bloco_console, estrutura)
        divergencias.extend(divergencias_fortes)
    divergencias = classificar_divergencias(divergencias)
    ok = not divergencias
    possui_material = any(div.material for div in divergencias)
    return AuditoriaParidadeConsole(
        auditado=True,
        fornecido=True,
        ok=ok,
        status='aprovado' if ok else ('reprovado' if possui_material else 'aprovado_com_ressalva'),
        secoes_esperadas=secoes_esperadas,
        secoes_observadas=secoes_observadas,
        ressalvas=[div.mensagem for div in divergencias if not div.material],
        divergencias=divergencias,
    )


def montar_metadados_paridade(
    auditoria_xlsx: AuditoriaParidadeXLSX,
    auditoria_console: AuditoriaParidadeConsole,
) -> MetadadosParidadeRenderizacao:
    return MetadadosParidadeRenderizacao(
        xlsx_auditado=auditoria_xlsx.auditado,
        console_auditado=auditoria_console.auditado,
    )


def consolidar_resultado_paridade(
    divergencias_entrada: list[DivergenciaParidadeRenderizacao],
    auditoria_xlsx: AuditoriaParidadeXLSX,
    auditoria_console: AuditoriaParidadeConsole,
) -> ResultadoParidadeRenderizacaoOficial:
    divergencias = classificar_divergencias(
        list(divergencias_entrada) + list(auditoria_xlsx.divergencias) + list(auditoria_console.divergencias)
    )
    qtd_materiais = sum(1 for div in divergencias if div.material)
    qtd_ressalvas = sum(1 for div in divergencias if not div.material)
    if divergencias_entrada or auditoria_xlsx.status == 'bloqueado':
        status = 'bloqueado'
    elif qtd_materiais:
        status = 'reprovado'
    elif qtd_ressalvas or not auditoria_console.auditado or not auditoria_xlsx.auditado:
        status = 'aprovado_com_ressalva'
    else:
        status = 'aprovado'
    ok = status == 'aprovado'
    resumo = ResumoParidadeRenderizacaoOficial(
        status=status,
        ok=ok,
        qtd_divergencias=len(divergencias),
        qtd_divergencias_materiais=qtd_materiais,
        qtd_ressalvas=qtd_ressalvas,
        qtd_abas_esperadas=auditoria_xlsx.qtd_abas_esperadas,
        qtd_abas_auditadas=auditoria_xlsx.qtd_abas_auditadas,
        xlsx_auditado=auditoria_xlsx.auditado,
        console_auditado=auditoria_console.auditado,
    )
    metadados = montar_metadados_paridade(auditoria_xlsx, auditoria_console)
    return ResultadoParidadeRenderizacaoOficial(
        artefato=ARTEFATO_PARIDADE,
        etapa=ETAPA_PARIDADE,
        status=status,
        ok=ok,
        entrada_formal=ENTRADA_FORMAL,
        divergencias=divergencias,
        resumo=resumo,
        auditoria_xlsx=auditoria_xlsx,
        auditoria_console=auditoria_console,
        metadados=metadados,
    )


def validar_paridade_renderizacao_oficial(
    pacote_saida_observavel: PacoteSaidaObservavelOficial,
    caminho_xlsx: Path | str | None = None,
    console_renderizado: object | None = None,
) -> ResultadoParidadeRenderizacaoOficial:
    divergencias_entrada = validar_entrada_paridade_renderizacao(pacote_saida_observavel)
    if divergencias_entrada:
        auditoria_xlsx = AuditoriaParidadeXLSX(auditado=False, status='bloqueado')
        auditoria_console = AuditoriaParidadeConsole(auditado=False, status='bloqueado')
        return consolidar_resultado_paridade(divergencias_entrada, auditoria_xlsx, auditoria_console)

    blocos_esperados = extrair_blocos_esperados_do_pacote(pacote_saida_observavel)
    auditoria_xlsx = auditar_paridade_xlsx(blocos_esperados, caminho_xlsx)
    auditoria_console = auditar_paridade_console(blocos_esperados, console_renderizado)
    return consolidar_resultado_paridade([], auditoria_xlsx, auditoria_console)
