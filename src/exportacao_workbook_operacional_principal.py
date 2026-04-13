
from __future__ import annotations

from datetime import datetime, date
from pathlib import Path
from zoneinfo import ZoneInfo

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _reference_date_brasilia() -> date:
    return datetime.now(ZoneInfo("America/Sao_Paulo")).date()


def _signed_business_days(start_date: date, ref_date: date) -> int:
    if start_date == ref_date:
        return 0
    if start_date < ref_date:
        return int(np.busday_count(start_date, ref_date))
    return -int(np.busday_count(ref_date, start_date))


def _classify_lote(investimento: str, data_aplicacao: date, ref_date: date) -> str:
    inv = (investimento or "").strip()
    if inv == "-":
        return "BLOQUEADOS_JA_GASTOS"
    if inv == "":
        if data_aplicacao > ref_date:
            return "LIVRES_FUTUROS"
        return "LIVRES_DISPONIVEIS"
    if data_aplicacao <= ref_date:
        return "INVESTIDOS_ATUAIS"
    return "INVESTIDOS_FUTUROS"


def build_lotes_operacionais(raw_workbook_path: str | Path, reference_workbook_path: str | Path) -> dict[str, pd.DataFrame]:
    raw_workbook_path = Path(raw_workbook_path)
    reference_workbook_path = Path(reference_workbook_path)

    ref_date = _reference_date_brasilia()

    lotes_raw = pd.read_excel(raw_workbook_path, sheet_name="Inventário de Lotes")
    lotes_raw.columns = ["Lote", "Data_Aplicacao", "Valor_Original_R$", "Investimento"]
    lotes_raw["Data_Aplicacao"] = pd.to_datetime(lotes_raw["Data_Aplicacao"])

    situacao = pd.read_excel(reference_workbook_path, sheet_name="Situacao Atual")
    situacao_small = situacao[[
        "Lote ID", "Saldo Bruto Atual (R$)", "Saldo Líquido Atual (R$)",
        "Total Bruto Sacado (R$)", "Total Líquido Sacado (R$)", "Vezes Usado (Futuro)"
    ]].copy().rename(columns={
        "Lote ID": "Lote",
        "Saldo Bruto Atual (R$)": "Saldo_Bruto_R$",
        "Saldo Líquido Atual (R$)": "Saldo_Liquido_R$",
        "Total Bruto Sacado (R$)": "Total_Bruto_Sacado_R$",
        "Total Líquido Sacado (R$)": "Total_Liquido_Sacado_R$",
        "Vezes Usado (Futuro)": "Vezes_Usado_Futuro",
    })

    lotes = lotes_raw.merge(situacao_small, on="Lote", how="left")
    lotes["Investimento"] = lotes["Investimento"].fillna("")
    lotes["Classe_Lote"] = lotes.apply(
        lambda r: _classify_lote(str(r["Investimento"]), r["Data_Aplicacao"].date(), ref_date),
        axis=1,
    )

    lotes["Dias_Corridos"] = lotes["Data_Aplicacao"].dt.date.map(lambda d: (ref_date - d).days)
    lotes["Dias_Uteis"] = lotes["Data_Aplicacao"].dt.date.map(lambda d: _signed_business_days(d, ref_date))

    # Fill operational values using Inventário as primary source
    for idx, row in lotes.iterrows():
        classe = row["Classe_Lote"]
        valor_original = float(row["Valor_Original_R$"]) if pd.notna(row["Valor_Original_R$"]) else 0.0

        if classe == "INVESTIDOS_ATUAIS":
            if pd.isna(row["Saldo_Bruto_R$"]):
                lotes.at[idx, "Saldo_Bruto_R$"] = valor_original
            if pd.isna(row["Saldo_Liquido_R$"]):
                lotes.at[idx, "Saldo_Liquido_R$"] = valor_original
            if pd.isna(row["Total_Bruto_Sacado_R$"]):
                lotes.at[idx, "Total_Bruto_Sacado_R$"] = 0.0
            if pd.isna(row["Total_Liquido_Sacado_R$"]):
                lotes.at[idx, "Total_Liquido_Sacado_R$"] = 0.0
            if pd.isna(row["Vezes_Usado_Futuro"]):
                lotes.at[idx, "Vezes_Usado_Futuro"] = 0

        elif classe in {"LIVRES_DISPONIVEIS", "LIVRES_FUTUROS", "INVESTIDOS_FUTUROS"}:
            lotes.at[idx, "Saldo_Bruto_R$"] = valor_original if pd.isna(row["Saldo_Bruto_R$"]) else row["Saldo_Bruto_R$"]
            lotes.at[idx, "Saldo_Liquido_R$"] = valor_original if pd.isna(row["Saldo_Liquido_R$"]) else row["Saldo_Liquido_R$"]
            lotes.at[idx, "Total_Bruto_Sacado_R$"] = 0.0 if pd.isna(row["Total_Bruto_Sacado_R$"]) else row["Total_Bruto_Sacado_R$"]
            lotes.at[idx, "Total_Liquido_Sacado_R$"] = 0.0 if pd.isna(row["Total_Liquido_Sacado_R$"]) else row["Total_Liquido_Sacado_R$"]
            lotes.at[idx, "Vezes_Usado_Futuro"] = 0 if pd.isna(row["Vezes_Usado_Futuro"]) else row["Vezes_Usado_Futuro"]

        elif classe == "BLOQUEADOS_JA_GASTOS":
            lotes.at[idx, "Saldo_Bruto_R$"] = 0.0 if pd.isna(row["Saldo_Bruto_R$"]) else row["Saldo_Bruto_R$"]
            lotes.at[idx, "Saldo_Liquido_R$"] = 0.0 if pd.isna(row["Saldo_Liquido_R$"]) else row["Saldo_Liquido_R$"]
            lotes.at[idx, "Total_Bruto_Sacado_R$"] = valor_original if pd.isna(row["Total_Bruto_Sacado_R$"]) else row["Total_Bruto_Sacado_R$"]
            lotes.at[idx, "Total_Liquido_Sacado_R$"] = valor_original if pd.isna(row["Total_Liquido_Sacado_R$"]) else row["Total_Liquido_Sacado_R$"]
            lotes.at[idx, "Vezes_Usado_Futuro"] = 0 if pd.isna(row["Vezes_Usado_Futuro"]) else row["Vezes_Usado_Futuro"]

    lotes["Status"] = lotes["Classe_Lote"]
    lotes["Carteira"] = lotes["Investimento"].replace("", "(livre)")
    lotes = lotes[[
        "Lote", "Data_Aplicacao", "Valor_Original_R$", "Investimento", "Carteira", "Classe_Lote", "Status",
        "Dias_Corridos", "Dias_Uteis", "Saldo_Bruto_R$", "Saldo_Liquido_R$",
        "Total_Bruto_Sacado_R$", "Total_Liquido_Sacado_R$", "Vezes_Usado_Futuro"
    ]].sort_values(["Data_Aplicacao", "Lote"]).reset_index(drop=True)

    return {
        "INVESTIDOS_ATUAIS": lotes[lotes["Classe_Lote"].eq("INVESTIDOS_ATUAIS")].copy(),
        "LIVRES_DISPONIVEIS": lotes[lotes["Classe_Lote"].eq("LIVRES_DISPONIVEIS")].copy(),
        "LIVRES_FUTUROS": lotes[lotes["Classe_Lote"].eq("LIVRES_FUTUROS")].copy(),
        "BLOQUEADOS_JA_GASTOS": lotes[lotes["Classe_Lote"].eq("BLOQUEADOS_JA_GASTOS")].copy(),
        "ALL": lotes.copy(),
        "REFERENCE_DATE": pd.DataFrame([{"Data_Referencia_Brasilia": pd.Timestamp(ref_date), "Fuso": "America/Sao_Paulo"}]),
    }


def _write_df(ws, df: pd.DataFrame, title: str, notes: str) -> None:
    fill_header = PatternFill("solid", fgColor="1F4E78")
    fill_section = PatternFill("solid", fgColor="D9EAF7")
    fill_title = PatternFill("solid", fgColor="0F243E")
    font_header = Font(color="FFFFFF", bold=True)
    font_title = Font(color="FFFFFF", bold=True, size=12)
    thin_top = Side(style="thin", color="808080")
    border_top = Border(top=thin_top)
    align_center = Alignment(horizontal="center", vertical="center")
    currency_fmt = 'R$ #,##0.00;[Red](R$ #,##0.00);-'

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(2, len(df.columns)))
    c = ws.cell(row=row, column=1, value=title)
    c.fill = fill_title
    c.font = font_title
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(2, len(df.columns)))
    c = ws.cell(row=row, column=1, value=notes)
    c.fill = fill_section
    c.alignment = Alignment(wrap_text=True)
    row += 2

    hdr_row = row
    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=hdr_row, column=j, value=str(col))
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center
        cell.border = border_top

    for i, (_, rec) in enumerate(df.iterrows(), start=hdr_row + 1):
        for j, col in enumerate(df.columns, start=1):
            val = rec[col]
            cell = ws.cell(row=i, column=j, value=None if pd.isna(val) else val)
            if hasattr(val, "to_pydatetime"):
                cell.value = val.to_pydatetime()
                cell.number_format = "dd/mm/yyyy"

    if len(df) > 0:
        ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(len(df.columns))}{hdr_row + len(df)}"
    ws.freeze_panes = f"A{hdr_row + 1}"

    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max(len(str(col_name)), 12)
        if len(df) > 0:
            sample = df[col_name].head(100).map(lambda x: "" if pd.isna(x) else str(x))
            max_len_sample = int(sample.map(len).max()) if len(sample) > 0 else 0
            max_len = max(max_len, min(60, max_len_sample))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

        header = str(col_name)
        if any(token in header for token in ["R$", "Valor", "Saldo", "Liquido", "Líquido", "Bruto", "Ganho", "Diferença"]):
            for r in range(hdr_row + 1, hdr_row + len(df) + 1):
                ws.cell(r, col_idx).number_format = currency_fmt


def build_operational_workbook(
    raw_workbook_path: str | Path,
    reference_workbook_path: str | Path,
    official_switchings_csv: str | Path,
    official_resgates_csv: str | Path,
    official_timeline_csv: str | Path,
    output_path: str | Path,
) -> Path:
    raw_workbook_path = Path(raw_workbook_path)
    reference_workbook_path = Path(reference_workbook_path)
    output_path = Path(output_path)

    gastos_raw = pd.read_excel(raw_workbook_path, sheet_name="Todos os Gastos", usecols=[0, 1, 2, 3, 4, 5])
    gastos_raw.columns = ["Data", "Descricao", "Valor", "Pago", "Lote_usado_1", "Lote_usado_2"]
    gastos_raw["Data"] = pd.to_datetime(gastos_raw["Data"])
    gastos_raw["Pago"] = gastos_raw["Pago"].fillna("")

    ref_extrato = pd.read_excel(reference_workbook_path, sheet_name="Extrato")
    ref_situacao = pd.read_excel(reference_workbook_path, sheet_name="Situacao Atual")

    switchings = pd.read_csv(official_switchings_csv)
    resgates = pd.read_csv(official_resgates_csv)
    timeline = pd.read_csv(official_timeline_csv)

    lotes_views = build_lotes_operacionais(raw_workbook_path, reference_workbook_path)

    gastos_hist = gastos_raw[gastos_raw["Pago"].astype(str).str.strip().str.upper().eq("OK")].copy()
    gastos_fut = gastos_raw[~gastos_raw["Pago"].astype(str).str.strip().str.upper().eq("OK")].copy()
    gastos_fut = gastos_fut.sort_values(["Data", "Descricao", "Valor"]).reset_index(drop=True)

    lotes_raw = pd.read_excel(raw_workbook_path, sheet_name="Inventário de Lotes")
    lotes_raw.columns = ["Lote ID", "Data Aplicacao", "Valor Original", "Investimento"]
    lote_to_carteira = lotes_raw.set_index("Lote ID")["Investimento"].fillna("").astype(str).to_dict()
    lote_to_valor_original = lotes_raw.set_index("Lote ID")["Valor Original"].to_dict()
    lote_to_data = lotes_raw.set_index("Lote ID")["Data Aplicacao"].to_dict()

    historico_pag = gastos_hist.copy()
    historico_pag["Tipo_Evento"] = "PAGAMENTO_REALIZADO"
    historico_pag["Carteira_Origem"] = historico_pag["Lote_usado_1"].map(lote_to_carteira)
    historico_pag["Valor_Original_Lote"] = historico_pag["Lote_usado_1"].map(lote_to_valor_original)
    historico_pag["Data_Aplicacao_Lote"] = historico_pag["Lote_usado_1"].map(lote_to_data)
    historico_pag = historico_pag[
        ["Data", "Tipo_Evento", "Descricao", "Valor", "Lote_usado_1", "Lote_usado_2", "Carteira_Origem", "Valor_Original_Lote", "Data_Aplicacao_Lote"]
    ].rename(columns={
        "Descricao": "Conta",
        "Valor": "Valor_Liquido_Pago_R$",
        "Lote_usado_1": "Lote_Origem_1",
        "Lote_usado_2": "Lote_Origem_2",
    })

    timeline["data"] = pd.to_datetime(timeline["data"])
    switchings["data_switching"] = pd.to_datetime(switchings["data_switching"])
    resgates["data"] = pd.to_datetime(resgates["data"])

    resg_group = resgates.groupby("data").agg(
        qtd_resgates=("id_lote", "count"),
        lote_principal=("id_lote", lambda s: ", ".join(pd.Series(s).dropna().astype(str).unique())),
        carteira_resgate=("carteira", lambda s: ", ".join(pd.Series(s).dropna().astype(str).unique())),
        valor_resgatado_total_centavos=("valor_resgatado_centavos", "sum"),
    ).reset_index()

    plano_futuro = gastos_fut.copy()
    plano_futuro["Estrategia_Dia"] = plano_futuro["Data"].map(dict(zip(timeline["data"], timeline["estrategia_escolhida"])))
    plano_futuro["Deficit_Dia_R$"] = plano_futuro["Data"].map(dict(zip(timeline["data"], timeline["deficit_sem_acao_centavos"] / 100)))
    plano_futuro["Fez_Switching?"] = plano_futuro["Data"].map(dict(zip(timeline["data"], timeline["switching_realizado"])))
    plano_futuro["Lote_Switching"] = plano_futuro["Data"].map(dict(zip(timeline["data"], timeline["id_lote_origem_switching"])))
    plano_futuro["Carteira_Destino_Switching"] = plano_futuro["Data"].map(dict(zip(timeline["data"], timeline["carteira_destino_switching"])))
    plano_futuro["Valor_Switching_R$"] = plano_futuro["Data"].map(dict(zip(timeline["data"], timeline["valor_switching_centavos"] / 100)))
    plano_futuro["Lote_Resgate"] = plano_futuro["Data"].map(resg_group.set_index("data")["lote_principal"].to_dict())
    plano_futuro["Carteira_Resgate"] = plano_futuro["Data"].map(resg_group.set_index("data")["carteira_resgate"].to_dict())
    plano_futuro["Valor_Resgatado_Dia_R$"] = plano_futuro["Data"].map(resg_group.set_index("data")["valor_resgatado_total_centavos"].div(100).to_dict())
    plano_futuro["Observacao_Dia"] = plano_futuro["Data"].map(dict(zip(timeline["data"], timeline["observacao"])))
    plano_futuro = plano_futuro.rename(columns={"Descricao": "Conta", "Valor": "Valor_Conta_R$"})[
        ["Data", "Conta", "Valor_Conta_R$", "Estrategia_Dia", "Deficit_Dia_R$", "Fez_Switching?", "Lote_Switching",
         "Carteira_Destino_Switching", "Valor_Switching_R$", "Lote_Resgate", "Carteira_Resgate", "Valor_Resgatado_Dia_R$", "Observacao_Dia"]
    ]

    switchings_prop = switchings.copy()
    switchings_prop["valor_switching_R$"] = switchings_prop["valor_switching_centavos"] / 100
    switchings_prop["ganho_vs_base_R$"] = switchings_prop["ganho_vs_base_centavos"] / 100
    switchings_prop["riqueza_terminal_base_R$"] = switchings_prop["riqueza_terminal_base_centavos"] / 100
    switchings_prop["riqueza_terminal_switch_R$"] = switchings_prop["riqueza_terminal_switch_centavos"] / 100
    switchings_prop["carteira_origem"] = switchings_prop["id_lote_origem"].map(lote_to_carteira)
    switchings_prop = switchings_prop.rename(columns={
        "data_switching": "Data",
        "id_lote_origem": "Lote_Origem",
        "id_lote_destino": "Lote_Destino",
        "carteira_destino": "Carteira_Destino",
        "tipo_switching": "Tipo",
        "motivo": "Motivo",
    })[
        ["Data", "Lote_Origem", "carteira_origem", "Lote_Destino", "Carteira_Destino", "Tipo",
         "valor_switching_R$", "ganho_vs_base_R$", "riqueza_terminal_base_R$", "riqueza_terminal_switch_R$", "Motivo"]
    ].rename(columns={"carteira_origem": "Carteira_Origem", "valor_switching_R$": "Valor_Switching_R$"})

    sw_group = switchings_prop.groupby(["Lote_Origem", "Carteira_Origem", "Carteira_Destino"], dropna=False).agg(
        Qtd_Eventos=("Data", "count"),
        Primeira_Data=("Data", "min"),
        Ultima_Data=("Data", "max"),
        Valor_Total_R=("Valor_Switching_R$", "sum"),
        Ganho_Total_R=("ganho_vs_base_R$", "sum"),
    ).reset_index().rename(columns={"Valor_Total_R": "Valor_Total_R$", "Ganho_Total_R": "Ganho_Total_R$"})
    sw_dates = switchings_prop.groupby(["Lote_Origem", "Carteira_Destino"])["Data"].apply(
        lambda s: ", ".join(pd.to_datetime(s).dt.strftime("%d/%m/%Y"))
    ).reset_index(name="Datas_Envolvidas")
    sw_group = sw_group.merge(sw_dates, on=["Lote_Origem", "Carteira_Destino"], how="left")
    sw_group["Fragmentado_em_varios_dias?"] = sw_group["Qtd_Eventos"].gt(1).map({True: "SIM", False: "NAO"})
    sw_group["Por_que_em_varios_dias"] = sw_group["Qtd_Eventos"].apply(
        lambda n: "Decisão recalculada por data crítica; o motor só aciona novo switching quando surge novo déficit." if n > 1 else "Switching único suficiente para a necessidade observada."
    )

    justificativa = pd.DataFrame([
        {
            "Lote_Origem": row["Lote_Origem"],
            "Carteira_Origem": row["Carteira_Origem"],
            "Carteira_Destino": row["Carteira_Destino"],
            "Datas_Envolvidas": row["Datas_Envolvidas"],
            "Valor_Total_R$": row["Valor_Total_R$"],
            "Qtd_Eventos": row["Qtd_Eventos"],
            "Por_que_essa_carteira": "Foi a carteira vencedora entre os destinos elegíveis nas datas críticas, maximizando a riqueza terminal no horizonte remanescente.",
            "Por_que_nao_tudo_de_uma_vez": "A política oficial v2 decide por data crítica. O mesmo lote pode ser reutilizado em novas datas quando um novo déficit aparece.",
            "Por_que_nao_outro_lote": "Os demais lotes elegíveis tiveram pior resultado intertemporal, foram bloqueados por regras do motor ou não melhoraram a riqueza terminal.",
            "Observacao": "Esta aba resume a lógica; a aba agrupada mostra a soma do plano por origem/destino.",
        }
        for _, row in sw_group.iterrows()
    ])

    all_lotes = lotes_views["ALL"].copy()
    carteira_final = all_lotes[[
        "Lote", "Data_Aplicacao", "Valor_Original_R$", "Investimento", "Classe_Lote",
        "Saldo_Bruto_R$", "Saldo_Liquido_R$", "Total_Bruto_Sacado_R$", "Total_Liquido_Sacado_R$",
        "Vezes_Usado_Futuro"
    ]].rename(columns={"Lote": "Lote_ID", "Investimento": "Carteira_Atual", "Classe_Lote": "Status_Final_Operacional"})

    future_res_by_lot = resgates.groupby("id_lote")["valor_resgatado_centavos"].sum().div(100).to_dict()
    future_sw_count = switchings.groupby("id_lote_origem")["valor_switching_centavos"].count().to_dict()
    future_sw_total = switchings.groupby("id_lote_origem")["valor_switching_centavos"].sum().div(100).to_dict()
    carteira_final["Resgates_Futuros_Oficiais_R$"] = carteira_final["Lote_ID"].map(future_res_by_lot).fillna(0)
    carteira_final["Qtd_Switchings_Futuros"] = carteira_final["Lote_ID"].map(future_sw_count).fillna(0).astype(int)
    carteira_final["Valor_Switchings_Futuros_R$"] = carteira_final["Lote_ID"].map(future_sw_total).fillna(0)
    carteira_final["Saldo_Liquido_Final_Estimado_R$"] = (carteira_final["Saldo_Liquido_R$"] - carteira_final["Resgates_Futuros_Oficiais_R$"]).round(2)

    first_critical = timeline.loc[timeline["deficit_sem_acao_centavos"] > 0, "data"].min()
    last_critical = timeline.loc[timeline["deficit_sem_acao_centavos"] > 0, "data"].max()
    lote_dominante = switchings_prop["Lote_Origem"].mode().iat[0]
    destino_dominante = switchings_prop["Carteira_Destino"].mode().iat[0]
    ref_date = _reference_date_brasilia()

    resumo = pd.DataFrame([
        ["Status da política oficial v2", "APPROVED"],
        ["Data de corte", "2026-04-10"],
        ["Data de referência (Brasília)", ref_date.strftime("%d/%m/%Y")],
        ["Fuso horário", "America/Sao_Paulo"],
        ["Horizonte final", "2027-03-31"],
        ["Caixa livre inicial (R$)", 8640.25],
        ["Recebidos futuros livres (R$)", 146680.00],
        ["Pagamentos futuros totais (R$)", round(float(gastos_fut["Valor"].sum()), 2)],
        ["Riqueza terminal política base (R$)", 10136.69],
        ["Riqueza terminal política oficial v2 (R$)", 10514.93],
        ["Ganho total vs base (R$)", 378.24],
        ["Total de switchings oficiais", int(len(switchings_prop))],
        ["Total de resgates oficiais", int(len(resgates))],
        ["Primeira data crítica", first_critical.strftime("%d/%m/%Y") if pd.notna(first_critical) else ""],
        ["Última data crítica", last_critical.strftime("%d/%m/%Y") if pd.notna(last_critical) else ""],
        ["Lote dominante na política", lote_dominante],
        ["Carteira destino dominante", destino_dominante],
    ], columns=["Indicador", "Valor"])

    conferencia = pd.DataFrame([
        ["Historico_Pagamentos carregado", len(historico_pag)],
        ["Pagamentos futuros carregados", len(plano_futuro)],
        ["Lotes investidos atuais", len(lotes_views["INVESTIDOS_ATUAIS"])],
        ["Lotes livres disponíveis", len(lotes_views["LIVRES_DISPONIVEIS"])],
        ["Lotes livres futuros", len(lotes_views["LIVRES_FUTUROS"])],
        ["Lotes bloqueados/já gastos", len(lotes_views["BLOQUEADOS_JA_GASTOS"])],
        ["Switchings oficiais carregados", len(switchings_prop)],
        ["Resgates oficiais carregados", len(resgates)],
        ["Valor total switchings oficiais (R$)", round(switchings_prop["Valor_Switching_R$"].sum(), 2)],
        ["Valor total resgates oficiais (R$)", round(resgates["valor_resgatado_centavos"].sum() / 100, 2)],
        ["Soma ganhos switchings (R$)", round(switchings_prop["ganho_vs_base_R$"].sum(), 2)],
        ["Soma agrupada de switchings (R$)", round(sw_group["Valor_Total_R$"].sum(), 2)],
        ["Diferença agrupado vs proposto (R$)", round(sw_group["Valor_Total_R$"].sum() - switchings_prop["Valor_Switching_R$"].sum(), 2)],
        ["Saldo líquido final estimado Carteira_Final (R$)", round(carteira_final["Saldo_Liquido_Final_Estimado_R$"].sum(), 2)],
        ["Riqueza terminal oficial v2 (R$)", 10514.93],
        ["Observação", "As classes de lotes usam o Inventário de Lotes como fonte primária. A planilha de referência entra apenas como apoio para saldos."],
    ], columns=["Conferencia", "Valor"])

    extrato_hist = ref_extrato.copy()
    extrato_hist["Tipo Evento"] = "PAGAMENTO_REALIZADO"
    extrato_hist["Carteira Origem"] = extrato_hist["Lote"].map(lote_to_carteira)
    extrato_hist["Carteira Destino"] = None
    extrato_hist["Motivo"] = "Histórico já realizado na base."
    extrato_hist["Status"] = "REALIZADO"

    future_events = []
    for _, row in switchings_prop.iterrows():
        future_events.append({
            "Data": row["Data"],
            "Ordem Evento": 1,
            "Tipo Evento": "SWITCHAR",
            "Conta": "Cobertura do déficit do dia",
            "Lote": row["Lote_Origem"],
            "Carteira Origem": row["Carteira_Origem"],
            "Carteira Destino": row["Carteira_Destino"],
            "Saldo Antes": None,
            "Bruto": row["Valor_Switching_R$"],
            "Imposto": None,
            "Liquido": row["Valor_Switching_R$"],
            "Dias Corridos": None,
            "Dias Úteis": None,
            "Saldo Remanescente": None,
            "Motivo": row["Motivo"],
            "Status": "PLANEJADO",
        })
    for _, row in resgates.iterrows():
        future_events.append({
            "Data": pd.to_datetime(row["data"]),
            "Ordem Evento": int(row["ordem_resgate_no_dia"]) + 10,
            "Tipo Evento": "RESGATAR",
            "Conta": "Cobertura do déficit do dia",
            "Lote": row["id_lote"],
            "Carteira Origem": row["carteira"],
            "Carteira Destino": None,
            "Saldo Antes": row["valor_liquido_antes_centavos"] / 100,
            "Bruto": (row["valor_bruto_antes_centavos"] - row["valor_bruto_depois_centavos"]) / 100,
            "Imposto": None,
            "Liquido": row["valor_resgatado_centavos"] / 100,
            "Dias Corridos": None,
            "Dias Úteis": None,
            "Saldo Remanescente": row["valor_liquido_depois_centavos"] / 100,
            "Motivo": "Resgate oficial confirmado pela baseline v2.",
            "Status": "PLANEJADO",
        })
    extrato_future = pd.DataFrame(future_events)
    extrato_frames = [
        extrato_hist[["Data", "Ordem Evento", "Tipo Evento", "Conta", "Lote", "Carteira Origem", "Carteira Destino",
                      "Saldo Antes", "Bruto", "Imposto", "Liquido", "Dias Corridos", "Dias Úteis",
                      "Saldo Remanescente", "Motivo", "Status"]]
    ]
    if not extrato_future.empty:
        extrato_frames.append(
            extrato_future[["Data", "Ordem Evento", "Tipo Evento", "Conta", "Lote", "Carteira Origem", "Carteira Destino",
                            "Saldo Antes", "Bruto", "Imposto", "Liquido", "Dias Corridos", "Dias Úteis",
                            "Saldo Remanescente", "Motivo", "Status"]]
        )
    extrato = pd.concat(extrato_frames, ignore_index=True).sort_values(["Data", "Ordem Evento", "Tipo Evento"]).reset_index(drop=True)

    wb = Workbook()
    wb.remove(wb.active)
    sheets = [
        ("Resumo", resumo, "Resumo executivo do plano oficial v2", "Visão curta para validar o que já foi feito e o que deve ser feito."),
        ("Extrato", extrato, "Extrato operacional", "Histórico realizado + eventos futuros oficiais (switchings e resgates)."),
        ("Investidos_Atuais", lotes_views["INVESTIDOS_ATUAIS"], "Lotes investidos atuais", "Fonte primária: Inventário de Lotes. Data de referência: hoje em Brasília."),
        ("Livres_Disponiveis", lotes_views["LIVRES_DISPONIVEIS"], "Lotes livres disponíveis", "Lotes sem carteira, já disponíveis na data de referência."),
        ("Livres_Futuros", lotes_views["LIVRES_FUTUROS"], "Lotes livres futuros", "Lotes sem carteira, ainda futuros em relação à data de referência."),
        ("Bloqueados_Ja_Gastos", lotes_views["BLOQUEADOS_JA_GASTOS"], "Lotes bloqueados / já gastos", "Lotes com Investimento == '-' . Não disponíveis para novas decisões."),
        ("Plano_Futuro", plano_futuro, "Plano futuro de pagamentos", "Lista as contas futuras e a ação oficial associada ao dia."),
        ("Switchings_Propostos", switchings_prop, "Switchings oficiais propostos", "Switchings realmente escolhidos pela baseline oficial v2."),
        ("Switchings_Agrupados", sw_group, "Switchings agrupados", "Agrupa repetições da mesma origem/destino para facilitar validação operacional."),
        ("Justificativa_Switching", justificativa, "Justificativa resumida do switching", "Explica por que a carteira foi escolhida e por que o switching foi fragmentado em mais de um dia."),
        ("Carteira_Final", carteira_final, "Carteira final operacional", "Visão operacional simplificada do saldo por lote após os eventos oficiais."),
        ("Conferencia", conferencia, "Conferência geral", "Checagens simples para validar a consistência do workbook principal."),
    ]
    for name, df, title, notes in sheets:
        ws = wb.create_sheet(title=name)
        _write_df(ws, df, title, notes)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
