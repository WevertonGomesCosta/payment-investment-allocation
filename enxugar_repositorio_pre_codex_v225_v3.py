from __future__ import annotations

"""
V225 — Enxugamento final pré-Codex v3, idempotente e tolerante a PermissionError.

Use este script depois da falha por WinError 5 no OneDrive.

Ele:
1. Continua a limpeza mesmo que a execução v2 tenha parado no meio.
2. Garante que helpers locais do console não usem nomes legados.
3. Remove, se ainda existirem:
   - aplicacao/console/secoes_financeiras.py
   - aplicacao/console/secoes_canonicas.py
4. Tenta remover diretórios auxiliares; se o OneDrive/Windows bloquear, registra aviso e continua.
5. Atualiza:
   - AGENTS.md
   - scripts/validacao/validar_rota_oficial_v225.py
   - relatorios/atuais/codex_ready/CODEX_READY_V225.md
   - relatorios/atuais/codex_ready/INVENTARIO_LEGADO_INATIVO_V225.md
   - relatorios/atuais/codex_ready/ENXUGAMENTO_FINAL_REPOSITORIO_PRE_CODEX_V225.md

Execute na raiz do repositório:
    python enxugar_repositorio_pre_codex_v225_v3.py

Depois valide:
    python scripts/validacao/validar_rota_oficial_v225.py
    python aplicacao/principal.py
"""

from pathlib import Path
from datetime import datetime
import ast
import os
import shutil
import stat
import sys


REPO = Path(".").resolve()

ARQ_AGENTS = REPO / "AGENTS.md"
ARQ_PRINCIPAL = REPO / "aplicacao" / "principal.py"
ARQ_CONSOLE = REPO / "aplicacao" / "console" / "principal.py"
ARQ_PLANILHA = REPO / "nucleo" / "gerar_planilha_operacional.py"
ARQ_OBSERVAVEL = REPO / "nucleo" / "saida_observavel.py"
ARQ_CANONICA = REPO / "nucleo" / "saida_canonica.py"
ARQ_CONFIG = REPO / "dados" / "config_atualizado.json"
ARQ_VALIDACAO = REPO / "scripts" / "validacao" / "validar_rota_oficial_v225.py"

ARQ_SECOES_FINANCEIRAS = REPO / "aplicacao" / "console" / "secoes_financeiras.py"
ARQ_SECOES_CANONICAS = REPO / "aplicacao" / "console" / "secoes_canonicas.py"

DIR_CODEX = REPO / "relatorios" / "atuais" / "codex_ready"
DIR_BACKUPS_LEGADO = DIR_CODEX / "legado_preservado"
DIR_SCRIPTS_TEMP = DIR_CODEX / "scripts_temporarios_removidos"

ARQ_CODEX = DIR_CODEX / "CODEX_READY_V225.md"
ARQ_INVENTARIO = DIR_CODEX / "INVENTARIO_LEGADO_INATIVO_V225.md"
ARQ_RELATORIO = DIR_CODEX / "ENXUGAMENTO_FINAL_REPOSITORIO_PRE_CODEX_V225.md"

TEMP_SCRIPTS_RAIZ = sorted(set([
    "aplicar_v225_contexto_unico.py",
    "auditar_secoes_financeiras_v225.py",
    "auditar_secoes_financeiras_v225_v2.py",
    "remover_legados_secoes_financeiras_lote01_v225.py",
    "migrar_amostras_pagamentos_saida_observavel_v225.py",
    "preparar_codex_ready_v225.py",
    "corrigir_migracao_amostras_codex_ready_v225.py",
    "forcar_migracao_amostras_observavel_v225.py",
    "concluir_codex_ready_amostras_v225.py",
    "concluir_codex_ready_amostras_v225_corrigido.py",
    "corrigir_estado_minimo_codex_ready_v225.py",
    "corrigir_classificacao_rota_codex_ready_v225.py",
    "renomear_helper_amostras_operacionais_v225.py",
    "limpeza_final_pre_codex_v225.py",
    "fechar_inconsistencias_legado_pre_codex_v225.py",
    "enxugar_repositorio_pre_codex_v225.py",
    "enxugar_repositorio_pre_codex_v225_v2.py",
]))


def fail(msg: str) -> None:
    print(f"ERRO: {msg}", file=sys.stderr)
    raise SystemExit(1)


def read_optional(path: Path) -> str:
    if not path.exists():
        return ""
    return path.read_text(encoding="utf-8")


def read_required(path: Path) -> str:
    if not path.exists():
        fail(f"Arquivo obrigatório ausente: {path}")
    return path.read_text(encoding="utf-8")


def write(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def check_py_text(text: str, label: str) -> None:
    try:
        ast.parse(text)
    except SyntaxError as exc:
        fail(f"Erro de sintaxe em {label}: {exc}")


def check_py_file(path: Path) -> None:
    check_py_text(read_required(path), str(path))


def status(cond: bool) -> str:
    return "SIM" if cond else "NÃO"


def rel(path: Path) -> str:
    return str(path.relative_to(REPO)).replace("\\", "/")


def renomear_helpers_locais_do_console() -> list[str]:
    texto = read_required(ARQ_CONSOLE)
    original = texto
    alteracoes: list[str] = []

    substituicoes = [
        (
            "def _render_secao_situacao_atual(contexto_baseline, saida_canonica, resumo_fechamento, resumo_recebidos) -> None:",
            "def _render_situacao_atual_operacional(contexto_baseline, saida_canonica, resumo_fechamento, resumo_recebidos) -> None:",
            "_render_secao_situacao_atual(",
            "_render_situacao_atual_operacional(",
            "_render_secao_situacao_atual -> _render_situacao_atual_operacional",
        ),
        (
            "def _render_secao_amostras_pagamentos(saida_canonica) -> None:",
            "def _render_amostras_pagamentos_operacionais(saida_canonica) -> None:",
            "_render_secao_amostras_pagamentos(",
            "_render_amostras_pagamentos_operacionais(",
            "_render_secao_amostras_pagamentos -> _render_amostras_pagamentos_operacionais",
        ),
    ]

    for def_antiga, def_nova, chamada_antiga, chamada_nova, rotulo in substituicoes:
        if def_antiga in texto or chamada_antiga in texto:
            texto = texto.replace(def_antiga, def_nova)
            texto = texto.replace(chamada_antiga, chamada_nova)
            alteracoes.append(rotulo)

    if texto != original:
        proibidos = [
            "render_secao_situacao_atual",
            "render_secao_amostras_pagamentos",
            "render_secao_canonicas",
            "aplicacao.console.secoes_financeiras",
            "aplicacao.console.secoes_canonicas",
        ]
        restantes = [p for p in proibidos if p in texto]
        if restantes:
            fail("Ainda restam strings legadas no console: " + ", ".join(restantes))

        check_py_text(texto, str(ARQ_CONSOLE))
        write(ARQ_CONSOLE, texto)

    return alteracoes


def remover_arquivo(path: Path) -> tuple[bool, str]:
    if not path.exists():
        return False, f"{rel(path)}: já ausente"

    try:
        path.unlink()
        return True, f"{rel(path)}: removido"
    except PermissionError as exc:
        return False, f"{rel(path)}: NÃO removido por PermissionError: {exc}"


def chmod_retry(func, path, exc_info):
    """Callback para shutil.rmtree em Windows."""
    try:
        os.chmod(path, stat.S_IWRITE)
        func(path)
    except Exception:
        raise


def remover_diretorio_tolerante(path: Path) -> tuple[bool, str]:
    if not path.exists():
        return False, f"{rel(path)}: já ausente"

    try:
        shutil.rmtree(path, onerror=chmod_retry)
        return True, f"{rel(path)}/: removido"
    except PermissionError as exc:
        return False, f"{rel(path)}/: NÃO removido por PermissionError; feche Explorer/Excel/OneDrive e remova manualmente se necessário. Detalhe: {exc}"
    except OSError as exc:
        return False, f"{rel(path)}/: NÃO removido por OSError; pode estar travado pelo OneDrive. Detalhe: {exc}"


def remover_legados_e_auxiliares() -> tuple[list[str], list[str], list[str]]:
    removidos_stubs: list[str] = []
    removidos_aux: list[str] = []
    avisos: list[str] = []

    for path in [ARQ_SECOES_FINANCEIRAS, ARQ_SECOES_CANONICAS]:
        ok, msg = remover_arquivo(path)
        if ok:
            removidos_stubs.append(rel(path))
        else:
            avisos.append(msg)

    for path in [DIR_BACKUPS_LEGADO, DIR_SCRIPTS_TEMP]:
        ok, msg = remover_diretorio_tolerante(path)
        if ok:
            removidos_aux.append(msg)
        elif "já ausente" not in msg:
            avisos.append(msg)

    for nome in TEMP_SCRIPTS_RAIZ:
        path = REPO / nome
        if path.exists() and path.is_file():
            ok, msg = remover_arquivo(path)
            if ok:
                removidos_aux.append(nome)
            else:
                avisos.append(msg)

    return removidos_stubs, removidos_aux, avisos


def auditar_estado() -> dict:
    principal = read_optional(ARQ_PRINCIPAL)
    console = read_optional(ARQ_CONSOLE)
    planilha = read_optional(ARQ_PLANILHA)
    observavel = read_optional(ARQ_OBSERVAVEL)

    contexto_unico = all([
        ARQ_PRINCIPAL.exists(),
        "from aplicacao.console.principal import render_console" in principal,
        "from nucleo.gerar_planilha_operacional import main as gerar_planilha_operacional" in principal,
        "def carregar_contexto_e_saida(" in principal,
        "construir_saida_canonica(contexto_baseline" in principal,
        "render_console(contexto_baseline, saida_canonica)" in principal,
        "contexto=contexto_baseline" in principal,
        "saida=saida_canonica" in principal,
    ])

    saida_observavel = all([
        "from nucleo.saida_observavel import" in console,
        "construir_amostras_pagamentos_operacionais" in console,
        "construir_linhas_lotes_id_curta" in console,
        "construir_resumo_patrimonio_total_lotes" in console,
        "construir_blocos_situacao_atual" in planilha,
        "def construir_amostras_pagamentos_operacionais(" in observavel,
        "def construir_blocos_situacao_atual(" in observavel,
    ])

    console_sem_financeiras = all([
        "aplicacao.console.secoes_financeiras" not in console,
        "render_secao_amostras_pagamentos" not in console,
        "render_secao_situacao_atual" not in console,
    ])

    console_sem_canonicas = all([
        "aplicacao.console.secoes_canonicas" not in console,
        "render_secao_canonicas" not in console,
    ])

    legados_removidos = all([
        not ARQ_SECOES_FINANCEIRAS.exists(),
        not ARQ_SECOES_CANONICAS.exists(),
    ])

    planilha_sem_validacao = all([
        "Validacao" not in planilha,
        "'Validacao'" not in planilha,
        '"Validacao"' not in planilha,
    ])

    estado_minimo = all([
        ARQ_PRINCIPAL.exists(),
        ARQ_CONFIG.exists(),
        contexto_unico,
        saida_observavel,
        console_sem_financeiras,
        console_sem_canonicas,
        legados_removidos,
        planilha_sem_validacao,
    ])

    return {
        "principal_existe": ARQ_PRINCIPAL.exists(),
        "config_existe": ARQ_CONFIG.exists(),
        "contexto_unico": contexto_unico,
        "saida_observavel": saida_observavel,
        "console_sem_financeiras": console_sem_financeiras,
        "console_sem_canonicas": console_sem_canonicas,
        "legados_removidos": legados_removidos,
        "planilha_sem_validacao": planilha_sem_validacao,
        "estado_minimo": estado_minimo,
    }


def criar_agents(a: dict) -> str:
    return f"""# AGENTS.md — payment-investment-allocation

## Objetivo

Este repositório está preparado para uso com agentes de código, incluindo Codex. Este arquivo define a rota oficial, os comandos de validação e as restrições operacionais que devem ser respeitadas.

## Baseline vigente

- Baseline operacional: V225
- Entrada oficial: `aplicacao/principal.py`
- Configuração canônica: `dados/config_atualizado.json`
- Saída operacional oficial: `saidas/oficial/relatorio_operacional_v225.xlsx`

## Comando oficial de execução

```bash
python aplicacao/principal.py
```

## Comando oficial de validação

```bash
python scripts/validacao/validar_rota_oficial_v225.py
```

## Arquitetura operacional obrigatória

```text
aplicacao/principal.py
├── carregar_contexto_baseline(...) uma única vez
├── construir_saida_canonica(...) uma única vez
├── render_console(contexto_baseline, saida_canonica)
└── gerar_planilha_operacional(contexto=contexto_baseline, saida=saida_canonica)
```

Estado auditado:

- contexto único: {status(a["contexto_unico"])}
- saída observável única: {status(a["saida_observavel"])}
- console sem dependência operacional de `secoes_financeiras.py`: {status(a["console_sem_financeiras"])}
- console sem dependência operacional de `secoes_canonicas.py`: {status(a["console_sem_canonicas"])}
- arquivos legados de console removidos: {status(a["legados_removidos"])}

## Fontes únicas

- `nucleo/saida_canonica.py`: saída canônica estruturada.
- `nucleo/saida_observavel.py`: fonte única para dados observáveis compartilhados entre console e planilha.
- `aplicacao/console/principal.py`: renderizador do console.
- `nucleo/gerar_planilha_operacional.py`: renderizador da planilha.

Alterações em dados observáveis compartilhados devem ser feitas primeiro em `nucleo/saida_observavel.py`.

## Restrições fortes

Não alterar sem solicitação explícita:

- motor econômico;
- replay;
- regra de pagamentos;
- switching;
- ranking;
- cache CDI/BCB;
- identidade da baseline V225;
- contratos matemáticos/econômicos;
- estrutura de leitura das abas de entrada;
- `dados/config_atualizado.json`.

## Abas de entrada autorizadas

A execução deve ler somente:

- `Carteira`;
- `Todos os Gastos`;
- `Inventário de Lotes`.

Qualquer estrutura derivada deve ser criada internamente pelo script.

## Arquivos legados de console

Os módulos antigos `aplicacao/console/secoes_financeiras.py` e `aplicacao/console/secoes_canonicas.py` foram removidos do repositório para evitar reuso acidental.

Não recriar renderizadores paralelos. Se alguma saída antiga precisar voltar ao console ou à planilha, migrar primeiro o contrato de dados para `nucleo/saida_observavel.py`.

## Proibições operacionais para agentes

- Não criar nova rota principal paralela.
- Não fazer console e planilha recalcularem os mesmos dados por funções diferentes.
- Não reabrir validações antigas já encerradas sem evidência concreta.
- Não recriar arquivos legados de apresentação removidos.
- Não alterar `dados/config_atualizado.json` sem necessidade contratual explícita.
- Não versionar `__pycache__`, `.pyc`, logs temporários ou artefatos locais não oficiais.

## Antes de propor commit

```bash
python scripts/validacao/validar_rota_oficial_v225.py
git status
```
"""


def criar_validacao() -> str:
    return """from __future__ import annotations

from pathlib import Path
import subprocess
import sys

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


REPO = Path(__file__).resolve().parents[2]

ARQUIVOS_PY_COMPILE = [
    "aplicacao/principal.py",
    "aplicacao/console/principal.py",
    "nucleo/saida_observavel.py",
    "nucleo/gerar_planilha_operacional.py",
]

ARQUIVOS_REMOVIDOS_ESPERADOS = [
    "aplicacao/console/secoes_financeiras.py",
    "aplicacao/console/secoes_canonicas.py",
]

ARQ_CONSOLE = REPO / "aplicacao" / "console" / "principal.py"
ARQ_PLANILHA = REPO / "nucleo" / "gerar_planilha_operacional.py"
ARQ_OBSERVAVEL = REPO / "nucleo" / "saida_observavel.py"
ARQ_CODEX = REPO / "relatorios" / "atuais" / "codex_ready" / "CODEX_READY_V225.md"
ARQ_XLSX = REPO / "saidas" / "oficial" / "relatorio_operacional_v225.xlsx"


def run(cmd: list[str]) -> None:
    print("$ " + " ".join(cmd))
    proc = subprocess.run(cmd, cwd=REPO, text=True)
    if proc.returncode != 0:
        raise SystemExit(proc.returncode)


def check(cond: bool, msg: str) -> None:
    if not cond:
        raise AssertionError(msg)


def main() -> None:
    for rel in ARQUIVOS_PY_COMPILE:
        check((REPO / rel).exists(), f"Arquivo ausente: {rel}")

    for rel in ARQUIVOS_REMOVIDOS_ESPERADOS:
        check(not (REPO / rel).exists(), f"Arquivo legado deveria ter sido removido: {rel}")

    run([sys.executable, "-m", "py_compile", *ARQUIVOS_PY_COMPILE])

    console = ARQ_CONSOLE.read_text(encoding="utf-8")
    planilha = ARQ_PLANILHA.read_text(encoding="utf-8")
    observavel = ARQ_OBSERVAVEL.read_text(encoding="utf-8")

    check("aplicacao.console.secoes_financeiras" not in console, "Console ainda importa secoes_financeiras.")
    check("aplicacao.console.secoes_canonicas" not in console, "Console ainda importa secoes_canonicas.")
    check("render_secao_situacao_atual" not in console, "Console ainda contém nome legado de Situação Atual.")
    check("render_secao_amostras_pagamentos" not in console, "Console ainda contém nome legado de amostras.")
    check("render_secao_canonicas" not in console, "Console ainda contém nome legado de canonicas.")

    check("construir_amostras_pagamentos_operacionais" in console, "Console não usa amostras observáveis.")
    check("construir_linhas_lotes_id_curta" in console, "Console não usa Situação Atual observável.")
    check("construir_blocos_situacao_atual" in planilha, "Planilha não usa blocos observáveis da Situação Atual.")
    check("def construir_amostras_pagamentos_operacionais(" in observavel, "saida_observavel não contém amostras de pagamentos.")
    check("def construir_blocos_situacao_atual(" in observavel, "saida_observavel não contém blocos da Situação Atual.")
    check("Validacao" not in planilha, "Planilha ainda contém referência à aba Validacao.")

    run([sys.executable, "aplicacao/principal.py"])

    check(ARQ_XLSX.exists(), f"Saída operacional não encontrada: {ARQ_XLSX}")

    if load_workbook is not None:
        wb = load_workbook(ARQ_XLSX, read_only=True)
        try:
            sheets = set(wb.sheetnames)
            check("Situação Atual" in sheets, "Aba Situação Atual ausente.")
            check("Validacao" not in sheets, "Aba Validacao ainda existe.")
        finally:
            wb.close()

    if ARQ_CODEX.exists():
        texto = ARQ_CODEX.read_text(encoding="utf-8")
        linhas = [linha.strip() for linha in texto.splitlines() if "Estado mínimo Codex-ready" in linha]
        check(linhas, "CODEX_READY_V225.md não contém Estado mínimo Codex-ready.")
        check(
            any("| SIM |" in linha or linha.endswith("| SIM") for linha in linhas),
            "CODEX_READY_V225.md não confirma Estado mínimo Codex-ready = SIM: " + " ; ".join(linhas),
        )

    print("")
    print("VALIDAÇÃO OFICIAL V225 CONCLUÍDA COM SUCESSO")


if __name__ == "__main__":
    main()
"""


def criar_codex(a: dict) -> str:
    return f"""# CODEX-ready V225

## Identificação

- Baseline: V225
- Data/hora local: {datetime.now().isoformat(timespec='seconds')}
- Tipo: enxugamento final do repositório pré-Codex
- Escopo: remoção de stubs legados e arquivos temporários
- Alteração de motor econômico: não
- Alteração de replay: não
- Alteração de pagamentos: não
- Alteração de switching: não
- Alteração de ranking: não
- Alteração de cache: não

## Estado auditado

| Item | Status |
|---|---:|
| `aplicacao/principal.py` existe | {status(a['principal_existe'])} |
| `dados/config_atualizado.json` existe | {status(a['config_existe'])} |
| Entrada oficial carrega contexto único | {status(a['contexto_unico'])} |
| Console e planilha usam dados observáveis centralizados | {status(a['saida_observavel'])} |
| Console sem dependência operacional de `secoes_financeiras.py` | {status(a['console_sem_financeiras'])} |
| Console sem dependência operacional de `secoes_canonicas.py` | {status(a['console_sem_canonicas'])} |
| Arquivos legados de console removidos | {status(a['legados_removidos'])} |
| Planilha não cria aba `Validacao` | {status(a['planilha_sem_validacao'])} |
| Estado mínimo Codex-ready | {status(a['estado_minimo'])} |

## Rota oficial

```text
aplicacao/principal.py
├── carregar_contexto_baseline(...) uma única vez
├── construir_saida_canonica(...) uma única vez
├── render_console(contexto_baseline, saida_canonica)
└── gerar_planilha_operacional(contexto=contexto_baseline, saida=saida_canonica)
```

## Comando oficial de validação

```bash
python scripts/validacao/validar_rota_oficial_v225.py
```

## Saída esperada

```text
saidas/oficial/relatorio_operacional_v225.xlsx
```
"""


def criar_inventario(a: dict, helpers: list[str], removidos_stubs: list[str], removidos_aux: list[str], avisos: list[str]) -> str:
    helpers_texto = "\n".join(helpers) if helpers else "nenhum helper local precisou ser renomeado"
    stubs_texto = "\n".join(removidos_stubs) if removidos_stubs else "nenhum stub encontrado para remoção"
    aux_texto = "\n".join(removidos_aux) if removidos_aux else "nenhum arquivo auxiliar removido"
    avisos_texto = "\n".join(avisos) if avisos else "nenhum aviso"

    return f"""# Inventário de legado inativo — V225

## Identificação

- Baseline: V225
- Data/hora local: {datetime.now().isoformat(timespec='seconds')}
- Escopo: enxugamento final do repositório pré-Codex

## Helpers locais renomeados para evitar falso positivo

```text
{helpers_texto}
```

## Arquivos legados removidos

```text
{stubs_texto}
```

## Arquivos/diretórios auxiliares removidos

```text
{aux_texto}
```

## Avisos de remoção

```text
{avisos_texto}
```

## Estado atual

- entrada oficial: `aplicacao/principal.py`
- console oficial: `aplicacao/console/principal.py`
- planilha oficial: `nucleo/gerar_planilha_operacional.py`
- saída canônica: `nucleo/saida_canonica.py`
- saída observável: `nucleo/saida_observavel.py`
- validação oficial: `scripts/validacao/validar_rota_oficial_v225.py`

## Regra para Codex

Qualquer alteração que afete dados mostrados simultaneamente no console e na planilha deve seguir esta ordem:

```text
1. alterar ou criar contrato em nucleo/saida_observavel.py
2. renderizar no console sem recalcular
3. renderizar na planilha sem recalcular
4. validar python scripts/validacao/validar_rota_oficial_v225.py
```
"""


def criar_relatorio(a_depois: dict, helpers: list[str], removidos_stubs: list[str], removidos_aux: list[str], avisos: list[str]) -> str:
    return f"""# Enxugamento final do repositório pré-Codex — V225

## Identificação

- Baseline: V225
- Data/hora local: {datetime.now().isoformat(timespec='seconds')}

## Ações executadas

1. Helpers locais com nomes parecidos com funções legadas renomeados: {len(helpers)}.
2. Arquivos legados reduzidos a stub removidos: {len(removidos_stubs)}.
3. Arquivos/diretórios auxiliares temporários removidos: {len(removidos_aux)}.
4. `AGENTS.md` atualizado para refletir remoção dos legados.
5. `scripts/validacao/validar_rota_oficial_v225.py` atualizado para validar ausência dos legados.
6. `CODEX_READY_V225.md` e `INVENTARIO_LEGADO_INATIVO_V225.md` regenerados.

## Avisos

```text
{chr(10).join(avisos) if avisos else 'nenhum aviso'}
```

## Estado depois

- contexto único: {status(a_depois['contexto_unico'])}
- saída observável: {status(a_depois['saida_observavel'])}
- console sem `secoes_financeiras`: {status(a_depois['console_sem_financeiras'])}
- console sem `secoes_canonicas`: {status(a_depois['console_sem_canonicas'])}
- legados removidos: {status(a_depois['legados_removidos'])}
- estado mínimo: {status(a_depois['estado_minimo'])}

## Restrições respeitadas

Não houve alteração em:

- motor econômico;
- replay;
- pagamentos;
- switching;
- ranking;
- cache;
- `dados/config_atualizado.json`;
- rota principal.

## Validação necessária

```bash
python scripts/validacao/validar_rota_oficial_v225.py
python aplicacao/principal.py
```
"""


def main() -> None:
    obrigatorios = [
        ARQ_PRINCIPAL,
        ARQ_CONSOLE,
        ARQ_PLANILHA,
        ARQ_OBSERVAVEL,
        ARQ_CANONICA,
        ARQ_CONFIG,
        ARQ_VALIDACAO,
    ]

    for path in obrigatorios:
        if not path.exists():
            fail(f"Arquivo obrigatório ausente: {path}")

    DIR_CODEX.mkdir(parents=True, exist_ok=True)

    helpers = renomear_helpers_locais_do_console()
    removidos_stubs, removidos_aux, avisos = remover_legados_e_auxiliares()

    a = auditar_estado()

    write(ARQ_AGENTS, criar_agents(a))
    write(ARQ_VALIDACAO, criar_validacao())
    check_py_file(ARQ_VALIDACAO)

    a = auditar_estado()

    write(ARQ_CODEX, criar_codex(a))
    write(ARQ_INVENTARIO, criar_inventario(a, helpers, removidos_stubs, removidos_aux, avisos))
    write(ARQ_RELATORIO, criar_relatorio(a, helpers, removidos_stubs, removidos_aux, avisos))

    print("ENXUGAMENTO FINAL PRÉ-CODEX V3 CONCLUÍDO")
    print("")
    print(f"- helpers locais renomeados: {len(helpers)}")
    for item in helpers:
        print(f"  - {item}")
    print(f"- stubs removidos: {len(removidos_stubs)}")
    for item in removidos_stubs:
        print(f"  - {item}")
    print(f"- auxiliares removidos: {len(removidos_aux)}")
    for item in removidos_aux:
        print(f"  - {item}")
    if avisos:
        print("- avisos:")
        for item in avisos:
            print(f"  - {item}")
    print(f"- Estado mínimo Codex-ready: {status(a['estado_minimo'])}")
    print("")
    print("Agora rode:")
    print("python scripts/validacao/validar_rota_oficial_v225.py")
    print("python aplicacao/principal.py")


if __name__ == "__main__":
    main()
