from __future__ import annotations

import hashlib
import json
import shutil
import sys
from pathlib import Path

RAIZ_REPOSITORIO = Path(__file__).resolve().parents[1]
if str(RAIZ_REPOSITORIO) not in sys.path:
    sys.path.insert(0, str(RAIZ_REPOSITORIO))

from aplicacao.console.principal import render_console
from nucleo.contexto_operacional_canonico import carregar_contexto_operacional_canonico
from nucleo.integracao_estado_motor_canonico import construir_integracao_estado_motor_canonico
from nucleo.gerar_planilha_operacional import main as gerar_planilha_operacional
from nucleo.estado_temporal_inicial import construir_estado_temporal_inicial
from nucleo.motor_temporal_conjunto import construir_resultado_motor_temporal_conjunto
from nucleo.ledger_temporal_canonico import construir_ledger_temporal_canonico
from nucleo.gates_validacao_nucleo import validar_gates_nucleo
from nucleo.saida_canonica import construir_saida_canonica
from nucleo.saida_canonica_oficial import construir_saida_canonica_oficial
from nucleo.saida_observavel_oficial import construir_pacote_saida_observavel_oficial
from nucleo.paridade_renderizacao_oficial import validar_paridade_renderizacao_oficial
from nucleo.governanca_residuos_pipeline import construir_resultado_governanca_residuos_pipeline
from nucleo.inventario_residuos_pipeline import construir_inventario_residuos_pipeline
from nucleo.identidade_baseline import VERSAO_BASELINE, caminho_artifact, caminho_saida_operacional, metadados_versao_operacional
from nucleo.situacao_atual_oficial import construir_situacao_atual_oficial



def _sha256_arquivo(caminho: Path) -> str:
    h = hashlib.sha256()
    with caminho.open('rb') as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b''):
            h.update(chunk)
    return h.hexdigest()


def _abas_xlsx(caminho_xlsx: Path) -> list[str]:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(caminho_xlsx, read_only=True, data_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()
    except Exception:
        return []


def _status_etapa9(pacote) -> dict:
    return {
        'status': getattr(pacote, 'status', None),
        'preparado': bool(getattr(pacote, 'preparado', False)),
        'ok': bool(getattr(pacote, 'ok', False)),
    }


def _status_objeto(resultado) -> dict:
    return {
        'status': getattr(resultado, 'status', None),
        'ok': bool(getattr(resultado, 'ok', False)),
    }


def _montar_manifest_execucao(
    *,
    pacote_saida_observavel_oficial,
    caminho_saida: Path,
    resultado_paridade_renderizacao,
    resultado_governanca_residuos,
) -> tuple[Path, Path, dict]:
    metadados = dict(getattr(pacote_saida_observavel_oficial, 'metadados', {}) or {})
    nome_manifest = str(metadados.get('manifest_execucao') or 'manifest_execucao.json')
    caminho_manifest_versionado = caminho_saida_operacional(RAIZ_REPOSITORIO, nome_manifest)
    caminho_manifest_estavel = caminho_saida_operacional(RAIZ_REPOSITORIO, 'manifest_execucao.json')
    arquivos_gerados = [
        {
            'tipo': 'xlsx_operacional',
            'caminho': str(caminho_saida),
            'nome': caminho_saida.name,
            'sha256': _sha256_arquivo(caminho_saida),
        },
    ]
    manifest = {
        'pr': metadados.get('pr_artefato'),
        'me': metadados.get('me'),
        'branch': metadados.get('branch'),
        'commit': metadados.get('commit'),
        'commit_curto': metadados.get('commit_curto'),
        'timestamp_execucao_utc': metadados.get('timestamp_execucao_utc'),
        'data_referencia': metadados.get('data_referencia'),
        'arquivos_gerados': arquivos_gerados,
        'abas_xlsx': _abas_xlsx(caminho_saida),
        'etapas': {
            'etapa9': _status_etapa9(pacote_saida_observavel_oficial),
            'etapa10': _status_objeto(resultado_paridade_renderizacao),
            'etapa11': _status_objeto(resultado_governanca_residuos),
        },
    }
    caminho_manifest_versionado.parent.mkdir(parents=True, exist_ok=True)
    caminho_manifest_versionado.write_text(json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + '\n', encoding='utf-8')
    if caminho_manifest_estavel != caminho_manifest_versionado:
        shutil.copyfile(caminho_manifest_versionado, caminho_manifest_estavel)
    artifact = caminho_artifact(caminho_manifest_versionado.name)
    try:
        if artifact.parent.exists():
            shutil.copyfile(caminho_manifest_versionado, artifact)
    except Exception:
        pass
    return caminho_manifest_versionado, caminho_manifest_estavel, manifest


def _render_identidade_artefatos(manifest: dict, caminho_manifest_versionado: Path, caminho_manifest_estavel: Path) -> None:
    print("\n=== IDENTIDADE DOS ARTEFATOS ===")
    print(f"- PR: {manifest.get('pr')}")
    print(f"- ME: {manifest.get('me')}")
    print(f"- branch: {manifest.get('branch')}")
    print(f"- commit curto: {manifest.get('commit_curto')}")
    print(f"- timestamp execução UTC: {manifest.get('timestamp_execucao_utc')}")
    print(f"- data de referência: {manifest.get('data_referencia')}")
    print(f"- manifest versionado: {caminho_manifest_versionado}")
    print(f"- manifest_execucao.json: {caminho_manifest_estavel}")
    print(f"- sha256 manifest: {_sha256_arquivo(caminho_manifest_versionado)}")
    print("- arquivos gerados:")
    for arquivo in list(manifest.get('arquivos_gerados') or []):
        print(f"  - {arquivo.get('tipo')}: {arquivo.get('nome')} | sha256={arquivo.get('sha256')}")
    print(f"- abas XLSX: {', '.join(manifest.get('abas_xlsx') or [])}")
    etapas = manifest.get('etapas') or {}
    for etapa, status in etapas.items():
        print(f"- {etapa}: status={status.get('status')} | ok={status.get('ok')}")

def _valor(objeto, campo, padrao=None):
    if isinstance(objeto, dict):
        return objeto.get(campo, padrao)
    return getattr(objeto, campo, padrao)


def _registrar_metadados_versao_operacional(pacote_saida_observavel_oficial, saida_canonica_oficial) -> None:
    if pacote_saida_observavel_oficial is None:
        return
    data_referencia = getattr(
        saida_canonica_oficial,
        'data_referencia',
        getattr(pacote_saida_observavel_oficial, 'data_referencia', None),
    )
    pacote_saida_observavel_oficial.metadados.update(
        metadados_versao_operacional(
            RAIZ_REPOSITORIO,
            data_referencia=data_referencia,
        )
    )


def _formatar_item_gate(item, indice):
    gate_id = _valor(item, 'gate_id', 'gate_indefinido')
    codigo = _valor(item, 'codigo', 'codigo_indefinido')
    mensagem = _valor(item, 'mensagem', '')
    data_referencia = _valor(item, 'data_referencia')
    entidade_tipo = _valor(item, 'entidade_tipo')
    entidade_id = _valor(item, 'entidade_id')

    partes = [f"{indice}. gate={gate_id}", f"codigo={codigo}"]

    if data_referencia is not None:
        partes.append(f"data={data_referencia}")
    if entidade_tipo is not None:
        partes.append(f"entidade={entidade_tipo}")
    if entidade_id is not None:
        partes.append(f"id={entidade_id}")
    if mensagem:
        partes.append(f"mensagem={mensagem}")

    return " | ".join(partes)


def _render_resumo_gates_bloqueados(resultado_gates_validacao_nucleo, limite_itens=8):
    resumo = _valor(resultado_gates_validacao_nucleo, 'resumo')
    bloqueios = list(_valor(resultado_gates_validacao_nucleo, 'bloqueios', []) or [])
    avisos = list(_valor(resultado_gates_validacao_nucleo, 'avisos', []) or [])

    print("\nResumo dos gates de validação de núcleo:")
    print(f"- gates executados: {_valor(resumo, 'qtd_gates_executados', 'NA')}/{_valor(resumo, 'qtd_gates', 'NA')}")
    print(f"- gates aprovados: {_valor(resumo, 'qtd_gates_aprovados', 'NA')}")
    print(f"- gates reprovados: {_valor(resumo, 'qtd_gates_reprovados', 'NA')}")
    print(f"- bloqueios: {_valor(resumo, 'qtd_bloqueios', len(bloqueios))}")
    print(f"- avisos: {_valor(resumo, 'qtd_avisos', len(avisos))}")
    print(f"- pronto_para_etapa8: {_valor(resumo, 'pronto_para_etapa8', resultado_gates_validacao_nucleo.pronto_para_etapa8)}")

    if bloqueios:
        print("\nPrincipais bloqueios:")
        for indice, bloqueio in enumerate(bloqueios[:limite_itens], start=1):
            print("- " + _formatar_item_gate(bloqueio, indice))
        if len(bloqueios) > limite_itens:
            print(f"- ... {len(bloqueios) - limite_itens} bloqueio(s) adicional(is) omitido(s).")

    if avisos:
        print("\nPrincipais avisos:")
        for indice, aviso in enumerate(avisos[:limite_itens], start=1):
            print("- " + _formatar_item_gate(aviso, indice))
        if len(avisos) > limite_itens:
            print(f"- ... {len(avisos) - limite_itens} aviso(s) adicional(is) omitido(s).")

    print("\nPróxima ação objetiva: corrigir os bloqueios acima antes de esperar console/XLSX oficiais.")


def _render_resultado_paridade_renderizacao(resultado_paridade) -> None:
    if resultado_paridade is None:
        return

    resumo = getattr(resultado_paridade, 'resumo', None)
    auditoria_xlsx = getattr(resultado_paridade, 'auditoria_xlsx', None)
    auditoria_console = getattr(resultado_paridade, 'auditoria_console', None)
    divergencias = list(getattr(resultado_paridade, 'divergencias', []) or [])

    print("\n=== PARIDADE DA RENDERIZAÇÃO OFICIAL — ETAPA 10 ===")
    print(f"- artefato: {getattr(resultado_paridade, 'artefato', None)}")
    print(f"- entrada formal: {getattr(resultado_paridade, 'entrada_formal', None)}")
    print(f"- status: {getattr(resultado_paridade, 'status', None)}")
    print(f"- ok: {getattr(resultado_paridade, 'ok', None)}")
    print(f"- xlsx auditado: {getattr(auditoria_xlsx, 'auditado', None)}")
    print(f"- xlsx status: {getattr(auditoria_xlsx, 'status', None)}")
    print(f"- console auditado: {getattr(auditoria_console, 'auditado', None)}")
    print(f"- console status: {getattr(auditoria_console, 'status', None)}")
    print(f"- divergências: {getattr(resumo, 'qtd_divergencias', len(divergencias))}")
    print(f"- divergências materiais: {getattr(resumo, 'qtd_divergencias_materiais', None)}")
    print(f"- ressalvas: {getattr(resumo, 'qtd_ressalvas', None)}")

    if divergencias:
        print("- primeiras divergências/ressalvas:")
        for divergencia in divergencias[:5]:
            categoria = getattr(divergencia, 'categoria', None)
            alvo = getattr(divergencia, 'alvo', None)
            material = getattr(divergencia, 'material', None)
            mensagem = getattr(divergencia, 'mensagem', None)
            print(f"  - categoria={categoria} | alvo={alvo} | material={material} | mensagem={mensagem}")
        if len(divergencias) > 5:
            print(f"  - ... {len(divergencias) - 5} divergência(s)/ressalva(s) adicional(is) omitida(s).")


def _itens_limpeza_por_classificacao(resultado_limpeza, *classificacoes: str):
    return [
        item
        for item in list(getattr(resultado_limpeza, 'artefatos_avaliados', []) or [])
        if getattr(item, 'classificacao', None) in classificacoes
    ]


def _render_itens_limpeza(titulo: str, itens: list, limite: int = 6) -> None:
    print(f"- {titulo}: {len(itens)}")
    for item in itens[:limite]:
        referencias = getattr(item, 'referencias', {}) or {}
        arquivo = referencias.get('arquivo') or 'arquivo_nao_informado'
        simbolo = referencias.get('simbolo_funcao_classe') or referencias.get('simbolo') or 'simbolo_nao_informado'
        decisao = referencias.get('decisao_recomendada') or getattr(item, 'motivo', '')
        print(f"  - {getattr(item, 'identificador', None)} | {arquivo} | {simbolo} | {decisao}")
    if len(itens) > limite:
        print(f"  - ... {len(itens) - limite} item(ns) adicional(is) omitido(s).")


def _render_resultado_governanca_residuos(resultado_limpeza) -> None:
    if resultado_limpeza is None:
        return

    resumo = getattr(resultado_limpeza, 'resumo', None)
    auditoria = getattr(resultado_limpeza, 'auditoria', None)
    oficiais = _itens_limpeza_por_classificacao(resultado_limpeza, 'rota_oficial_preservada')
    candidatos = _itens_limpeza_por_classificacao(resultado_limpeza, 'residuo_candidato_tratamento')
    bloqueados = list(getattr(resultado_limpeza, 'residuos_bloqueados_tratamento', []) or [])
    historicos_diagnosticos = _itens_limpeza_por_classificacao(
        resultado_limpeza,
        'referencia_historica_preservada',
        'diagnostico_preservado_fora_pipeline',
    )
    fallbacks = _itens_limpeza_por_classificacao(resultado_limpeza, 'rota_alternativa_temporaria_bloqueada_tratamento')

    print("\n=== GOVERNANÇA DE RESÍDUOS DO PIPELINE — ETAPA 11 ===")
    print(f"- artefato: {getattr(resultado_limpeza, 'artefato', None)}")
    print(f"- entrada formal: {getattr(resultado_limpeza, 'entrada_formal', None)}")
    print(f"- origem formal: {getattr(resultado_limpeza, 'origem_formal', None)}")
    print(f"- status: {getattr(resultado_limpeza, 'status', None)}")
    print(f"- ok: {getattr(resultado_limpeza, 'ok', None)}")
    print(f"- inventario_residuos_auxiliar_fornecido: {getattr(auditoria, 'inventario_residuos_auxiliar_fornecido', None)}")
    print(f"- artefatos avaliados: {getattr(resumo, 'qtd_artefatos_avaliados', None)}")
    print(f"- rotas oficiais preservadas: {getattr(resumo, 'qtd_rotas_oficiais_preservadas', len(oficiais))}")
    print(f"- resíduos candidatos a tratamento: {getattr(resumo, 'qtd_residuos_candidatos_tratamento', None)}")
    print(f"- resíduos bloqueados para tratamento: {getattr(resumo, 'qtd_residuos_bloqueados', None)}")
    print(f"- históricos/diagnósticos preservados: {getattr(resumo, 'qtd_historicos_diagnosticos_preservados', len(historicos_diagnosticos))}")
    print(f"- rotas alternativas temporárias bloqueadas para tratamento: {getattr(resumo, 'qtd_fallbacks_temporarios_bloqueados', len(fallbacks))}")
    print(f"- remoção automática autorizada: {getattr(resumo, 'acao_automatica_autorizada', None)}")
    print(
        "- classificação limitada por ausência de inventário: "
        f"{getattr(auditoria, 'classificacao_limitada_por_ausencia_inventario', None)}"
    )

    print("- classificação explícita do inventário:")
    _render_itens_limpeza('rotas oficiais preservadas', oficiais)
    _render_itens_limpeza('resíduos candidatos a tratamento', candidatos)
    _render_itens_limpeza('resíduos bloqueados por dependência ativa', bloqueados)
    _render_itens_limpeza('históricos/diagnósticos preservados', historicos_diagnosticos)
    _render_itens_limpeza('rotas alternativas temporárias bloqueadas para tratamento nesta etapa', fallbacks)

    bloqueios = list(getattr(resultado_limpeza, 'bloqueios_governanca', []) or [])
    if bloqueios:
        print("- bloqueios/ressalvas de limpeza:")
        for bloqueio in bloqueios[:5]:
            print(f"  - {bloqueio}")
        if len(bloqueios) > 5:
            print(f"  - ... {len(bloqueios) - 5} bloqueio(s)/ressalva(s) adicional(is) omitido(s).")


def carregar_contexto_e_saida():
    """Carrega as Etapas 1-8 e só prepara saídas posteriores quando os gates aprovam."""
    contexto_operacional_canonico = carregar_contexto_operacional_canonico(
        raiz_repositorio=RAIZ_REPOSITORIO,
        instalar_automaticamente=False,
    )
    estado_temporal_inicial = construir_estado_temporal_inicial(contexto_operacional_canonico)
    integracao_estado_motor_canonico = construir_integracao_estado_motor_canonico(
        contexto_operacional_canonico,
        estado_temporal_inicial,
        raiz_repositorio=RAIZ_REPOSITORIO,
    )
    _ = integracao_estado_motor_canonico
    resultado_motor_temporal_conjunto = construir_resultado_motor_temporal_conjunto(estado_temporal_inicial)
    ledger_temporal_canonico = construir_ledger_temporal_canonico(resultado_motor_temporal_conjunto)
    resultado_gates_validacao_nucleo = validar_gates_nucleo(ledger_temporal_canonico)

    if not resultado_gates_validacao_nucleo.pronto_para_etapa8:
        return (
            contexto_operacional_canonico,
            estado_temporal_inicial,
            resultado_motor_temporal_conjunto,
            ledger_temporal_canonico,
            resultado_gates_validacao_nucleo,
            None,
            None,
            None,
        )

    saida_canonica = construir_saida_canonica(contexto_operacional_canonico, versao=VERSAO_BASELINE)
    situacao_atual_origem = construir_situacao_atual_oficial(
        contexto_operacional_canonico,
        saida_canonica,
        estado_temporal_inicial=estado_temporal_inicial,
    )

    saida_canonica_oficial = construir_saida_canonica_oficial(
        ledger=ledger_temporal_canonico,
        gates=resultado_gates_validacao_nucleo,
        ranking_carteira=getattr(contexto_operacional_canonico, 'ranking_carteira', None),
        situacao_atual_origem=situacao_atual_origem,
    )
    pacote_saida_observavel_oficial = construir_pacote_saida_observavel_oficial(saida_canonica_oficial)
    _registrar_metadados_versao_operacional(
        pacote_saida_observavel_oficial,
        saida_canonica_oficial,
    )
    return (
        contexto_operacional_canonico,
        estado_temporal_inicial,
        resultado_motor_temporal_conjunto,
        ledger_temporal_canonico,
        resultado_gates_validacao_nucleo,
        saida_canonica,
        saida_canonica_oficial,
        pacote_saida_observavel_oficial,
    )


def main():
    (
        contexto_operacional_canonico,
        estado_temporal_inicial,
        resultado_motor_temporal_conjunto,
        ledger_temporal_canonico,
        resultado_gates_validacao_nucleo,
        saida_canonica,
        saida_canonica_oficial,
        pacote_saida_observavel_oficial,
    ) = carregar_contexto_e_saida()

    _ = resultado_motor_temporal_conjunto
    _ = ledger_temporal_canonico
    _ = saida_canonica_oficial

    if not resultado_gates_validacao_nucleo.pronto_para_etapa8:
        print(
            "Execução bloqueada pelos gates de validação de núcleo: "
            "ResultadoGatesValidacaoNucleo.pronto_para_etapa8=False. "
            "Console e XLSX oficiais não foram gerados."
        )
        _render_resumo_gates_bloqueados(resultado_gates_validacao_nucleo)
        return None

    console_auditavel = render_console(
        contexto_operacional_canonico,
        saida_canonica,
        estado_temporal_inicial=estado_temporal_inicial,
        pacote_saida_observavel_oficial=pacote_saida_observavel_oficial,
    )

    caminho_saida = gerar_planilha_operacional(
        contexto=contexto_operacional_canonico,
        saida=saida_canonica,
        estado_temporal_inicial=estado_temporal_inicial,
        pacote_saida_observavel_oficial=pacote_saida_observavel_oficial,
    )

    print(f"Saída operacional gerada em: {caminho_saida}")

    resultado_paridade_renderizacao = validar_paridade_renderizacao_oficial(
        pacote_saida_observavel=pacote_saida_observavel_oficial,
        caminho_xlsx=caminho_saida,
        console_renderizado=console_auditavel,
    )
    _render_resultado_paridade_renderizacao(resultado_paridade_renderizacao)

    inventario_residuos_pipeline = construir_inventario_residuos_pipeline()
    resultado_governanca_residuos = construir_resultado_governanca_residuos_pipeline(
        resultado_paridade_renderizacao,
        evidencias_auxiliares=inventario_residuos_pipeline,
    )
    _render_resultado_governanca_residuos(resultado_governanca_residuos)

    caminho_manifest_versionado, caminho_manifest_estavel, manifest = _montar_manifest_execucao(
        pacote_saida_observavel_oficial=pacote_saida_observavel_oficial,
        caminho_saida=Path(caminho_saida),
        resultado_paridade_renderizacao=resultado_paridade_renderizacao,
        resultado_governanca_residuos=resultado_governanca_residuos,
    )
    _render_identidade_artefatos(manifest, caminho_manifest_versionado, caminho_manifest_estavel)

    return caminho_saida


if __name__ == "__main__":
    main()
