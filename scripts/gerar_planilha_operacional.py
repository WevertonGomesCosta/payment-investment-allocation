from __future__ import annotations

from pathlib import Path
from datetime import date
from typing import Iterable
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

RAIZ = Path(__file__).resolve().parents[1]
if str(RAIZ) not in sys.path:
    sys.path.insert(0, str(RAIZ))

from nucleo.carregador_config import carregar_config
from nucleo.ambiente import bootstrap_ambiente
from nucleo.calendario_financeiro import construir_calendario_financeiro, contar_dias_rendimento
from nucleo.leitor_planilha import carregar_planilha
from nucleo.carteira_canonica import carregar_carteira_canonica
from nucleo.dados_operacionais_canonicos import carregar_dados_operacionais_canonicos
from nucleo.cache_cdi_bcb import carregar_cache_cdi_diario
from nucleo.triagem_motor import carregar_triagem_motor
from nucleo.nucleo_financeiro_minimo import carregar_nucleo_financeiro_minimo, construir_tabela_iof, construir_faixas_ir
from nucleo.replay_passado_controlado import carregar_replay_passado_controlado


SAIDA_INTERNA = RAIZ / 'saidas' / 'relatorio_operacional_v41.xlsx'
SAIDA_EXTERNA = Path('/mnt/data/payment-investment-allocation_relatorio_operacional_v41.xlsx')


def _limiar(config: dict) -> float:
    auditoria_cfg = config.get('auditoria') if isinstance(config.get('auditoria'), dict) else {}
    replay_cfg = config.get('replay') if isinstance(config.get('replay'), dict) else {}
    valor = auditoria_cfg.get('limiar_residuo_resolvido')
    if valor is None:
        valor = replay_cfg.get('valor_minimo_lote_ativo', 0.01)
    try:
        return float(valor)
    except Exception:
        return 0.01


def _as_rows(iterable: Iterable[dict], columns: list[tuple[str, str]]):
    for item in iterable:
        yield [item.get(src) for src, _ in columns]


def _apply_table_style(ws, headers: list[str], rows: list[list]):
    ws.sheet_view.showGridLines = False
    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(color='FFFFFF', bold=True)
    thin_gray = Side(style='thin', color='D9E1F2')
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(bottom=thin_gray)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(len(rows)+1, 2)}"

    currency_cols = {'Valor', 'Saldo Antes', 'Bruto', 'Imposto', 'Líquido', 'Saldo Remanescente', 'Aplicação Mínima', 'Bruto Atual', 'Líquido Atual', 'Saldo rem', 'Score Final'}
    percent_cols = {'Taxa Base CDI', 'Taxa Bônus CDI'}
    int_cols = {'Dias Corridos', 'Dias Úteis', 'Dias até evento', 'Rank Global', 'Rank Família', 'Dias Bônus', 'Carência Dias'}

    for col_idx, header in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        max_len = len(header)
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
            if header in currency_cols and isinstance(value, (int, float)):
                cell.number_format = 'R$ #,##0.00;[Red](R$ #,##0.00);-'
            elif header in percent_cols and isinstance(value, (int, float)):
                cell.number_format = '0.0%'
            elif header in int_cols and isinstance(value, (int, float)):
                cell.number_format = '0'
            elif 'Data' in header and hasattr(value, 'year'):
                cell.number_format = 'dd/mm/yyyy'
        ws.column_dimensions[letter].width = min(max_len + 2, 32)


def main() -> None:
    cfg = carregar_config(raiz_repositorio=RAIZ)
    ctx = bootstrap_ambiente(cfg.conteudo, grupos_extras=['financeiro'], instalar_automaticamente=False)
    cal = construir_calendario_financeiro(cfg.conteudo, data_referencia=ctx.data_referencia)
    plan = carregar_planilha(cfg.conteudo, raiz_repositorio=cfg.raiz_repositorio)
    cart = carregar_carteira_canonica(plan, cfg.conteudo)
    dados = carregar_dados_operacionais_canonicos(plan, cfg.conteudo, data_referencia=ctx.data_referencia, carteira_canonica=cart)
    cache = carregar_cache_cdi_diario(dados, cfg.conteudo, data_referencia=ctx.data_referencia, raiz_repositorio=cfg.raiz_repositorio)
    tri = carregar_triagem_motor(cart, dados, cal, cfg.conteudo, data_referencia=ctx.data_referencia)
    nuc = carregar_nucleo_financeiro_minimo(dados, cart, cal, cfg.conteudo, data_referencia=ctx.data_referencia, serie_cdi=cache.serie_cdi)
    rep = carregar_replay_passado_controlado(dados, nuc, cal, cfg.conteudo, data_referencia=ctx.data_referencia, serie_cdi=cache.serie_cdi)

    tabela_iof = construir_tabela_iof(cfg.conteudo)
    faixas_ir = construir_faixas_ir(cfg.conteudo)
    limiar = _limiar(cfg.conteudo)

    wb = Workbook()
    ws_passado = wb.active
    ws_passado.title = 'Extrato passado'
    log = rep.log_passado.copy().sort_values(by=['Data', 'Sequencia Saque'], kind='stable')
    cols_passado = [
        ('Data', 'Data'), ('Conta', 'Conta'), ('Despesa ID', 'Despesa ID'), ('Lote', 'Lote'),
        ('Saldo Antes', 'Saldo Antes'), ('Bruto', 'Bruto'), ('Imposto', 'Imposto'), ('Liquido', 'Líquido'),
        ('Dias Corridos', 'Dias Corridos'), ('Dias Úteis', 'Dias Úteis'), ('Saldo Remanescente', 'Saldo Remanescente'),
        ('Fase Operacional Lote', 'Fase')
    ]
    rows_passado = list(_as_rows(log.to_dict('records'), cols_passado))
    _apply_table_style(ws_passado, [dst for _, dst in cols_passado], rows_passado)

    ws_futuro = wb.create_sheet('Extrato futuro')
    gastos_futuros = dados.gastos_canonicos[dados.gastos_canonicos['futuro_ou_pendente_na_data_referencia'] == True].copy().sort_values(by=['data', 'despesa_id'], kind='stable')
    rows_futuro = []
    for item in gastos_futuros.to_dict('records'):
        data_evt = item.get('data')
        dias_ate = max((data_evt - ctx.data_referencia).days, 0) if isinstance(data_evt, date) else None
        rows_futuro.append([
            data_evt,
            item.get('descricao'),
            item.get('despesa_id'),
            item.get('valor'),
            item.get('pago'),
            item.get('lote_usado_1'),
            item.get('lote_usado_2'),
            dias_ate,
            'futuro/pendente',
        ])
    headers_futuro = ['Data', 'Conta', 'Despesa ID', 'Valor', 'Pago', 'Lote usado 1', 'Lote usado 2', 'Dias até evento', 'Status']
    _apply_table_style(ws_futuro, headers_futuro, rows_futuro)

    ws_melhores = wb.create_sheet('Melhores produtos')
    candidatos = tri.quadro_candidatos.copy().sort_values(by=['score_final', 'score_retorno'], ascending=[False, False], kind='stable')
    rows_melhores = []
    for _, row in candidatos.iterrows():
        rows_melhores.append([
            row.get('nome'), row.get('familia_produto'), row.get('regime_taxa'), row.get('taxa_base_cdi'), row.get('taxa_bonus_cdi'),
            row.get('dias_bonus'), row.get('carencia_dias'), row.get('aplicacao_minima'), row.get('score_final'), row.get('rank_global'), row.get('rank_familia')
        ])
    headers_melhores = ['Produto', 'Família', 'Regime', 'Taxa Base CDI', 'Taxa Bônus CDI', 'Dias Bônus', 'Carência Dias', 'Aplicação Mínima', 'Score Final', 'Rank Global', 'Rank Família']
    _apply_table_style(ws_melhores, headers_melhores, rows_melhores)

    ws_atual = wb.create_sheet('Situação atual')
    rows_atual = []
    for lote in sorted(rep.lotes_apos_replay, key=lambda x: (x.data_recebimento, x.data_aplicacao, x.id)):
        saldo_bruto = round(float(lote.valor_bruto_em_data(
            ctx.data_referencia,
            cal,
            serie_cdi=cache.serie_cdi,
            data_base_referencia=ctx.data_referencia,
        ) or 0.0), 2)
        if lote.esgotado or saldo_bruto <= limiar:
            continue
        saldo_liquido = round(float(lote.valor_liquido_em_data(
            ctx.data_referencia,
            cal,
            tabela_iof=tabela_iof,
            faixas_ir=faixas_ir,
            serie_cdi=cache.serie_cdi,
            data_base_referencia=ctx.data_referencia,
        ) or 0.0), 2)
        saldo_rem = round(float(getattr(lote, 'principal_remanescente', 0.0) or 0.0), 2)
        dias_corridos = max((ctx.data_referencia - lote.data_recebimento).days, 0)
        dias_uteis = 0 if ctx.data_referencia < lote.data_aplicacao else contar_dias_rendimento(
            lote.data_base_fiscal,
            ctx.data_referencia,
            cal,
            serie_cdi=cache.serie_cdi,
            data_fechamento_referencia=ctx.data_referencia,
        )
        rows_atual.append([
            lote.id, lote.data_recebimento, lote.data_aplicacao, lote.investimento, round(float(getattr(lote, 'valor_inicial', 0.0) or 0.0), 2), dias_corridos, dias_uteis,
            saldo_bruto, saldo_liquido, saldo_rem
        ])
    headers_atual = ['Lote', 'Recebimento', 'Aplicação', 'Produto', 'Valor Original', 'Dias Corridos', 'Dias Úteis', 'Bruto Atual', 'Líquido Atual', 'Saldo rem']
    _apply_table_style(ws_atual, headers_atual, rows_atual)

    SAIDA_INTERNA.parent.mkdir(parents=True, exist_ok=True)
    wb.save(SAIDA_INTERNA)
    wb.save(SAIDA_EXTERNA)
    print(SAIDA_INTERNA)
    print(SAIDA_EXTERNA)


if __name__ == '__main__':
    main()
