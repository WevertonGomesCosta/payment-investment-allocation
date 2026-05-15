from __future__ import annotations

import math
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parents[2]
DIR_DIAG = BASE_DIR / "saidas/diagnostico"

ARQ_U3_LINHAS = DIR_DIAG / "saida_operacional_pagamentos_v17_f0_u3_linhas.csv"
ARQ_U3_PAGAMENTOS = DIR_DIAG / "saida_operacional_pagamentos_v17_f0_u3_pagamentos.csv"
ARQ_U3_RESUMO = DIR_DIAG / "saida_operacional_pagamentos_v17_f0_u3_resumo.csv"

ARQ_XLSX = DIR_DIAG / "saida_operacional_pagamentos_v17_f0_u4.xlsx"
ARQ_RESUMO_U4 = DIR_DIAG / "exportacao_auxiliar_pagamentos_v17_f0_u4_resumo.csv"
ARQ_LOG = BASE_DIR / "logs/iteracoes/ME-V17-F0-U4_EXPORTACAO_AUXILIAR_PAGAMENTOS.md"

STATUS_GERAL = "exportacao_auxiliar_pagamentos_v17_f0_u4_gerada"


def to_float(x: Any) -> float:
    if pd.isna(x):
        return 0.0
    if isinstance(x, (int, float)):
        f = float(x)
        return 0.0 if math.isnan(f) else f
    s = str(x).strip()
    if not s or s.lower() in {"nan", "none", "n/d", "na", "-"}:
        return 0.0
    s = s.replace("R$", "").replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def round2(x: Any) -> float:
    return round(to_float(x) + 1e-12, 2)


def carregar_csv_obrigatorio(path: Path, nome: str) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(
            f"Arquivo obrigatório ausente para U.4: {path}\n"
            "Execute antes: python -B scripts/diagnostico/integrar_saida_operacional_pagamentos_multifonte_v17_f0_u3.py"
        )
    df = pd.read_csv(path)
    if df.empty:
        raise ValueError(f"Arquivo obrigatório vazio para U.4 ({nome}): {path}")
    return df


def normalizar_valores_numericos(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    campos_numericos = [
        c for c in out.columns
        if any(t in c.lower() for t in [
            "valor", "saldo", "resgate", "diferenca", "soma",
            "qtd", "maior", "pagamento_idx", "ordem"
        ])
    ]
    for c in campos_numericos:
        out[c] = out[c].map(lambda x: round2(x) if pd.notna(x) and str(x).strip() != "" else x)
    return out


def montar_resumo_u4(
    pagamentos: pd.DataFrame,
    linhas: pd.DataFrame,
    multifonte: pd.DataFrame,
    pendencias: pd.DataFrame,
) -> dict[str, Any]:
    resumo = {
        "qtd_pagamentos_u4": int(pagamentos["chave_pagamento"].nunique()),
        "qtd_linhas_operacionais_u4": int(len(linhas)),
        "qtd_linhas_multifonte_u4": int(len(multifonte)),
        "qtd_pagamentos_multifonte_u4": int(multifonte["chave_pagamento"].nunique()),
        "qtd_pagamentos_bloqueados_u4": int((pagamentos["bloqueio_u3"].astype(str) == "sim").sum()),
        "qtd_pagamentos_executaveis_u4": int((pagamentos["executavel_operacionalmente_u3"].astype(str) == "sim").sum()),
        "qtd_candidatos_fifo_diagnosticos_u4": int(
            (linhas["origem_linha_u3"].astype(str) == "candidato_fifo_apenas_diagnostico").sum()
        ),
        "soma_valores_pagamentos_unicos_u4": round2(pagamentos["valor_pagamento"].sum()),
        "soma_resgates_operacionais_u4": round2(linhas["valor_resgate_operacional_u3"].sum()),
        "soma_resgates_multifonte_u4": round2(multifonte["valor_resgate_operacional_u3"].sum()),
        "maior_diferenca_cobertura_multifonte_u4": round2(abs(multifonte["diferenca_cobertura_u3"]).max()),
        "arquivo_xlsx_auxiliar_u4": str(ARQ_XLSX.relative_to(BASE_DIR)),
        "status_geral_u4": STATUS_GERAL,
    }
    return resumo


def validar_criterios(
    pagamentos: pd.DataFrame,
    linhas: pd.DataFrame,
    multifonte: pd.DataFrame,
    pendencias: pd.DataFrame,
    resumo: dict[str, Any],
) -> None:
    assert resumo["qtd_pagamentos_u4"] == 159, f"Esperado 159 pagamentos, obtido {resumo['qtd_pagamentos_u4']}"
    assert resumo["qtd_linhas_operacionais_u4"] == 175, f"Esperado 175 linhas, obtido {resumo['qtd_linhas_operacionais_u4']}"
    assert resumo["qtd_linhas_multifonte_u4"] == 32, f"Esperado 32 linhas multifonte, obtido {resumo['qtd_linhas_multifonte_u4']}"
    assert resumo["qtd_pagamentos_multifonte_u4"] == 16, f"Esperado 16 pagamentos multifonte, obtido {resumo['qtd_pagamentos_multifonte_u4']}"
    assert resumo["qtd_pagamentos_bloqueados_u4"] == 110, f"Esperado 110 bloqueados, obtido {resumo['qtd_pagamentos_bloqueados_u4']}"
    assert len(pendencias) == 110, f"Esperado 110 linhas na aba Pendencias, obtido {len(pendencias)}"
    assert resumo["qtd_candidatos_fifo_diagnosticos_u4"] == 109, (
        f"Esperado 109 FIFO diagnósticos, obtido {resumo['qtd_candidatos_fifo_diagnosticos_u4']}"
    )
    assert resumo["maior_diferenca_cobertura_multifonte_u4"] <= 0.01, (
        f"Diferença multifonte acima de 0.01: {resumo['maior_diferenca_cobertura_multifonte_u4']}"
    )


def montar_metadados() -> pd.DataFrame:
    linhas = [
        ("microetapa", "V17-F0-U.4"),
        ("classe", "DIAGNOSTICO / EXPORTACAO AUXILIAR CONTROLADA / PAGAMENTOS"),
        ("baseline", "main pos-merge da PR #333"),
        ("merge_commit_base", "5a5676d597a2417bd8bd6604a8843675d95b9a09"),
        ("microetapa_anterior", "V17-F0-U.3"),
        ("arquivo_xlsx_auxiliar", str(ARQ_XLSX.relative_to(BASE_DIR))),
        ("fonte_1", str(ARQ_U3_LINHAS.relative_to(BASE_DIR))),
        ("fonte_2", str(ARQ_U3_PAGAMENTOS.relative_to(BASE_DIR))),
        ("fonte_3", str(ARQ_U3_RESUMO.relative_to(BASE_DIR))),
        ("restricao", "XLSX auxiliar diagnostico; nao e saida oficial"),
        ("restricao", "motor economico nao alterado"),
        ("restricao", "recomendador oficial nao alterado"),
        ("restricao", "exportador oficial nao alterado"),
        ("restricao", "XLSX oficial nao alterado"),
        ("restricao", "dados e cache nao alterados"),
        ("restricao", "contrato e modelo oficial nao alterados"),
        ("restricao", "S.7 e T.0-T.8 nao reabertos"),
        ("restricao", "FIFO nao promovido"),
        ("restricao", "110 pagamentos sem lote nao refactibilizados"),
        ("status_geral_u4", STATUS_GERAL),
    ]
    return pd.DataFrame(linhas, columns=["campo", "valor"])


def ajustar_planilha(writer: pd.ExcelWriter, abas: dict[str, pd.DataFrame]) -> None:
    wb = writer.book

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    campos_monetarios = (
        "valor", "saldo", "resgate", "diferenca", "soma", "maior"
    )

    for nome_aba, df in abas.items():
        ws = wb[nome_aba]
        ws.freeze_panes = "A2"

        max_row = ws.max_row
        max_col = ws.max_column
        if max_row >= 1 and max_col >= 1:
            ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = wrap_alignment

        for col_idx, col_name in enumerate(df.columns, start=1):
            letra = get_column_letter(col_idx)
            serie = df[col_name].astype(str).head(500)
            max_len = max([len(str(col_name))] + [len(x) for x in serie.tolist()])
            largura = min(max(max_len + 2, 10), 45)
            if nome_aba == "Linhas_Operacionais":
                largura = min(largura, 38)
            ws.column_dimensions[letra].width = largura

            if any(t in col_name.lower() for t in campos_monetarios):
                for row_idx in range(2, max_row + 1):
                    ws[f"{letra}{row_idx}"].number_format = '#,##0.00'

        for row in ws.iter_rows(min_row=2, max_row=min(max_row, 200), max_col=max_col):
            for cell in row:
                cell.alignment = wrap_alignment


def exportar_xlsx(
    resumo_u4: pd.DataFrame,
    pagamentos: pd.DataFrame,
    linhas: pd.DataFrame,
    multifonte: pd.DataFrame,
    pendencias: pd.DataFrame,
    metadados: pd.DataFrame,
) -> None:
    abas = {
        "Resumo_U4": resumo_u4,
        "Pagamentos": pagamentos,
        "Linhas_Operacionais": linhas,
        "Multifonte": multifonte,
        "Pendencias": pendencias,
        "Metadados": metadados,
    }

    with pd.ExcelWriter(ARQ_XLSX, engine="openpyxl") as writer:
        for nome, df in abas.items():
            df.to_excel(writer, sheet_name=nome, index=False)
        ajustar_planilha(writer, abas)


def main() -> int:
    DIR_DIAG.mkdir(parents=True, exist_ok=True)
    ARQ_LOG.parent.mkdir(parents=True, exist_ok=True)

    linhas = normalizar_valores_numericos(carregar_csv_obrigatorio(ARQ_U3_LINHAS, "U3 linhas"))
    pagamentos = normalizar_valores_numericos(carregar_csv_obrigatorio(ARQ_U3_PAGAMENTOS, "U3 pagamentos"))
    resumo_u3 = carregar_csv_obrigatorio(ARQ_U3_RESUMO, "U3 resumo")

    multifonte = linhas.loc[
        linhas["origem_linha_u3"].astype(str) == "fonte_multifonte_decomposta_u2"
    ].copy()

    pendencias = pagamentos.loc[
        (pagamentos["bloqueio_u3"].astype(str) == "sim")
        | (pagamentos["tipo_pagamento_operacional_u3"].astype(str).str.contains("sem_lote", na=False))
    ].copy()

    resumo = montar_resumo_u4(pagamentos, linhas, multifonte, pendencias)
    validar_criterios(pagamentos, linhas, multifonte, pendencias, resumo)

    resumo_u4 = pd.DataFrame([{"metrica": k, "valor": v} for k, v in resumo.items()])
    metadados = montar_metadados()

    resumo_u4.to_csv(ARQ_RESUMO_U4, index=False)
    exportar_xlsx(resumo_u4, pagamentos, linhas, multifonte, pendencias, metadados)

    if not ARQ_XLSX.exists() or ARQ_XLSX.stat().st_size <= 0:
        raise RuntimeError(f"XLSX auxiliar não foi gerado corretamente: {ARQ_XLSX}")

    fontes_lidas = {
        "U3_LINHAS": linhas.shape,
        "U3_PAGAMENTOS": pagamentos.shape,
        "U3_RESUMO": resumo_u3.shape,
    }

    linhas_fontes = "\n".join(
        f"- `{nome}`: `{shape[0]} x {shape[1] if len(shape) > 1 else 0}`"
        for nome, shape in fontes_lidas.items()
    )
    linhas_resumo = "\n".join(f"- `{k}`: `{v}`" for k, v in resumo.items())

    data_exec = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log = f"""# ME-V17-F0-U4 — Exportação auxiliar controlada da saída operacional de pagamentos

- MICROETAPA: V17-F0-U.4
- CLASSE: DIAGNÓSTICO / EXPORTAÇÃO AUXILIAR CONTROLADA / PAGAMENTOS
- DATA_EXECUCAO_LOCAL: {data_exec}
- BASELINE: main pós-merge da PR #333
- MICROETAPA_ANTERIOR: V17-F0-U.3
- STATUS_GERAL_U4: `{STATUS_GERAL}`

## Objetivo

Gerar um XLSX auxiliar diagnóstico, baseado exclusivamente nos CSVs da U.3, para tornar a saída operacional de pagamentos consumível visualmente.

A U.4 não altera recomendador oficial, motor econômico, exportador oficial, XLSX oficial, dados, cache, contrato, modelo, logs anteriores ou módulos `nucleo/`.

## Fontes lidas

{linhas_fontes}

## Artefatos diagnósticos locais gerados

- `{ARQ_XLSX.relative_to(BASE_DIR)}`
- `{ARQ_RESUMO_U4.relative_to(BASE_DIR)}`

## Abas do XLSX auxiliar

- `Resumo_U4`
- `Pagamentos`
- `Linhas_Operacionais`
- `Multifonte`
- `Pendencias`
- `Metadados`

## Contadores principais

{linhas_resumo}

## Interpretação operacional

A U.4 promove a saída U.3 apenas para um XLSX auxiliar diagnóstico em `saidas/diagnostico/`. A exportação torna a decomposição operacional consumível, mas não integra o conteúdo ao fluxo oficial.

A aba `Pagamentos` preserva 159 pagamentos únicos. A aba `Linhas_Operacionais` preserva 175 linhas. A aba `Multifonte` preserva 32 linhas fonte-a-fonte e 16 pagamentos. A aba `Pendencias` preserva 110 pagamentos bloqueados/pendentes.

## Decisão normativa preservada

- O XLSX gerado é auxiliar e diagnóstico.
- O XLSX oficial não é alterado.
- O exportador oficial não é alterado.
- O motor econômico não é alterado.
- O recomendador oficial não é alterado.
- Os 110 pagamentos sem lote sugerido permanecem bloqueados/pendentes.
- Os 109 candidatos FIFO permanecem diagnósticos.
- Nenhuma fonte FIFO é promovida automaticamente.
- Nenhum recebido é transformado em fonte oficial.
- S.7 e T.0–T.8 não são reabertos.

## Restrições preservadas

- `aplicacao/principal.py` não alterado.
- Motor econômico não alterado.
- Recomendador oficial não alterado.
- Exportador oficial não alterado.
- XLSX oficial não alterado.
- Dados e cache não alterados.
- Contrato e modelo oficial não alterados.
- Logs anteriores não alterados.
- Scripts existentes não alterados.
- Módulos `nucleo/` não alterados.

## Status

`{STATUS_GERAL}`
"""
    ARQ_LOG.write_text(log, encoding="utf-8")

    print("=== U4 GERADA ===")
    for k, v in resumo.items():
        print(f"{k}: {v}")

    print("\nXLSX:")
    print(ARQ_XLSX.relative_to(BASE_DIR))
    print("xlsx_tamanho_bytes:", ARQ_XLSX.stat().st_size)

    print("\nResumo CSV:")
    print(ARQ_RESUMO_U4.relative_to(BASE_DIR))

    print("\nLog:")
    print(ARQ_LOG.relative_to(BASE_DIR))

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
