from __future__ import annotations
import glob
from pathlib import Path
from datetime import datetime, timezone
import pandas as pd
from openpyxl import load_workbook

OUT_DIR = Path('saidas/diagnostico')
OUT_DIR.mkdir(parents=True, exist_ok=True)

BLOCK_STATUS = {
    'sem_saldo_temporal_auditavel','sem_fonte_auditavel','switch_then_pay_sem_materializacao','fonte_pos_switching_nao_materializada'
}

def norm(v): return str(v or '').strip().lower()
def is_nd(v): return norm(v) in {'', 'n/d', 'nd', 'não determinado', 'nao determinado', 'nan', 'none'}
def is_cobertura_integral(v) -> bool:
    tok = norm(v)
    coberto = {'sim', 's', 'true', 'verdadeiro', '1', 'yes'}
    sem_cobertura = {'não', 'nao', 'n', 'false', 'falso', '0', '', 'n/d', 'nd', 'nan', 'none', 'ausente', 'não determinado', 'nao determinado'}
    if tok in coberto:
        return True
    if tok in sem_cobertura:
        return False
    return False

def _inferir_causa(row):
    status = norm(row.get('Status recomendação')); motivo = norm(row.get('Motivo bloqueio lote'))
    reserva = row.get('_reserva_preenchida', False)
    if status == 'switch_then_pay_sem_materializacao' or motivo == 'switch_then_pay_sem_materializacao':
        return ('lote pós-switching não injetado no fluxo','ledger_temporal_conjunto','registrada_pipeline')
    if reserva and is_nd(row.get('Saldo Antes')): return ('reserva não promovida por saldo','promocao_reserva_para_lote','inferida')
    if reserva and row.get('_reserva_futura_sinal', False): return ('reserva não promovida por data','elegibilidade_temporal','inferida')
    if reserva and row.get('_reserva_prazo_sinal', False): return ('reserva não promovida por liquidez/carência','elegibilidade_liquidez_carencia','inferida')
    if reserva and motivo in {'', 'n/d', 'nd', 'não determinado', 'nao determinado'}:
        return ('reserva descartada sem motivo estruturado','promocao_reserva_para_lote','causa_nao_rastreada_no_pipeline')
    if status in BLOCK_STATUS: return ('sem fonte elegível real','motor_recomendacao','registrada_pipeline')
    return ('outro','nao_classificado','causa_nao_rastreada_no_pipeline')

def _non_empty(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    if len(df)==0: return df
    c=[x for x in cols if x in df.columns]
    if not c: return df.iloc[0:0]
    mask=df[c].apply(lambda r: any(not is_nd(v) for v in r), axis=1)
    return df[mask].copy()

def main()->int:
    candidates=sorted(glob.glob('saidas/oficial/*.xlsx'), key=lambda p: Path(p).stat().st_mtime)
    if not candidates:
        print('sem xlsx em saidas/oficial'); return 2
    print('candidatos_xlsx_ordenados_por_mtime:')
    for p in candidates:
        mt=datetime.fromtimestamp(Path(p).stat().st_mtime,tz=timezone.utc).isoformat()
        print(f'- {Path(p).resolve()} | mtime_utc={mt}')
    xlsx=Path(candidates[-1]).resolve()
    mtime=datetime.fromtimestamp(xlsx.stat().st_mtime,tz=timezone.utc).isoformat()

    xl=pd.ExcelFile(xlsx)
    abas=list(xl.sheet_names)
    print(f'xlsx_escolhido={xlsx}')
    print(f'xlsx_mtime_utc={mtime}')
    print(f'abas_disponiveis={abas}')

    extrato=pd.read_excel(xl, sheet_name='Extrato Futuro')
    extrato=extrato[~extrato.apply(lambda r: all(is_nd(v) for v in r), axis=1)].copy()

    aba_shapes=[]
    def ler(nome):
        if nome in abas:
            df=pd.read_excel(xl, sheet_name=nome)
            df=df[~df.apply(lambda r: all(is_nd(v) for v in r), axis=1)].copy()
            if nome == 'Switching' and len(df.columns) == 0:
                wb = load_workbook(filename=xlsx, read_only=True, data_only=True)
                ws = wb[nome]
                headers = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)) if c not in (None, '')]
                wb.close()
                if headers:
                    df = pd.DataFrame(columns=list(headers))
            aba_shapes.append({'aba':nome,'linhas':len(df),'colunas':len(df.columns)})
            return df
        return pd.DataFrame()
    aba_shapes.append({'aba':'Extrato Futuro','linhas':len(extrato),'colunas':len(extrato.columns)})
    estado=ler('Estado Pos-Switching'); sint=ler('Lotes Sinteticos Pos-Sw'); sw=ler('Switching'); sws=ler('Switchings'); aud_fontes=ler('Auditoria Fontes'); aud_fifo=ler('Auditoria FIFO'); aud_fifo_cand=ler('Auditoria FIFO Candidatos'); saida_can=ler('Saida Canonica')


    estruturado = pd.DataFrame()
    base_struct = aud_fontes if len(aud_fontes) else saida_can
    if len(base_struct):
        cand_cols = ['Despesa ID','fonte_candidata_id','tipo_fonte_candidata','origem_fonte_candidata','elegivel_temporalmente','saldo_liquido_disponivel','elegivel_liquidez_carencia','promovida_para_lote_sugerido','etapa_descarte_fonte','motivo_descarte_fonte','origem_motivo_descarte','evento_switching_id']
        has=[c for c in cand_cols if c in base_struct.columns]
        if has and 'Despesa ID' in has:
            estruturado = base_struct[has].copy()
            estruturado = estruturado.rename(columns={'Despesa ID':'_join_despesa_id'})
            extrato['_join_despesa_id']=extrato['Despesa ID'].astype(str)
            extrato = extrato.merge(estruturado, how='left', on='_join_despesa_id')

    extrato['_lote_nd']=extrato['Lote sugerido'].apply(is_nd)
    col_cobertura = 'Cobertura integral' if 'Cobertura integral' in extrato.columns else ('Cobertura Integral?' if 'Cobertura Integral?' in extrato.columns else None)
    extrato['_sem_cobertura_integral'] = (~extrato[col_cobertura].apply(is_cobertura_integral)) if col_cobertura else extrato['_lote_nd']
    extrato['_reserva_preenchida']=~extrato['Lote reserva'].apply(is_nd)
    extrato['_reserva_futura_sinal']=extrato['Lote reserva'].astype(str).str.contains(r'mai\.|jun\.|jul\.|ago\.|set\.|out\.|nov\.|dez\.', case=False, na=False)
    extrato['_reserva_prazo_sinal']=extrato['Lote reserva'].astype(str).str.contains('cdb|lc[ai]|tesouro|prazo|carência|carencia', case=False, na=False)
    extrato['_consumiu_lote_pos_sw']=~extrato.get('Lote pós-switching', pd.Series(['']*len(extrato))).apply(is_nd)

    sem_lote=extrato[extrato['_sem_cobertura_integral']].copy()
    if len(sem_lote):
        causas=sem_lote.apply(_inferir_causa, axis=1, result_type='expand')
        sem_lote[['causa_raiz','etapa_descarte_fonte','tipo_causa']]=causas
    else:
        sem_lote['causa_raiz'] = pd.Series(dtype=str)
        sem_lote['etapa_descarte_fonte'] = pd.Series(dtype=str)
        sem_lote['tipo_causa'] = pd.Series(dtype=str)
    sem_lote['fontes_candidatas_motor']=sem_lote.get('fonte_candidata_id', sem_lote['Lote reserva']).fillna('n/d')
    sem_lote['saldo_liquido_reserva_data']=sem_lote.get('saldo_liquido_disponivel', sem_lote['Saldo Antes']).fillna('n/d')
    sem_lote['elegibilidade_temporal_reserva']=sem_lote['_reserva_futura_sinal'].map({True:'ineligivel_data_inferido',False:'sem_evidencia_ineligibilidade_data'})
    sem_lote['liquidez_carencia_reserva']=sem_lote['_reserva_prazo_sinal'].map({True:'sinal_produto_com_carencia_liquidez',False:'n/d'})
    sem_lote['reserva_descartada']=sem_lote['_reserva_preenchida'].map({True:'sim',False:'não'})
    sem_lote['motivo_descarte_fonte_estruturado']=sem_lote.get('motivo_descarte_fonte', pd.Series(['']*len(sem_lote))).fillna('')
    sem_lote['origem_motivo_descarte_estruturado']=sem_lote.get('origem_motivo_descarte', pd.Series(['']*len(sem_lote))).fillna('')

    sem_lote[['Data','Conta','Valor','Lote reserva','Status recomendação','Motivo bloqueio lote','fontes_candidatas_motor','saldo_liquido_reserva_data','elegibilidade_temporal_reserva','liquidez_carencia_reserva','reserva_descartada','etapa_descarte_fonte','motivo_descarte_fonte_estruturado','origem_motivo_descarte_estruturado','causa_raiz','tipo_causa']].to_csv(OUT_DIR/'diagnostico_baixa_resolutividade_detalhe.csv',index=False)
    causas_agg=sem_lote['causa_raiz'].value_counts().rename_axis('causa_raiz').reset_index(name='qtd')
    causas_agg.to_csv(OUT_DIR/'diagnostico_baixa_resolutividade_causas.csv',index=False)

    materiais=[]
    if len(estado):
        base=_non_empty(estado,['Novo lote','Data','Lotes origem'])
        for _,r in base.iterrows(): materiais.append({'lote':r.get('Novo lote'),'data':r.get('Data'),'origem':r.get('Lotes origem'),'destino':r.get('Produto destino') or r.get('Destino'),'valor':r.get('Valor inicial') or r.get('Valor líquido total')})
    elif len(sint):
        base=_non_empty(sint,['Novo lote','Data','Lotes origem'])
        for _,r in base.iterrows(): materiais.append({'lote':r.get('Novo lote'),'data':r.get('Data'),'origem':r.get('Lotes origem'),'destino':r.get('Destino'),'valor':r.get('Valor líquido total')})
    elif len(sw):
        base=_non_empty(sw,['Novo lote','Lote pós-switching','Data'])
        for _,r in base.iterrows(): materiais.append({'lote':r.get('Novo lote') or r.get('Lote pós-switching'),'data':r.get('Data') or r.get('Data sugerida'),'origem':r.get('Lote origem'),'destino':r.get('Destino') or r.get('Produto destino switching'),'valor':r.get('Valor líquido total') or r.get('Valor líquido origem')})
    elif len(sws):
        base=_non_empty(sws,['Novo lote','Lote pós-switching','Data'])
        for _,r in base.iterrows(): materiais.append({'lote':r.get('Novo lote') or r.get('Lote pós-switching'),'data':r.get('Data') or r.get('Data sugerida'),'origem':r.get('Lote origem'),'destino':r.get('Destino') or r.get('Produto destino switching'),'valor':r.get('Valor líquido total') or r.get('Valor líquido origem')})

    mat=pd.DataFrame(materiais)
    if len(mat):
        mat=mat[~mat['lote'].apply(is_nd)].drop_duplicates(subset=['lote','data'])
    sw_rows=[]
    for _,r in mat.iterrows():
        lote=str(r['lote']); data=str(r['data'] or '')
        eleg=extrato[extrato['Data'].astype(str)>=data] if data else extrato
        entrou=bool((extrato['Lote sugerido'].astype(str).str.strip().str.lower()==lote.strip().lower()).any())
        sw_rows.append({'evento_switching_id':f"sw::{data or 'n/d'}::{r.get('origem') or 'n/d'}::{lote}",'lote_pos_switching':lote,'data_materializacao':data or 'n/d','lotes_origem':r.get('origem') or 'n/d','produto_destino':r.get('destino') or 'n/d','valor_inicial_materializado':r.get('valor') or 'n/d','qtd_pagamentos_elegiveis_apos_materializacao':len(eleg),'entrou_como_fonte_no_extrato_futuro':'sim' if entrou else 'não','motivo_inferido_se_nao_entrou':'fonte disponível não propagada motor → ledger' if not entrou else 'n/d'})
    sw_df=pd.DataFrame(sw_rows)
    sw_df.to_csv(OUT_DIR/'diagnostico_switchings_materializados.csv',index=False)


    diverg_status = 0
    diverg_motivo = 0
    diverg_pacote = 0
    promovida_invalida = 0
    sem_motivo_estruturado = 0
    if len(aud_fontes):
        base = aud_fontes.copy()
        key = 'Despesa ID' if 'Despesa ID' in base.columns else None
        if key is not None:
            ef = extrato[['Despesa ID','Status recomendação','Motivo bloqueio lote','Pacote do dia','Lote sugerido']].copy()
            m = ef.merge(base, on='Despesa ID', how='left', suffixes=('_ef','_af'))
            diverg_status = int((m['Status recomendação'].fillna('').astype(str) != m.get('status_ledger', pd.Series(['']*len(m))).fillna('').astype(str)).sum())
            diverg_motivo = int((m['Motivo bloqueio lote'].fillna('').astype(str) != m.get('motivo_bloqueio_ledger', pd.Series(['']*len(m))).fillna('').astype(str)).sum())
            diverg_pacote = int((m['Pacote do dia'].fillna('').astype(str) != m.get('pacote_do_dia_ledger', pd.Series(['']*len(m))).fillna('').astype(str)).sum())
            lote_col = 'Lote sugerido' if 'Lote sugerido' in m.columns else ('Lote sugerido_ef' if 'Lote sugerido_ef' in m.columns else None)
            lote_nd = m[lote_col].fillna('').astype(str).str.lower().isin(['','n/d','nd','não determinado','nao determinado']) if lote_col else pd.Series([False]*len(m))
            promovida = m.get('promovida_para_lote_sugerido', pd.Series([False]*len(m))).fillna(False).astype(bool)
            promovida_invalida = int((lote_nd & promovida).sum())
            sem_motivo_estruturado = int((lote_nd & m.get('etapa_descarte_fonte', pd.Series(['']*len(m))).fillna('').astype(str).str.strip().eq('')).sum())

    resumo=pd.DataFrame([{
        'xlsx_escolhido':str(xlsx),'xlsx_mtime_utc':mtime,'abas_disponiveis':' | '.join(abas),
        'total_pagamentos_futuros':len(extrato),'total_lote_sugerido_determinado':int((~extrato['_lote_nd']).sum()),'total_lote_sugerido_nao_determinado':int(extrato['_lote_nd'].sum()),'total_sem_cobertura_integral':int(extrato['_sem_cobertura_integral'].sum()),
        'total_reserva_preenchida':int(extrato['_reserva_preenchida'].sum()),'total_reserva_preenchida_e_lote_nd':int((extrato['_reserva_preenchida'] & extrato['_lote_nd']).sum()),
        'total_lotes_pos_switching_materializados':len(sw_df),'total_pagamentos_que_consumiram_lote_pos_switching':int(extrato['_consumiu_lote_pos_sw'].sum()),
        'divergencias_status_extrato_vs_auditoria': diverg_status,
        'divergencias_motivo_extrato_vs_auditoria': diverg_motivo,
        'divergencias_pacote_extrato_vs_auditoria': diverg_pacote,
        'linhas_promovida_invalida_lote_nd': promovida_invalida,
        'linhas_lote_nd_sem_motivo_estruturado': sem_motivo_estruturado
    }])
    resumo.to_csv(OUT_DIR/'diagnostico_baixa_resolutividade_resumo.csv',index=False)
    pd.DataFrame(aba_shapes).to_csv(OUT_DIR/'diagnostico_baixa_resolutividade_abas_shapes.csv', index=False)


    # Transição causal: fonte promovida -> sem saldo temporal auditável
    trans_cols = ['Data','Conta','Despesa ID','Valor','Lote sugerido','Lote reserva','Saldo temp. ant.','Bruto','Líquido','Saldo Remanescente','Status recomendação','Motivo bloqueio lote']
    trans = extrato[[c for c in trans_cols if c in extrato.columns]].copy()
    if len(aud_fontes):
        af_cols = ['Despesa ID','fonte_candidata_id','origem_fonte_candidata','saldo_liquido_disponivel','promovida_para_lote_sugerido','status_ledger','motivo_bloqueio_ledger']
        trans = trans.merge(aud_fontes[[c for c in af_cols if c in aud_fontes.columns]], on='Despesa ID', how='left', suffixes=('_extrato','_aud'))
    trans = trans.rename(columns={
        'Data':'data','Conta':'conta','Despesa ID':'despesa_id','Valor':'valor','Lote sugerido':'lote_sugerido','Lote reserva':'lote_reserva',
        'saldo_liquido_disponivel':'saldo_liquido_disponivel_auditoria','Saldo temp. ant.':'saldo_temporal_antes_extrato','Bruto':'bruto_extrato','Líquido':'liquido_extrato',
        'Saldo Remanescente':'saldo_remanescente_extrato','Status recomendação':'status_extrato','Motivo bloqueio lote':'motivo_extrato',
        'status_ledger_aud':'status_ledger','motivo_bloqueio_ledger':'motivo_ledger'
    })
    def _causa_transicao(r):
        st = norm(r.get('status_ledger')); mt = norm(r.get('motivo_ledger')); prom = bool(r.get('promovida_para_lote_sugerido'))
        if prom and st == 'sem_saldo_temporal_auditavel' and mt == 'saldo_temporal_insuficiente_cumulativo':
            return 'rebaixamento_por_saldo_temporal_cumulativo'
        if prom and st == 'ok':
            return 'sem_transicao_bloqueante'
        return 'outro'
    trans['causa_transicao'] = trans.apply(_causa_transicao, axis=1)
    trans['saldo_temporal_real_antes'] = trans['saldo_temporal_antes_extrato']
    trans['consumo_evento'] = trans['valor']
    trans['saldo_temporal_real_depois'] = trans['saldo_remanescente_extrato']
    trans['saldo_local_ou_valor_promovido'] = trans.get('saldo_liquido_disponivel_auditoria', pd.Series([pd.NA] * len(trans)))
    saldo_local_num = pd.to_numeric(trans['saldo_local_ou_valor_promovido'], errors='coerce')
    saldo_temporal_num = pd.to_numeric(trans['saldo_temporal_real_antes'], errors='coerce')
    trans['diferenca_saldo_local_vs_temporal'] = (saldo_local_num - saldo_temporal_num).round(2)
    trans['interpretacao_saldo_liquido_disponivel'] = trans.apply(
        lambda r: 'saldo_liquido_disponivel aparenta valor local/promovido e não saldo cumulativo real'
        if norm(r.get('status_ledger')) == 'sem_saldo_temporal_auditavel' else 'n/d',
        axis=1,
    )
    trans['interpretacao'] = trans['causa_transicao'].map({
        'rebaixamento_por_saldo_temporal_cumulativo':'saldo_liquido_disponivel_auditoria representa elegibilidade local/promocao; ledger valida saldo cumulativo cronologico e rebaixa quando insuficiente',
        'sem_transicao_bloqueante':'evento permaneceu ok apos validacao cumulativa'
    }).fillna('transicao fora do padrao principal')
    trans.to_csv(OUT_DIR/'auditoria_transicao_fonte_promovida_para_sem_saldo.csv', index=False)
    sem_cobertura = trans[extrato['_sem_cobertura_integral'].values].copy()
    total_promovida_true_e_sem_saldo = int((sem_cobertura.get('promovida_para_lote_sugerido', pd.Series([False]*len(sem_cobertura))).fillna(False).astype(bool) & sem_cobertura['status_ledger'].astype(str).eq('sem_saldo_temporal_auditavel')).sum())
    total_promovida_false_e_sem_saldo = int(((~sem_cobertura.get('promovida_para_lote_sugerido', pd.Series([False]*len(sem_cobertura))).fillna(False).astype(bool)) & sem_cobertura['status_ledger'].astype(str).eq('sem_saldo_temporal_auditavel')).sum())
    print(f"transicao_total_sem_cobertura_integral={len(sem_cobertura)}")
    print(f"transicao_total_promovida_true_e_sem_saldo={total_promovida_true_e_sem_saldo}")
    print(f"transicao_total_promovida_false_e_sem_saldo={total_promovida_false_e_sem_saldo}")
    if 'origem_fonte_candidata' in sem_cobertura.columns:
        print('transicao_total_por_origem_fonte_candidata:')
        print(sem_cobertura['origem_fonte_candidata'].fillna('n/d').astype(str).value_counts().to_string())
    print('transicao_total_por_lote_sugerido:')
    print(sem_cobertura['lote_sugerido'].fillna('n/d').astype(str).value_counts().to_string())
    print('transicao_total_por_causa_transicao:')
    print(sem_cobertura['causa_transicao'].fillna('n/d').astype(str).value_counts().to_string())
    primeira_quebra = sem_cobertura[sem_cobertura['causa_transicao'].eq('rebaixamento_por_saldo_temporal_cumulativo')].sort_values(['data','despesa_id']).head(1)
    primeira_quebra.to_csv(OUT_DIR/'primeira_quebra_causal.csv', index=False)

    print('shapes_lidos:')
    print(pd.DataFrame(aba_shapes).to_string(index=False))
    print(resumo.to_string(index=False))
    qtd_estruturada = int((~sem_lote['motivo_descarte_fonte_estruturado'].astype(str).str.strip().eq('')).sum()) if 'motivo_descarte_fonte_estruturado' in sem_lote.columns else 0
    qtd_nao_rastreada = int((sem_lote['tipo_causa'] == 'causa_nao_rastreada_no_pipeline').sum())
    qtd_inferida = int((sem_lote['tipo_causa'] == 'inferida').sum())

    fifo_avaliados = fifo_resolvidos = 0

    pagamentos_estado_sem_avaliacao_sem_motivo = 0
    if len(aud_fifo):
        motivos = aud_fifo.get('fifo_motivo_nao_promocao', pd.Series(['']*len(aud_fifo))).fillna('').astype(str)
        est = aud_fifo.get('fifo_qtd_lotes_estado', pd.Series([0]*len(aud_fifo))).fillna(0).astype(float)
        av = aud_fifo.get('fifo_qtd_lotes_avaliados', pd.Series([0]*len(aud_fifo))).fillna(0).astype(float)
        motivo_ok = motivos.str.startswith('fifo_nao_aplicavel_') | (~motivos.str.strip().isin(['', 'n/d', 'nd']))
        pagamentos_estado_sem_avaliacao_sem_motivo = int(((est > 0) & (av == 0) & (~motivo_ok)).sum())
    dist_motivos_fifo = pd.Series(dtype='int64')
    estado_vazio = saldo_todos = data_todos = car_todos = mig_todos = 0
    total_candidatos_fifo = 0
    dist_bloq_cand = pd.Series(dtype='int64')
    total_candidatos_elegiveis = 0
    pagamentos_com_candidato_elegivel = 0
    if len(aud_fifo):
        fifo_avaliados = int((aud_fifo.get('fifo_qtd_lotes_avaliados', pd.Series([0]*len(aud_fifo))).fillna(0).astype(float) > 0).sum())
        fifo_resolvidos = int((aud_fifo.get('origem_fonte_candidata', pd.Series(['']*len(aud_fifo))).fillna('').astype(str) == 'pay_only_fifo_v1').sum())
        dist_motivos_fifo = aud_fifo.get('fifo_motivo_nao_promocao', pd.Series([], dtype='object')).fillna('n/d').astype(str).value_counts()
        estado_vazio = int((aud_fifo.get('fifo_qtd_lotes_estado', pd.Series([0]*len(aud_fifo))).fillna(0).astype(float) == 0).sum())
    if len(aud_fifo_cand):
        total_candidatos_fifo = len(aud_fifo_cand)
        total_candidatos_elegiveis = int(aud_fifo_cand.get('elegivel_fifo', pd.Series([False]*len(aud_fifo_cand))).fillna(False).astype(bool).sum())
        pagamentos_com_candidato_elegivel = int(aud_fifo_cand[aud_fifo_cand.get('elegivel_fifo', pd.Series([False]*len(aud_fifo_cand))).fillna(False).astype(bool)]['Despesa ID'].astype(str).nunique()) if 'Despesa ID' in aud_fifo_cand.columns else 0
        motivos = aud_fifo_cand.get('motivo_bloqueio_fifo', pd.Series(['']*len(aud_fifo_cand))).fillna('elegivel').astype(str)
        dist_bloq_cand = motivos.value_counts()
        saldo_todos = int((motivos == 'saldo').sum())
        data_todos = int((motivos == 'data').sum())
        car_todos = int((motivos == 'carencia').sum())
        mig_todos = int((motivos == 'migracao').sum())

    print('causas_raiz_agrupadas:')
    print(causas_agg.to_string(index=False))
    resolvidos_fifo = 0
    if len(aud_fontes) and 'origem_fonte_candidata' in aud_fontes.columns:
        resolvidos_fifo = int((aud_fontes['origem_fonte_candidata'].fillna('').astype(str) == 'pay_only_fifo_v1').sum())
    print(f'causas_estruturadas={qtd_estruturada} | causas_inferidas={qtd_inferida} | causas_nao_rastreadas={qtd_nao_rastreada}')
    print(f'pagamentos_resolvidos_por_pay_only_fifo_v1={resolvidos_fifo}')
    print(f'fifo_total_avaliado={fifo_avaliados} | fifo_total_resolvido={fifo_resolvidos}')
    print(f'fifo_total_candidatos_pagamento_x_lote={total_candidatos_fifo}')
    if len(dist_motivos_fifo):
        print('fifo_distribuicao_motivos_nao_promocao:')
        print(dist_motivos_fifo.to_string())
    print(f'fifo_estado_lotes_vazio={estado_vazio} | fifo_todos_bloq_saldo={saldo_todos} | fifo_todos_bloq_data={data_todos} | fifo_todos_bloq_carencia={car_todos} | fifo_todos_bloq_migracao={mig_todos}')
    if len(dist_bloq_cand):
        print('fifo_distribuicao_bloqueios_por_candidato:')
        print(dist_bloq_cand.to_string())
    print(f'fifo_total_candidatos_elegiveis={total_candidatos_elegiveis} | fifo_pagamentos_com_ao_menos_um_elegivel={pagamentos_com_candidato_elegivel}')
    print(f'pagamentos_estado_lotes_gt0_avaliados_zero_sem_motivo={pagamentos_estado_sem_avaliacao_sem_motivo}')
    print(f'divergencias_status={diverg_status} | divergencias_motivo={diverg_motivo} | divergencias_pacote={diverg_pacote}')
    print(f'linhas_promovida_invalida={promovida_invalida} | linhas_sem_motivo_estruturado={sem_motivo_estruturado}')
    print('lotes_pos_switching_materializados_encontrados:')
    if len(sw_df): print(sw_df[['lote_pos_switching','entrou_como_fonte_no_extrato_futuro']].to_string(index=False))
    else: print('nenhum')
    return 0

if __name__=='__main__': raise SystemExit(main())
