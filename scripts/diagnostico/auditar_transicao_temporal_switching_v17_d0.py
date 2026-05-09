from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
import csv
import re
import sys
import unicodedata
import xml.etree.ElementTree as ET
import zipfile

RAIZ = Path(__file__).resolve().parents[2]
XLSX = RAIZ / 'saidas' / 'oficial' / 'relatorio_operacional_v225.xlsx'
OUT = RAIZ / 'saidas' / 'diagnostico' / 'v17_d0'
OUT.mkdir(parents=True, exist_ok=True)
COLS_USO_POS_SWITCHING = [
    'lote_origem',
    'data_switching',
    'data_pagamento',
    'descricao_pagamento',
    'valor_pagamento',
    'origem_ocorrencia',
    'dias_apos_switching',
    'status_comparacao_temporal',
    'violacao_uso_pos_switching',
    'justificativa',
]


@dataclass
class Switching:
    lote_origem: str
    data_switching: str
    lote_destino: str
    produto_destino_switching: str
    valor_liquido_migrado: float


def _norm(s: str) -> str:
    s = (s or '').strip().lower()
    s = ''.join(ch for ch in unicodedata.normalize('NFD', s) if unicodedata.category(ch) != 'Mn')
    s = re.sub(r'[^a-z0-9]+', ' ', s)
    s = re.sub(r'\s+', ' ', s)
    return s


def _parse_float(v: object) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    t = str(v).strip()
    if not t:
        return 0.0
    t = t.replace('R$', '').replace(' ', '')
    if ',' in t and '.' in t:
        t = t.replace('.', '').replace(',', '.')
    elif ',' in t:
        t = t.replace(',', '.')
    try:
        return float(t)
    except Exception:
        return 0.0


def _xlsx_sheets(path: Path) -> dict[str, list[list[str]]]:
    if not path.exists():
        raise FileNotFoundError(f'planilha operacional obrigatoria ausente: {path}')
    # Preferência V17-D0.2: leitura por openpyxl quando disponível.
    try:
        from openpyxl import load_workbook  # type: ignore
        wb = load_workbook(path, data_only=True)
        out: dict[str, list[list[str]]] = {}
        for ws in wb.worksheets:
            rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                rows.append(['' if cell is None else str(cell) for cell in row])
            out[ws.title] = rows
        if out:
            return out
    except Exception:
        # fallback XML nativo
        pass

    ns = {'m': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main', 'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}
    with zipfile.ZipFile(path) as z:
        shared = []
        if 'xl/sharedStrings.xml' in z.namelist():
            root = ET.fromstring(z.read('xl/sharedStrings.xml'))
            for si in root.findall('m:si', ns):
                txt = ''.join(t.text or '' for t in si.findall('.//m:t', ns))
                shared.append(txt)

        wb = ET.fromstring(z.read('xl/workbook.xml'))
        rels = ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))
        rel_map = {r.attrib['Id']: r.attrib['Target'] for r in rels}
        out: dict[str, list[list[str]]] = {}
        for sh in wb.findall('m:sheets/m:sheet', ns):
            nm = sh.attrib.get('name', '')
            rid = sh.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id', '')
            target = rel_map.get(rid, '')
            part = 'xl/' + target.lstrip('/')
            if part not in z.namelist():
                continue
            ws = ET.fromstring(z.read(part))
            rows = []
            for row in ws.findall('m:sheetData/m:row', ns):
                vals = {}
                maxc = 0
                for c in row.findall('m:c', ns):
                    ref = c.attrib.get('r', 'A1')
                    col = ''.join(ch for ch in ref if ch.isalpha())
                    idx = 0
                    for ch in col:
                        idx = idx * 26 + (ord(ch) - 64)
                    idx -= 1
                    maxc = max(maxc, idx)
                    t = c.attrib.get('t', '')
                    value = ''
                    if t == 'inlineStr':
                        inline = c.find('m:is', ns)
                        if inline is not None:
                            value = ''.join(tn.text or '' for tn in inline.findall('.//m:t', ns))
                    else:
                        v = c.find('m:v', ns)
                        value = v.text if v is not None and v.text is not None else ''
                        if t == 's' and value.isdigit():
                            value = shared[int(value)] if int(value) < len(shared) else ''
                    vals[idx] = value
                line = [''] * (maxc + 1)
                for i, v in vals.items():
                    line[i] = v
                rows.append(line)
            out[nm] = rows
    return out


def _extract_switchings(sheet: list[list[str]]) -> list[Switching]:
    if not sheet:
        return []
    header = [_norm(c) for c in sheet[0]]
    def idx(opts: list[str]) -> int:
        for o in opts:
            if o in header:
                return header.index(o)
        return -1
    i_origem = idx([
        'lote id antes',
        'lote antes',
        'lote origem',
        'lote origem switching',
        'lote_origem',
        'origem',
    ])
    i_data = idx([
        'data sugerida',
        'data switching',
        'data_switching',
        'data',
    ])
    i_dest = idx([
        'lote id depois',
        'lote depois',
        'lote destino',
        'lote_destino',
        'lote pos switching',
        'lote pos-switching',
        'lote destino pos switching',
        'lote destino pos-switching',
        'novo lote',
        'lote novo',
        'lote id novo',
        'lote id pos switching',
        'lote id pos-switching',
        'novo lote destino',
        'lote destino novo',
        'destino switching',
        'destino_switching',
    ])
    i_prod_dest = idx([
        'produto destino switching',
        'produto destino',
        'carteira destino',
        'destino produto',
    ])
    i_val = idx([
        'valor liquido migrado',
        'valor_liquido_migrado',
        'valor liquido origem',
        'valor_liquido_origem',
        'valor liquido',
        'valor_liquido',
    ])
    # schema oficial pode ter apenas produto destino sem lote destino explicito
    if len(sheet) > 1 and i_dest < 0 and i_prod_dest < 0:
        raise RuntimeError('falha_schema_switching: coluna de lote/produto destino nao reconhecida')
    out = []
    for r in sheet[1:]:
        origem = r[i_origem] if i_origem >= 0 and i_origem < len(r) else ''
        if not origem:
            continue
        out.append(Switching(
            lote_origem=str(origem),
            data_switching=str(r[i_data] if i_data >= 0 and i_data < len(r) else ''),
            lote_destino=str(r[i_dest] if i_dest >= 0 and i_dest < len(r) else ''),
            produto_destino_switching=str(r[i_prod_dest] if i_prod_dest >= 0 and i_prod_dest < len(r) else ''),
            valor_liquido_migrado=_parse_float(str(r[i_val] if i_val >= 0 and i_val < len(r) else '0')),
        ))
    return out


def _linhas_situacao_lotes_ativos(situacao: list[list[str]]) -> list[list[str]]:
    blocos_ativos: list[list[str]] = []
    em_ativo = False
    for row in situacao:
        linha = ' '.join(str(c) for c in row if str(c).strip())
        n = _norm(linha)
        if n:
            if 'lotes ativos' in n:
                em_ativo = True
                continue
            if 'lotes exauridos' in n and em_ativo:
                em_ativo = False
            if em_ativo and any(k in n for k in ['patrimonio total dos lotes', 'recebidos auditaveis', 'fechamento economico', 'resumo de recebidos']):
                em_ativo = False
        if em_ativo:
            blocos_ativos.append(row)
    return blocos_ativos


def _tokens_lote(celula: str) -> list[str]:
    raw = str(celula or '')
    partes = re.split(r'\s*(?:\+|;|,|\||/)\s*', raw)
    tokens = []
    for parte in partes:
        token = _norm(parte)
        if token:
            tokens.append(token)
    return tokens


def _cell_matches_lote(celula: str, lote: str) -> bool:
    alvo = _norm(lote)
    if not alvo:
        return False
    cel = _norm(celula)
    if cel == alvo:
        return True
    return alvo in _tokens_lote(celula)


def _linhas_com_lote(rows: list[list[str]], lote: str) -> list[list[str]]:
    linhas: list[list[str]] = []
    for r in rows:
        if any(_cell_matches_lote(str(c), lote) for c in r):
            linhas.append(r)
    return linhas


def _contains_lote(rows: list[list[str]], lote: str) -> bool:
    return len(_linhas_com_lote(rows, lote)) > 0


def _write_csv(path: Path, cols: list[str], data: list[dict]) -> None:
    with path.open('w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for d in data:
            w.writerow(d)


def _sheet_por_alias(
    sheets: dict[str, list[list[str]]],
    aliases: list[str],
    *,
    excluir: list[str] | None = None,
    descricao: str = 'aba',
) -> tuple[str, list[list[str]]]:
    excl = {_norm(x) for x in (excluir or [])}
    alias_set = {_norm(x) for x in aliases}
    for nome, rows in sheets.items():
        n = _norm(nome)
        if n in excl:
            continue
        if n in alias_set:
            return nome, rows
    raise RuntimeError(f'falha_schema_switching: {descricao} nao encontrada; abas_disponiveis={list(sheets.keys())}')


def _linhas_evidencia_lote_sintetico(sheets: dict[str, list[list[str]]]) -> list[list[str]]:
    aliases = {
        _norm('Lotes Sinteticos Pos-Sw'),
        _norm('Lotes Sintéticos Pos-Sw'),
        _norm('Estado Pos-Switching'),
        _norm('Estado Pós-Switching'),
    }
    linhas: list[list[str]] = []
    for nome, rows in sheets.items():
        if _norm(nome) in aliases:
            linhas.extend(rows)
    return linhas


def _idx_coluna(header: list[str], aliases: list[str]) -> int:
    h = [_norm(c) for c in header]
    for a in aliases:
        n = _norm(a)
        if n in h:
            return h.index(n)
    return -1


def _parse_date(v: object) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        # serial Excel (base 1899-12-30)
        try:
            base = datetime(1899, 12, 30).date()
            return base + timedelta(days=int(float(v)))
        except Exception:
            return None
    t = str(v).strip()
    if not t:
        return None
    t = t.replace('T', ' ')
    if re.fullmatch(r'\d+(\.\d+)?', t):
        try:
            base = datetime(1899, 12, 30).date()
            return base + timedelta(days=int(float(t)))
        except Exception:
            return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d/%m/%Y %H:%M:%S', '%Y-%m-%d %H:%M', '%d/%m/%Y %H:%M'):
        try:
            return datetime.strptime(t, fmt).date()
        except Exception:
            pass
    m = re.search(r'(\d{4}-\d{2}-\d{2}|\d{2}/\d{2}/\d{4})', t)
    if m:
        return _parse_date(m.group(1))
    return None


def _usos_lote_pos_switching(extrato_futuro: list[list[str]], lote_origem: str, data_switching: str) -> list[dict]:
    if not extrato_futuro:
        return []
    header = [str(c) for c in extrato_futuro[0]]
    i_data = _idx_coluna(header, ['Data', 'data', 'data_pagamento', 'data pagamento'])
    i_desc = _idx_coluna(header, ['Descrição', 'Descricao', 'Conta', 'conta', 'descricao_pagamento'])
    i_valor = _idx_coluna(header, ['Valor', 'valor', 'valor_pagamento', 'valor pagamento'])
    i_lote = _idx_coluna(header, ['Lote', 'Lotes usados', 'Lote usado', 'Fonte', 'fonte', 'lote_origem', 'lote origem'])
    dt_sw = _parse_date(data_switching)
    usos = []
    for row in extrato_futuro[1:]:
        has_lote = False
        if i_lote >= 0 and i_lote < len(row):
            has_lote = _cell_matches_lote(str(row[i_lote]), lote_origem)
        else:
            has_lote = any(_cell_matches_lote(str(c), lote_origem) for c in row)
        if not has_lote:
            continue
        data_raw = row[i_data] if i_data >= 0 and i_data < len(row) else ''
        dt_pg = _parse_date(data_raw)
        desc = str(row[i_desc]) if i_desc >= 0 and i_desc < len(row) else ''
        valor = _parse_float(row[i_valor]) if i_valor >= 0 and i_valor < len(row) else 0.0
        viol = False
        dias = ''
        status = 'indeterminado_datas_nao_interpretaveis'
        just = 'datas de pagamento e switching nao interpretaveis; ocorrencia mantida para auditoria'
        if dt_pg is not None and dt_sw is not None:
            d = (dt_pg - dt_sw).days
            dias = d
            if d > 0:
                viol = True
                status = 'violacao_pos_switching'
                just = 'origem aparece em pagamento posterior ao switching'
            elif d == 0:
                status = 'intradiario_sem_classificacao_automatica'
                just = 'ocorrencia no mesmo dia do switching; caso intradiario sem violacao automatica'
            else:
                status = 'ocorrencia_anterior_ao_switching'
                just = 'ocorrencia anterior ao switching; nao e violacao pos-switching'
        elif dt_pg is not None and dt_sw is None:
            status = 'indeterminado_data_switching_nao_interpretavel'
            just = 'data_switching nao interpretavel; ocorrencia nao classificada temporalmente'
        elif dt_pg is None and dt_sw is not None:
            status = 'indeterminado_data_pagamento_nao_interpretavel'
            just = 'data_pagamento nao interpretavel; ocorrencia nao classificada temporalmente'
        usos.append({
            'data_pagamento': str(dt_pg if dt_pg is not None else data_raw),
            'descricao_pagamento': desc,
            'valor_pagamento': valor,
            'origem_ocorrencia': 'Extrato Futuro',
            'dias_apos_switching': dias,
            'status_comparacao_temporal': status,
            'violacao_uso_pos_switching': viol,
            'justificativa': just,
        })
    return usos


def _contar_lotes_destino_auditados(switchings: list[Switching]) -> int:
    return len({str(s.lote_destino).strip() for s in switchings if str(s.lote_destino).strip()})


def main() -> int:
    sheets = _xlsx_sheets(XLSX)
    _, switching_sheet = _sheet_por_alias(
        sheets,
        ['Switching', 'Switchings', 'Eventos Switching', 'Eventos de Switching', 'Auditoria Switching', 'Tabela Switching'],
        excluir=['Resumo Switching'],
        descricao='aba de eventos switching',
    )
    situacao = sheets.get('Situação Atual', []) or sheets.get('Situacao Atual', [])
    inventario = sheets.get('Inventário de Lotes', []) or sheets.get('Inventario de Lotes', [])
    extrato_futuro = sheets.get('Extrato Futuro', [])
    linhas_lotes_sinteticos = _linhas_evidencia_lote_sintetico(sheets)

    switchings = _extract_switchings(switching_sheet)
    if switching_sheet and len(switching_sheet) > 1 and not switchings:
        raise RuntimeError(
            'falha_extracao_switching: aba Switching possui linhas, mas nenhum switching foi extraido'
        )
    origens = []
    usos = []
    destinos = []
    matriz = []

    for s in switchings:
        linhas_ativos = _linhas_situacao_lotes_ativos(situacao)
        ativo = _contains_lote(linhas_ativos, s.lote_origem)
        linhas_lote = _linhas_com_lote(inventario, s.lote_origem)
        exaurido = any('exaur' in _norm(' '.join(str(c) for c in row)) for row in linhas_lote)
        viol_ativo = bool(ativo)
        origens.append({
            'lote_origem': s.lote_origem, 'data_switching': s.data_switching, 'destino_switching': s.lote_destino,
            'valor_liquido_migrado': s.valor_liquido_migrado, 'aparece_em_situacao_atual_ativo': ativo,
            'aparece_em_lotes_exauridos': exaurido, 'status_observado': 'ativo_pos_switching' if ativo else 'nao_ativo_pos_switching',
            'status_esperado_contrato': 'origem_migrada_nao_ativa', 'violacao_lote_origem_ativo': viol_ativo,
            'justificativa': 'lote de origem ainda visível após switching' if viol_ativo else 'sem indício de atividade indevida',
        })

        usos_lote = _usos_lote_pos_switching(extrato_futuro, s.lote_origem, s.data_switching)
        if not usos_lote:
            usos.append({
                'lote_origem': s.lote_origem, 'data_switching': s.data_switching, 'data_pagamento': '', 'descricao_pagamento': '',
                'valor_pagamento': 0.0, 'origem_ocorrencia': 'Extrato Futuro', 'dias_apos_switching': '',
                'status_comparacao_temporal': 'sem_uso_detectado',
                'violacao_uso_pos_switching': False, 'justificativa': 'sem uso detectado no extrato futuro para o lote origem',
            })
        else:
            for u in usos_lote:
                usos.append({'lote_origem': s.lote_origem, 'data_switching': s.data_switching, **u})

        ap_inv = _contains_lote(inventario, s.lote_destino) if s.lote_destino else False
        ap_sit = _contains_lote(situacao, s.lote_destino) if s.lote_destino else False
        ap_sint = _contains_lote(linhas_lotes_sinteticos, s.lote_destino) if s.lote_destino else False
        if s.lote_destino:
            mat = 'sim' if (ap_inv or ap_sit) else 'nao'
            viol_dest = not (ap_inv or ap_sit)
            just_dest = 'destino sem materialização auditável' if viol_dest else 'destino identificado em estado temporal'
        elif s.produto_destino_switching:
            mat = 'indeterminada_por_schema'
            viol_dest = False
            just_dest = 'schema possui produto destino, mas nao lote destino explicito; materializacao de lote nao auditavel nesta aba'
        else:
            mat = 'nao'
            viol_dest = False
            just_dest = 'sem lote/produto destino explicito na linha de switching'
        destinos.append({
            'lote_origem': s.lote_origem, 'lote_destino': s.lote_destino, 'produto_destino': s.produto_destino_switching, 'data_recebimento': '', 'data_aplicacao': '',
            'valor_liquido_migrado': s.valor_liquido_migrado, 'aparece_no_inventario': ap_inv, 'aparece_na_situacao_atual': ap_sit,
            'aparece_como_lote_sintetico': ap_sint, 'materializacao_observada': mat, 'materializacao_esperada': 'sim',
            'violacao_destino_nao_materializado': viol_dest,
            'justificativa': just_dest,
        })

    origens_ativas = sum(1 for x in origens if x['violacao_lote_origem_ativo'])
    usos_pos = sum(1 for x in usos if x['violacao_uso_pos_switching'])
    usos_indet = sum(1 for x in usos if str(x.get('status_comparacao_temporal', '')).startswith('indeterminado'))
    dest_nao = sum(1 for x in destinos if x['violacao_destino_nao_materializado'])
    dupla = any(o['violacao_lote_origem_ativo'] and (not d['violacao_destino_nao_materializado']) for o, d in zip(origens, destinos))

    if usos_pos:
        matriz.append({'regra_contrato_modelo': 'origem_migrada_nao_pode_pagar', 'evidencia_observada': f'{usos_pos} origens usadas pós-switching', 'status_aderencia': 'violado', 'severidade': 'alta', 'impacto_potencial': 'pagamento indevido', 'tipo_correcao_futura': 'v17_d1_transicao_temporal', 'observacao': 'uso pós-switching detectado'})
    if origens_ativas:
        matriz.append({'regra_contrato_modelo': 'origem_migrada_nao_pode_estar_ativa', 'evidencia_observada': f'{origens_ativas} origens ativas', 'status_aderencia': 'violado', 'severidade': 'alta', 'impacto_potencial': 'dupla contagem patrimonial', 'tipo_correcao_futura': 'v17_d1_transicao_temporal', 'observacao': 'origem ainda ativa'})
    if dest_nao:
        matriz.append({'regra_contrato_modelo': 'destino_pos_switching_deve_materializar', 'evidencia_observada': f'{dest_nao} destinos não materializados', 'status_aderencia': 'violado', 'severidade': 'alta', 'impacto_potencial': 'quebra de rastreabilidade', 'tipo_correcao_futura': 'v17_d1_transicao_temporal', 'observacao': 'destino ausente'})
    if not matriz:
        matriz.append({'regra_contrato_modelo': 'aderencia_geral', 'evidencia_observada': 'sem violacoes detectadas', 'status_aderencia': 'aderente', 'severidade': 'baixa', 'impacto_potencial': 'nenhum', 'tipo_correcao_futura': 'sem_correcao', 'observacao': ''})

    lotes_destino_auditados = _contar_lotes_destino_auditados(switchings)
    resumo = [
        {'metrica': 'status_global_v17_d0', 'valor': 'ok_diagnostico'},
        {'metrica': 'switchings_auditados', 'valor': len(switchings)},
        {'metrica': 'lotes_origem_auditados', 'valor': len({s.lote_origem for s in switchings})},
        {'metrica': 'lotes_destino_auditados', 'valor': lotes_destino_auditados},
        {'metrica': 'origens_ativas_pos_switching', 'valor': origens_ativas},
        {'metrica': 'origens_usadas_pagamento_pos_switching', 'valor': usos_pos},
        {'metrica': 'usos_origem_pos_switching_indeterminados', 'valor': usos_indet},
        {'metrica': 'destinos_nao_materializados', 'valor': dest_nao},
        {'metrica': 'possivel_dupla_contagem', 'valor': dupla},
        {'metrica': 'violacoes_contrato_modelo_total', 'valor': sum(1 for m in matriz if m['status_aderencia'] == 'violado')},
        {'metrica': 'decisao_correcao_funcional', 'valor': 'abrir_v17_d1_se_confirmadas_violacoes_temporais' if any(m['status_aderencia']=='violado' for m in matriz) else 'sem_correcao_funcional_necessaria'},
        {'metrica': 'confirmacao_sem_alterar_motor', 'valor': True},
        {'metrica': 'confirmacao_sem_alterar_contrato_modelo', 'valor': True},
        {'metrica': 'confirmacao_sem_alterar_ranking', 'valor': True},
        {'metrica': 'confirmacao_sem_alterar_pagamentos', 'valor': True},
        {'metrica': 'confirmacao_sem_alterar_switching', 'valor': True},
        {'metrica': 'confirmacao_sem_alterar_saida_operacional', 'valor': True},
    ]

    _write_csv(OUT / 'v17_d0_origens_switching_estado_atual.csv', list(origens[0].keys()) if origens else ['lote_origem','data_switching','destino_switching','valor_liquido_migrado','aparece_em_situacao_atual_ativo','aparece_em_lotes_exauridos','status_observado','status_esperado_contrato','violacao_lote_origem_ativo','justificativa'], origens)
    _write_csv(OUT / 'v17_d0_uso_pos_switching_pagamentos.csv', COLS_USO_POS_SWITCHING, usos)
    _write_csv(OUT / 'v17_d0_destinos_switching_materializacao.csv', list(destinos[0].keys()) if destinos else ['lote_origem','lote_destino','produto_destino','data_recebimento','data_aplicacao','valor_liquido_migrado','aparece_no_inventario','aparece_na_situacao_atual','aparece_como_lote_sintetico','materializacao_observada','materializacao_esperada','violacao_destino_nao_materializado','justificativa'], destinos)
    _write_csv(OUT / 'v17_d0_matriz_aderencia_contrato_modelo.csv', list(matriz[0].keys()), matriz)
    _write_csv(OUT / 'v17_d0_resumo.csv', ['metrica', 'valor'], resumo)

    print('=== V17-D0 — AUDITORIA TRANSICAO TEMPORAL POS-SWITCHING ===')
    for r in resumo:
        print(f"{r['metrica']}={r['valor']}")
    print(f'output_dir={OUT}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
