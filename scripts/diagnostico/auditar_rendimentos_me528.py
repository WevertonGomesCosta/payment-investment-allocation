from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


RAIZ = Path(__file__).resolve().parents[2]
XLSX = RAIZ / "saidas" / "oficial" / "relatorio_operacional_pr532.xlsx"
OUT_DIR = RAIZ / "saidas" / "diagnostico"
OUT_CSV = OUT_DIR / "auditoria_rendimentos_me528.csv"
OUT_MD = OUT_DIR / "auditoria_rendimentos_me528.md"

BLOCOS = {
    "Lotes exauridos — identificação": ("exauridos", "id"),
    "Lotes exauridos — valores e patrimônio": ("exauridos", "valores"),
    "Lotes ativos — identificação": ("ativos", "id"),
    "Lotes ativos — valores e patrimônio": ("ativos", "valores"),
    "Patrimônio total dos lotes": ("patrimonio_total", "metricas"),
}

NUMERICOS = {
    "Orig.", "Bruto sac.", "Líq. sac.", "Bruto atual", "Líq. atual",
    "Patr. líq.", "Rend. líq.", "Rend. líq. motor", "Dif. rend.",
    "Valor",
}

SENTINELAS = {
    "Lote 10342 fev.",
    "Lote 6630,64 fev.",
    "Lote 4124,75 fev.",
    "Lote 7600 jun.",
    "Lote 4876 jun",
    "Lote 5680 abr.",
    "Lote 3120 mai",
    "Lote 2800 mai.",
}


def norm(v: Any) -> str:
    return str(v or "").strip()


def fnum(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return round(float(v), 2)
    except Exception:
        return 0.0


def ler_blocos(ws) -> dict[str, list[dict[str, Any]]]:
    rows = list(ws.iter_rows(values_only=True))
    saida: dict[str, list[dict[str, Any]]] = {k: [] for k in BLOCOS}

    i = 0
    while i < len(rows):
        titulo = norm(rows[i][0] if rows[i] else "")
        if titulo not in BLOCOS:
            i += 1
            continue

        header = [norm(x) for x in rows[i + 1]]
        j = i + 2

        while j < len(rows):
            primeira = norm(rows[j][0] if rows[j] else "")
            if not primeira:
                break
            if primeira in BLOCOS:
                break

            item = {}
            for col, val in zip(header, rows[j]):
                if not col:
                    continue
                item[col] = fnum(val) if col in NUMERICOS else val
            saida[titulo].append(item)
            j += 1

        i = j + 1

    return saida


def indexar_por_lote(linhas: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    return {norm(row.get("Lote")): row for row in linhas if norm(row.get("Lote"))}


def classe_linha(row: dict[str, Any]) -> tuple[str, str, str]:
    status = norm(row.get("Status ciclo")).lower()
    dif = fnum(row.get("Dif. rend."))
    bruto_sacado = fnum(row.get("Bruto sac."))
    liquido_sacado = fnum(row.get("Líq. sac."))
    liquido_atual = fnum(row.get("Líq. atual"))

    if abs(dif) <= 0.01:
        return "sem_divergencia_material", "não", "bases coincidem dentro de tolerância"

    if status == "migrado_por_switching":
        return (
            "divergencia_origem_migrada_por_switching",
            "não",
            "linha é origem de switching; deve ser auditável, mas não necessariamente somável no patrimônio principal",
        )

    if status == "ativo_pos_switching":
        return (
            "divergencia_destino_pos_switching",
            "sim",
            "linha é destino pós-switching ativo; auditar base fiscal e valor original sintético",
        )

    if bruto_sacado > 0 or liquido_sacado > 0:
        if liquido_atual > 0:
            return (
                "divergencia_saque_parcial",
                "sim",
                "base observável usa líquido sacado + líquido atual; motor calcula valor líquido do lote em data-alvo",
            )
        return (
            "divergencia_exaurido_por_saque",
            "sim",
            "base observável usa líquido efetivamente sacado; motor calcula valor financeiro teórico na data-alvo",
        )

    return (
        "divergencia_ativo_sem_saque",
        "sim",
        "auditar líquido atual renderizado contra valor_liquido_em_data; possível janela CDI, IR/IOF ou base fiscal",
    )


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(XLSX, data_only=True)
    ws = wb["Situação Atual"]
    blocos = ler_blocos(ws)

    ex_id = indexar_por_lote(blocos["Lotes exauridos — identificação"])
    ex_val = indexar_por_lote(blocos["Lotes exauridos — valores e patrimônio"])
    at_id = indexar_por_lote(blocos["Lotes ativos — identificação"])
    at_val = indexar_por_lote(blocos["Lotes ativos — valores e patrimônio"])

    auditadas = []

    for lote, val in ex_val.items():
        ident = ex_id.get(lote, {})
        row = {**ident, **val}
        row["grupo"] = "exauridos"
        auditadas.append(row)

    for lote, val in at_val.items():
        ident = at_id.get(lote, {})
        row = {**ident, **val}
        row["grupo"] = "ativos"
        auditadas.append(row)

    linhas_csv = []
    for row in auditadas:
        classe, entra_base_somavel_estimativa, acao = classe_linha(row)
        original = fnum(row.get("Orig."))
        patrimonio = fnum(row.get("Patr. líq."))
        rend_saida = fnum(row.get("Rend. líq."))
        rend_motor = fnum(row.get("Rend. líq. motor"))
        dif = fnum(row.get("Dif. rend."))

        linhas_csv.append({
            "lote": norm(row.get("Lote")),
            "grupo": norm(row.get("grupo")),
            "status_ciclo": norm(row.get("Status ciclo")),
            "produto": norm(row.get("Carteira")),
            "data_aplicacao": norm(row.get("Aplic.")),
            "base_fiscal": norm(row.get("Base fiscal")),
            "data_termino": norm(row.get("Data término")),
            "dias_corridos": norm(row.get("Dias corr.")),
            "dias_uteis": norm(row.get("Dias úteis")),
            "valor_original": original,
            "bruto_sacado": fnum(row.get("Bruto sac.")),
            "liquido_sacado": fnum(row.get("Líq. sac.")),
            "bruto_atual": fnum(row.get("Bruto atual")),
            "liquido_atual": fnum(row.get("Líq. atual")),
            "patrimonio_liquido": patrimonio,
            "rendimento_liquido_saida": rend_saida,
            "rendimento_liquido_motor": rend_motor,
            "dif_rendimento": dif,
            "recalculo_rend_saida": round(patrimonio - original, 2),
            "erro_formula_saida": round(rend_saida - (patrimonio - original), 2),
            "formula_saida_identificada": "Patr. líq. - Orig.",
            "formula_motor_identificada": "valor_liquido_em_data(data_alvo) - valor_inicial",
            "classe_causa_provavel": classe,
            "entra_base_somavel_estimativa": entra_base_somavel_estimativa,
            "acao_recomendada": acao,
            "sentinela": "sim" if norm(row.get("Lote")) in SENTINELAS else "não",
        })

    patrimonio_total = {
        norm(row.get("Métrica")): fnum(row.get("Valor"))
        for row in blocos["Patrimônio total dos lotes"]
        if norm(row.get("Métrica"))
    }

    campos = list(linhas_csv[0].keys()) if linhas_csv else []
    with OUT_CSV.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=campos)
        w.writeheader()
        w.writerows(linhas_csv)

    total_original_linhas = round(sum(r["valor_original"] for r in linhas_csv), 2)
    total_patrimonio_linhas = round(sum(r["patrimonio_liquido"] for r in linhas_csv), 2)
    total_rend_saida_linhas = round(sum(r["rendimento_liquido_saida"] for r in linhas_csv), 2)
    total_rend_motor_linhas = round(sum(r["rendimento_liquido_motor"] for r in linhas_csv), 2)
    total_dif_linhas = round(sum(r["dif_rendimento"] for r in linhas_csv), 2)

    agg_valor_original = patrimonio_total.get("Valor original total", 0.0)
    agg_patrimonio = patrimonio_total.get("Patrimônio líquido atual", 0.0)
    agg_rendimento = patrimonio_total.get("Rendimento líquido atual", 0.0)

    divergentes = [r for r in linhas_csv if abs(r["dif_rendimento"]) > 0.01]
    sentinelas = [r for r in linhas_csv if r["sentinela"] == "sim"]

    classes = {}
    for r in linhas_csv:
        classes[r["classe_causa_provavel"]] = classes.get(r["classe_causa_provavel"], 0) + 1

    md = []
    md.append("# ME-528 — Auditoria de rendimentos e reconciliação econômico-fiscal\n\n")
    md.append("## Fórmulas identificadas\n\n")
    md.append("- `Rend. líq.` = `Patr. líq.` - `Orig.`\n")
    md.append("- `Rend. líq. motor` = `lote.valor_liquido_em_data(data_alvo) - lote.valor_inicial`\n")
    md.append("- `Dif. rend.` = `Rend. líq.` - `Rend. líq. motor`\n\n")

    md.append("## Totais das linhas exibidas\n\n")
    md.append(f"- lotes auditados: {len(linhas_csv)}\n")
    md.append(f"- lotes com divergência material: {len(divergentes)}\n")
    md.append(f"- soma valor original das linhas: {total_original_linhas:.2f}\n")
    md.append(f"- soma patrimônio líquido das linhas: {total_patrimonio_linhas:.2f}\n")
    md.append(f"- soma Rend. líq. das linhas: {total_rend_saida_linhas:.2f}\n")
    md.append(f"- soma Rend. líq. motor das linhas: {total_rend_motor_linhas:.2f}\n")
    md.append(f"- soma Dif. rend. das linhas: {total_dif_linhas:.2f}\n\n")

    md.append("## Agregado oficial do bloco Patrimônio total dos lotes\n\n")
    md.append(f"- Valor original total: {agg_valor_original:.2f}\n")
    md.append(f"- Patrimônio líquido atual: {agg_patrimonio:.2f}\n")
    md.append(f"- Rendimento líquido atual: {agg_rendimento:.2f}\n")
    md.append(f"- checagem agregado: Patrimônio - Original = {round(agg_patrimonio - agg_valor_original, 2):.2f}\n\n")

    md.append("## Diferença entre linhas exibidas e agregado oficial\n\n")
    md.append(f"- diferença valor original: {round(total_original_linhas - agg_valor_original, 2):.2f}\n")
    md.append(f"- diferença patrimônio líquido: {round(total_patrimonio_linhas - agg_patrimonio, 2):.2f}\n")
    md.append(f"- diferença rendimento líquido: {round(total_rend_saida_linhas - agg_rendimento, 2):.2f}\n\n")

    md.append("## Classes prováveis\n\n")
    for k, v in sorted(classes.items()):
        md.append(f"- {k}: {v}\n")

    md.append("\n## Casos sentinela\n\n")
    for r in sentinelas:
        md.append(
            f"- {r['lote']} | {r['status_ciclo']} | "
            f"Orig={r['valor_original']:.2f} | Patr={r['patrimonio_liquido']:.2f} | "
            f"Rend. saída={r['rendimento_liquido_saida']:.2f} | "
            f"Rend. motor={r['rendimento_liquido_motor']:.2f} | "
            f"Dif={r['dif_rendimento']:.2f} | {r['classe_causa_provavel']}\n"
        )

    md.append("\n## Condição de parada provisória\n\n")
    md.append(
        "A auditoria deve agora separar formalmente duas questões: "
        "(1) divergência entre base patrimonial observável e base financeira do motor; "
        "(2) diferença entre a soma das linhas exibidas e o agregado oficial do patrimônio total. "
        "Não há evidência suficiente para corrigir motor ou saída antes dessa separação.\n"
    )

    OUT_MD.write_text("".join(md), encoding="utf-8")

    print(f"[OK] CSV: {OUT_CSV}")
    print(f"[OK] MD:  {OUT_MD}")
    print(f"[OK] lotes auditados: {len(linhas_csv)}")
    print(f"[OK] divergências materiais: {len(divergentes)}")
    print(f"[OK] soma Dif. rend. linhas: {total_dif_linhas:.2f}")
    print(f"[OK] diferença patrimônio linhas - agregado: {round(total_patrimonio_linhas - agg_patrimonio, 2):.2f}")


if __name__ == "__main__":
    main()
