from __future__ import annotations

from pathlib import Path
from typing import Iterable, Any
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

RAIZ = Path(__file__).resolve().parents[2]
if str(RAIZ) not in sys.path:
    sys.path.insert(0, str(RAIZ))

from nucleo.contexto_baseline import carregar_contexto_baseline
from nucleo.identidade_baseline import caminho_artifact, caminho_saida_operacional, nome_relatorio_operacional, VERSAO_BASELINE
from nucleo.saida_canonica import construir_saida_canonica


SAIDA_INTERNA = caminho_saida_operacional(RAIZ, nome_relatorio_operacional())
SAIDA_EXTERNA = caminho_artifact(nome_relatorio_operacional())


def _valor(item: dict[str, Any], chave: str) -> Any:
    return item.get(chave)


def _rows(itens: Iterable[dict[str, Any]], headers: list[str]) -> list[list[Any]]:
    return [[_valor(item, header) for header in headers] for item in itens]


def _apply_table_style(ws, headers: list[str], rows: list[list[Any]], *, start_row: int = 1, title: str | None = None, freeze: bool = False) -> int:
    ws.sheet_view.showGridLines = False
    header_fill = PatternFill('solid', fgColor='D9EAF7')
    header_font = Font(color='1F1F1F', bold=True)
    title_fill = PatternFill('solid', fgColor='EDF4FA')
    title_font = Font(color='1F1F1F', bold=True, size=12)
    thin_gray = Side(style='thin', color='D9E1F2')

    header_row = start_row
    if title:
        title_row = start_row
        ws.cell(row=title_row, column=1, value=title).fill = title_fill
        ws.cell(row=title_row, column=1).font = title_font
        ws.cell(row=title_row, column=1).alignment = Alignment(horizontal='left')
        header_row = start_row + 1

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(bottom=thin_gray)

    for row_offset, row in enumerate(rows, start=1):
        row_idx = header_row + row_offset
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = Border(bottom=thin_gray)

    if freeze:
        ws.freeze_panes = f'A{header_row + 1}'

    if headers:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{max(header_row + len(rows), header_row)}"

    currency_cols = {
        'Valor', 'Saldo Antes', 'Bruto', 'Imposto', 'Líquido', 'Saldo Remanescente',
        'Ganho estimado', 'Valor líquido origem', 'Score', 'Proxy terminal', 'Ticket mín.',
        'Valor original', 'Saldo rem', 'Valor bruto', 'Valor líquido', 'Valor vinculado',
        'Residual aplicação'
    }
    int_cols = {'Dias corridos', 'Dias úteis', 'Rank', 'Liquidez', 'Carência', 'Pagamentos vinculados'}

    for col_idx, header in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        max_len = len(str(header))
        for row_idx in range(header_row + 1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
            if header in currency_cols and isinstance(value, (int, float)):
                cell.number_format = 'R$ #,##0.00;[Red](R$ #,##0.00);-'
            elif header in int_cols and isinstance(value, (int, float)):
                cell.number_format = '0'
            elif 'Data' in header and hasattr(value, 'year'):
                cell.number_format = 'dd/mm/yyyy'
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 38)
    return header_row + len(rows)


def _adicionar_abas_ranking(wb, contexto) -> None:
    ranking = getattr(contexto, 'ranking_carteira', None)
    if ranking is None:
        return

    ws_carteira = wb.create_sheet('Carteira')
    quadro_carteira = ranking.quadro_destinos_switch.copy()
    cols_carteira = [
        'rank_destino', 'nome', 'score_final', 'proxy_terminal_destino', 'retorno_anual_proxy',
        'liquidez_dias', 'carencia_dias', 'aplicacao_minima', 'aplicacao_maxima',
        'tipo_produto', 'somente_combo', 'Status_Confirmação', 'Campos_Pendentes'
    ]
    cols_carteira = [c for c in cols_carteira if c in quadro_carteira.columns]
    quadro_carteira = quadro_carteira[cols_carteira].copy()
    headers_carteira = [
        'Rank', 'Produto', 'Score Final', 'Proxy Terminal', 'Retorno Proxy aa', 'Liquidez Dias',
        'Carência Dias', 'Aplicação Mínima', 'Aplicação Máxima', 'Tipo Produto',
        'Somente Combo', 'Status Confirmação', 'Campos Pendentes'
    ][:len(cols_carteira)]
    rows_carteira = quadro_carteira.astype(object).where(quadro_carteira.notna(), '').values.tolist()
    _apply_table_style(ws_carteira, headers_carteira, rows_carteira, freeze=True)

    ws_top30 = wb.create_sheet('Top30')
    top30 = ranking.top30.copy()
    rows_top30 = top30.astype(object).where(top30.notna(), '').values.tolist()
    _apply_table_style(ws_top30, list(top30.columns), rows_top30, freeze=True)

    ws_resumo = wb.create_sheet('Resumo Switching')
    resumo_rows = [
        ['produtos_total', ranking.resumo.get('produtos_total')],
        ['produtos_ativos_ranqueados', ranking.resumo.get('produtos_ativos_ranqueados')],
        ['qtd_destinos_switch', ranking.auditoria.get('qtd_destinos_switch')],
        ['destino_top1', ranking.auditoria.get('destino_top1')],
        ['qtd_diffs_materiais_nucleo', ranking.validacao.get('qtd_diffs_materiais_nucleo')],
        ['aceite_nucleo', ranking.validacao.get('aceite_nucleo')],
    ]
    _apply_table_style(ws_resumo, ['indicador', 'valor'], resumo_rows)

    ws_val = wb.create_sheet('Validacao')
    _apply_table_style(
        ws_val,
        ['colunas', 'qtd_diffs_materiais_nucleo', 'aceite_nucleo', 'metodo'],
        [[str(ranking.validacao.get('colunas')), ranking.validacao.get('qtd_diffs_materiais_nucleo'), ranking.validacao.get('aceite_nucleo'), ranking.auditoria.get('metodo')]],
    )


def _adicionar_situacao_atual(wb, saida) -> None:
    ws = wb.create_sheet('Situação Atual')
    r = 1
    r = _apply_table_style(
        ws,
        ['Lote', 'Recebimento', 'Aplicação', 'Último uso', 'Produto', 'Dias corridos', 'Dias úteis', 'Valor original', 'Bruto', 'Líquido', 'Saldo rem'],
        _rows(saida.lotes_exauridos, ['Lote', 'Recebimento', 'Aplicação', 'Último uso', 'Produto', 'Dias corridos', 'Dias úteis', 'Valor original', 'Bruto', 'Líquido', 'Saldo rem']),
        start_row=r,
        title='Lotes exauridos',
    )
    r = _apply_table_style(
        ws,
        ['Lote', 'Recebimento', 'Aplicação', 'Produto', 'Dias corridos', 'Dias úteis', 'Valor original', 'Bruto', 'Líquido', 'Saldo rem'],
        _rows(saida.lotes_ativos, ['Lote', 'Recebimento', 'Aplicação', 'Produto', 'Dias corridos', 'Dias úteis', 'Valor original', 'Bruto', 'Líquido', 'Saldo rem']),
        start_row=r + 3,
        title='Lotes ativos',
    )
    r = _apply_table_style(
        ws,
        ['Recebido', 'Lote origem', 'Recebimento', 'Aplicação', 'Valor bruto', 'Valor líquido', 'Status', 'Destino', 'Pagamentos vinculados', 'Valor vinculado', 'Residual aplicação', 'Disponível ref', 'Observação'],
        _rows(saida.recebidos_atuais, ['Recebido', 'Lote origem', 'Recebimento', 'Aplicação', 'Valor bruto', 'Valor líquido', 'Status', 'Destino', 'Pagamentos vinculados', 'Valor vinculado', 'Residual aplicação', 'Disponível ref', 'Observação']),
        start_row=r + 3,
        title='Recebidos auditáveis',
    )
    r = _apply_table_style(
        ws,
        ['Métrica', 'Valor'],
        _rows(saida.fechamento_atual, ['Métrica', 'Valor']),
        start_row=r + 3,
        title='Fechamento econômico',
    )
    _apply_table_style(
        ws,
        ['Métrica', 'Valor'],
        _rows(saida.resumo_recebidos, ['Métrica', 'Valor']),
        start_row=r + 3,
        title='Resumo de recebidos',
    )


def _adicionar_auditoria_saida_canonica(wb, saida) -> None:
    ws = wb.create_sheet('Saida Canonica')
    linhas = [{'Métrica': k, 'Valor': v} for k, v in saida.auditoria.items()]
    _apply_table_style(ws, ['Métrica', 'Valor'], _rows(linhas, ['Métrica', 'Valor']), freeze=True)


def main() -> Path:
    contexto = carregar_contexto_baseline(
        raiz_repositorio=RAIZ,
        instalar_automaticamente=False,
        incluir_resolver_hibrido_5p_shadow=False,
        incluir_benchmark_agrupado_individual_shadow=False,
        incluir_benchmark_runner_futuro_shadow=False,
        incluir_auditoria_primeira_quebra_runner_futuro_shadow=False,
    )
    saida = construir_saida_canonica(contexto, versao=VERSAO_BASELINE)

    wb = Workbook()
    ws_passado = wb.active
    ws_passado.title = 'Extrato Passado'
    headers_passado = ['Data', 'Conta', 'Despesa ID', 'Lote', 'Saldo Antes', 'Bruto', 'Imposto', 'Líquido', 'Saldo Remanescente']
    _apply_table_style(ws_passado, headers_passado, _rows(saida.extrato_passado, headers_passado), freeze=True)

    ws_futuro = wb.create_sheet('Extrato Futuro')
    headers_futuro = ['Data', 'Conta', 'Despesa ID', 'Valor', 'Lote sugerido', 'Saldo Antes', 'Bruto', 'Imposto', 'Líquido', 'Saldo Remanescente', 'Cobertura integral', 'Estratégia', 'Lote reserva', 'Necessita switching']
    _apply_table_style(ws_futuro, headers_futuro, _rows(saida.extrato_futuro, headers_futuro), freeze=True)

    ws_switching = wb.create_sheet('Switching')
    headers_switching = ['Data sugerida', 'Lote origem', 'Produto origem', 'Produto destino switching', 'Ganho estimado', 'Valor líquido origem', 'Status']
    _apply_table_style(ws_switching, headers_switching, _rows(saida.switchings, headers_switching), freeze=True)

    _adicionar_abas_ranking(wb, contexto)
    _adicionar_situacao_atual(wb, saida)
    _adicionar_auditoria_saida_canonica(wb, saida)

    SAIDA_INTERNA.parent.mkdir(parents=True, exist_ok=True)
    wb.save(SAIDA_INTERNA)
    try:
        if SAIDA_EXTERNA.parent.exists():
            wb.save(SAIDA_EXTERNA)
    except Exception as exc:
        print(f"[AVISO] cópia externa não gerada: {type(exc).__name__}:{exc}")
    return SAIDA_INTERNA


if __name__ == '__main__':
    print(main())
