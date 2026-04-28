from __future__ import annotations

import json
import shutil
from pathlib import Path
from typing import Any, Mapping

from openpyxl import load_workbook

from nucleo.identidade_baseline import (
    VERSAO_BASELINE,
    VERSAO_SLUG,
    caminho_artifact,
    caminho_saida_operacional,
)
from scripts.operacional.gerar_planilha_operacional import main as gerar_planilha_operacional_base


RAIZ = Path(__file__).resolve().parents[2]
CONFIG_CANONICO = RAIZ / 'dados' / 'config_atualizado.json'

ABAS_ATUAIS_PLANILHA_OPERACIONAL = {
    'extrato_passado': 'Extrato Passado',
    'extrato_futuro': 'Extrato Futuro',
    'switching': 'Switching',
    'carteira': 'Carteira',
    'top30': 'Top30',
    'resumo_switching': 'Resumo Switching',
    'validacao': 'Validacao',
    'situacao_atual': 'Situação Atual',
    'saida_canonica': 'Saida Canonica',
}


def _cfg_get(config: Mapping[str, Any], *caminho: str, padrao: Any = None) -> Any:
    atual: Any = config
    for chave in caminho:
        if not isinstance(atual, Mapping) or chave not in atual:
            return padrao
        atual = atual[chave]
    return atual


def _ler_config_canonico() -> dict[str, Any]:
    try:
        dados = json.loads(CONFIG_CANONICO.read_text(encoding='utf-8'))
    except Exception:
        return {}
    return dados if isinstance(dados, dict) else {}


def _config_planilha_operacional(config: Mapping[str, Any]) -> Mapping[str, Any]:
    cfg = _cfg_get(config, 'saidas', 'planilha_operacional', padrao={})
    return cfg if isinstance(cfg, Mapping) else {}


def _nome_aba_configurado(cfg_planilha: Mapping[str, Any], chave: str) -> str:
    abas = cfg_planilha.get('abas') if isinstance(cfg_planilha.get('abas'), Mapping) else {}
    valor = abas.get(chave) if isinstance(abas, Mapping) else None
    return str(valor).strip() if valor not in (None, '') else ABAS_ATUAIS_PLANILHA_OPERACIONAL[chave]


def _nome_arquivo_configurado(cfg_planilha: Mapping[str, Any], caminho_base: Path) -> str:
    valor = cfg_planilha.get('arquivo') or cfg_planilha.get('nome_arquivo')
    if valor in (None, ''):
        return caminho_base.name
    nome = str(valor).strip()
    try:
        nome = nome.format(versao=VERSAO_BASELINE, versao_slug=VERSAO_SLUG)
    except Exception:
        pass
    return nome if nome.lower().endswith('.xlsx') else f'{nome}.xlsx'


def _renomear_abas_operacionais(caminho: Path, cfg_planilha: Mapping[str, Any]) -> bool:
    wb = load_workbook(caminho)
    alterou = False
    for chave, nome_atual in ABAS_ATUAIS_PLANILHA_OPERACIONAL.items():
        nome_novo = _nome_aba_configurado(cfg_planilha, chave)
        if nome_novo == nome_atual:
            continue
        if nome_atual not in wb.sheetnames:
            continue
        if nome_novo in wb.sheetnames:
            continue
        wb[nome_atual].title = nome_novo
        alterou = True
    if alterou:
        wb.save(caminho)
    return alterou


def _copiar_para_artifact(caminho_saida: Path, nome_arquivo: str) -> None:
    destino = caminho_artifact(nome_arquivo)
    try:
        if destino.parent.exists() and destino.resolve() != caminho_saida.resolve():
            shutil.copy2(caminho_saida, destino)
    except Exception as exc:
        print(f"[AVISO] cópia externa configurada não gerada: {type(exc).__name__}:{exc}")


def main() -> Path:
    caminho_base = gerar_planilha_operacional_base()
    config = _ler_config_canonico()
    cfg_planilha = _config_planilha_operacional(config)

    _renomear_abas_operacionais(caminho_base, cfg_planilha)

    nome_arquivo = _nome_arquivo_configurado(cfg_planilha, caminho_base)
    caminho_final = caminho_saida_operacional(RAIZ, nome_arquivo)
    if caminho_final.resolve() != caminho_base.resolve():
        caminho_final.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(caminho_base, caminho_final)

    _copiar_para_artifact(caminho_final, nome_arquivo)
    return caminho_final


if __name__ == '__main__':
    print(main())
