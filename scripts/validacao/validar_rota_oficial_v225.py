from __future__ import annotations

from pathlib import Path
import subprocess
import sys

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


REPO = Path(__file__).resolve().parents[2]

ARQUIVOS_PY_COMPILE = [
    "aplicacao/principal.py",
    "aplicacao/console/principal.py",
    "aplicacao/console/secoes_financeiras.py",
    "aplicacao/console/secoes_canonicas.py",
    "nucleo/saida_observavel.py",
    "nucleo/gerar_planilha_operacional.py",
]

ARQ_CONSOLE = REPO / "aplicacao" / "console" / "principal.py"
ARQ_PLANILHA = REPO / "nucleo" / "gerar_planilha_operacional.py"
ARQ_OBSERVAVEL = REPO / "nucleo" / "saida_observavel.py"
ARQ_RELATORIO_CODEX = REPO / "relatorios" / "atuais" / "codex_ready" / "CODEX_READY_V225.md"
ARQ_XLSX = REPO / "saidas" / "oficial" / "relatorio_operacional_v225.xlsx"


def run(cmd: list[str]) -> None:
    print("$ " + " ".join(cmd))
    proc = subprocess.run(cmd, cwd=REPO, text=True)
    if proc.returncode != 0:
        raise SystemExit(proc.returncode)


def check(cond: bool, msg: str) -> None:
    if not cond:
        raise AssertionError(msg)


def main() -> None:
    for rel in ARQUIVOS_PY_COMPILE:
        check((REPO / rel).exists(), f"Arquivo ausente: {rel}")

    run([sys.executable, "-m", "py_compile", *ARQUIVOS_PY_COMPILE])

    console = ARQ_CONSOLE.read_text(encoding="utf-8")
    planilha = ARQ_PLANILHA.read_text(encoding="utf-8")
    observavel = ARQ_OBSERVAVEL.read_text(encoding="utf-8")

    check("aplicacao.console.secoes_financeiras" not in console, "Console ainda importa secoes_financeiras.")
    check("construir_amostras_pagamentos_operacionais" in console, "Console não usa amostras observáveis.")
    check("construir_linhas_lotes_id_curta" in console, "Console não usa Situação Atual observável.")
    check("construir_blocos_situacao_atual" in planilha, "Planilha não usa blocos observáveis da Situação Atual.")
    check("def construir_amostras_pagamentos_operacionais(" in observavel, "saida_observavel não contém amostras de pagamentos.")
    check("def construir_blocos_situacao_atual(" in observavel, "saida_observavel não contém blocos da Situação Atual.")
    check("Validacao" not in planilha, "Planilha ainda contém referência à aba Validacao.")

    run([sys.executable, "aplicacao/principal.py"])

    check(ARQ_XLSX.exists(), f"Saída operacional não encontrada: {ARQ_XLSX}")

    if load_workbook is not None:
        wb = load_workbook(ARQ_XLSX, read_only=True)
        try:
            sheets = set(wb.sheetnames)
            check("Situação Atual" in sheets, "Aba Situação Atual ausente.")
            check("Validacao" not in sheets, "Aba Validacao ainda existe.")
        finally:
            wb.close()

    if ARQ_RELATORIO_CODEX.exists():
        texto = ARQ_RELATORIO_CODEX.read_text(encoding="utf-8")
        linhas_estado = [
            linha.strip()
            for linha in texto.splitlines()
            if "Estado mínimo Codex-ready" in linha
        ]

        if linhas_estado:
            estado_ok = any("| SIM |" in linha or linha.endswith("| SIM") for linha in linhas_estado)
            if not estado_ok:
                print("[AVISO] CODEX_READY_V225.md ainda não confirma Estado mínimo Codex-ready = SIM.")
                print("[AVISO] Linha encontrada: " + " ; ".join(linhas_estado))
                print("[AVISO] A rota operacional passou; regenere CODEX_READY_V225.md após esta validação.")

    print("")
    print("VALIDAÇÃO OFICIAL V225 CONCLUÍDA COM SUCESSO")


if __name__ == "__main__":
    main()
